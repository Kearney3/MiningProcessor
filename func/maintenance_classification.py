"""维修记录分类引擎

提供故障判定、大类/小类分类、噪声过滤，以及 Excel 配置模板的导入/导出。
分类规则从 config.json 读取；配置为空时使用硬编码默认值。
"""
import logging
import re
import unicodedata
from functools import lru_cache

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

MAINTENANCE_CLASSIFICATION_SCHEMA_VERSION = 2


# ── 默认分类规则 ──────────────────────────────────────────────

_DEFAULT_CLASSIFICATIONS: list[dict] = [
    {"major": "发动机系统", "minor": "发动机总成/大修", "keywords": ["发动机在ub维修", "发动机在维修中", "发动机准备发到", "发动机备件", "发动机安装", "发动机拆解", "诊断发动机", "检修发动机", "待修发动机", "发动机大修", "发动机在大修", "维修发动机", "发动机维修", "发动机总成", "发动机报废", "发动机抱死", "发动机坏", "更换发动机", "等待发动机", "拆卸发动机", "发动机发到", "发动机运到", "装发动机", "大修"]},
    {"major": "发动机系统", "minor": "内部机械", "keywords": ["拉缸", "曲轴", "凸轮轴", "活塞", "缸套", "缸盖", "气门", "连杆", "止推瓦", "缸垫", "大瓦", "小瓦"]},
    {"major": "发动机系统", "minor": "启动机械/飞轮", "keywords": ["飞轮撞击", "撞飞轮", "盘动飞轮", "旋转飞轮", "转飞轮", "撞击飞轮", "飞轮齿圈", "齿圈损坏"], "all_keywords": [["飞轮", "卡"], ["飞轮", "损坏"]]},
    {"major": "发动机系统", "minor": "燃油供给与喷射", "keywords": ["喷油器", "喷油嘴", "高压泵", "共轨", "柴油泵", "燃油泵", "燃油管", "柴油管", "燃油滤芯", "柴油滤芯"]},
    {"major": "发动机系统", "minor": "进气与增压", "keywords": ["增压器", "涡轮增压", "涡轮", "中冷器", "进气管", "空滤堵", "空气滤芯堵", "进气系统"]},
    {"major": "发动机系统", "minor": "冷却系统", "keywords": ["发动机漏水", "水箱堵", "水箱漏", "散热器漏水", "水泵", "节温器", "风扇皮带", "冷却液漏", "防冻液漏", "水温高", "水温报警", "冷却系统"]},
    {"major": "发动机系统", "minor": "润滑与机油系统", "keywords": ["发动机漏油", "发动机渗油", "机油泵", "机油压力", "油底壳", "机油漏", "机油进", "机油报警", "机油管", "缸盖漏油"]},
    {"major": "发动机系统", "minor": "排气与尾气后处理", "keywords": ["排气管", "消声器", "下排气", "排温", "排气温度", "scr", "dpf", "尿素系统", "后处理"]},
    {"major": "发动机系统", "minor": "性能/工况异常", "keywords": ["发动机黄灯", "发动机黄色灯", "发动机出现问题", "发动机没劲", "发动机没劲儿", "发动机无力", "没劲", "发动机异响", "发动机冒烟", "发动机冒黑烟", "发动机冒白烟", "发动机冒蓝烟", "发动机熄火", "发动机高温", "发动机报警", "发动机故障"], "all_keywords": [["发动机", "报警"], ["发动机", "异常"]], "priority": -5},
    {"major": "变速箱与变矩器", "minor": "变速箱总成/内部机械", "keywords": ["变速箱在修复", "准备拆变速箱", "等待变速箱备件", "变速箱已大修", "变速箱已拆", "变速箱大修", "变速箱在大修", "变速箱维修", "待变速箱维修", "维修变速箱", "检修变速箱", "拆卸变速箱", "变速箱坏", "变速箱内部", "变速箱齿轮", "变速箱轴承", "变速箱异响"]},
    {"major": "变速箱与变矩器", "minor": "变矩器", "keywords": ["液力变矩器", "变矩器"]},
    {"major": "变速箱与变矩器", "minor": "换挡/离合器", "keywords": ["挂不上前进挡", "挂不上档", "挂不上挡", "无档", "无挡", "不增档", "不减档", "换挡", "挡位", "档位", "离合器", "前进挡", "倒挡"]},
    {"major": "变速箱与变矩器", "minor": "油路/滤清/冷却", "keywords": ["变速箱漏油", "变速箱渗油", "变速箱滤芯", "变速箱油冷却器", "变速箱油温", "变速箱压力"]},
    {"major": "变速箱与变矩器", "minor": "控制/报警", "keywords": ["变速箱报警", "变速箱故障灯", "变速箱控制器", "变速箱电磁阀", "变速箱报码", "变速箱故障"]},
    {"major": "传动与车桥", "minor": "传动轴/万向节", "keywords": ["传动轴", "十字轴", "万向节", "传动法兰"]},
    {"major": "传动与车桥", "minor": "车桥/差速器/半轴", "keywords": ["前桥", "后桥", "驱动桥", "差速器", "半轴"]},
    {"major": "传动与车桥", "minor": "终传动/轮边减速", "keywords": ["背靠背油封", "轮边减速", "终传动", "轮边", "太阳轮", "行星齿轮"]},
    {"major": "传动与车桥", "minor": "联轴器/通用传动", "keywords": ["维修传动系统", "传动系统", "联轴器", "耦合器", "传动装置"]},
    {"major": "液压系统", "minor": "液压泵/马达", "keywords": ["液压泵", "主泵", "先导泵", "液压马达"]},
    {"major": "液压系统", "minor": "通用液压油缸", "keywords": ["液压油缸", "液压缸", "油缸漏油", "油缸渗油", "油缸密封"], "priority": 30},
    {"major": "液压系统", "minor": "举升/工作装置油缸", "keywords": ["举升油缸", "举升缸", "提升油缸", "提升缸", "转斗油缸", "转斗缸", "大臂油缸", "大臂缸", "铲斗油缸", "铲斗缸", "推土缸"], "priority": 30},
    {"major": "液压系统", "minor": "转向油缸", "keywords": ["转向油缸", "转向缸"], "priority": 30},
    {"major": "液压系统", "minor": "悬挂油缸", "keywords": ["悬挂油缸", "悬挂缸", "悬挂漏油", "悬挂渗油"], "priority": 30},
    {"major": "液压系统", "minor": "控制阀/分配器", "keywords": ["液压控制阀", "液压电磁阀", "多路阀", "分配器", "溢流阀", "阀块"]},
    {"major": "液压系统", "minor": "管路/接头/密封", "keywords": ["液压油胶管", "高压油管", "液压软管", "液压接头", "液压管", "漏液压油", "液压油漏"]},
    {"major": "液压系统", "minor": "油箱/滤清/冷却", "keywords": ["液压油冷却器", "液压散热器", "液压油箱", "液压滤芯", "液压油滤", "液压油污染"]},
    {"major": "液压系统", "minor": "压力/功能异常", "keywords": ["液压油压力", "液压系统不工作", "液压压力", "液压无力", "液压异响", "液压报警", "液压油温"]},
    {"major": "低压电气与控制", "minor": "蓄电池/启动供电", "keywords": ["外接电源启动", "外接启动", "搭接着车", "塔接着车", "电池馈电", "电瓶馈电", "蓄电池", "电瓶", "搭电"], "priority": 25},
    {"major": "低压电气与控制", "minor": "启动机/启动回路", "keywords": ["启动继电器", "启动马达", "启动回路", "启动机", "起动机"], "priority": 25},
    {"major": "低压电气与控制", "minor": "发电机/充电", "keywords": ["发电机不发电", "发电机皮带", "发电机坏", "充电机", "不充电"], "priority": 20},
    {"major": "低压电气与控制", "minor": "线束/配电保护", "keywords": ["接触不良", "保险盒", "保险丝", "继电器", "线束", "电线", "线路", "插头", "短路", "断路"]},
    {"major": "低压电气与控制", "minor": "传感器/开关", "keywords": ["限位开关", "压力开关", "温度开关", "接近开关", "传感器"]},
    {"major": "低压电气与控制", "minor": "控制器/模块/故障码", "keywords": ["电气故障", "电气报警", "控制模块", "故障代码", "故障码", "控制器", "电脑板", "ecm", "ecu", "报码"], "priority": 10},
    {"major": "低压电气与控制", "minor": "灯光/喇叭/信号", "keywords": ["倒车报警器", "转向灯", "刹车灯", "警示灯", "蜂鸣器", "大灯", "灯光", "灯泡", "喇叭"]},
    {"major": "低压电气与控制", "minor": "显示器/仪表/监控", "keywords": ["称重显示", "仪表模块", "显示器", "显示屏", "仪表盘", "监控器"]},
    {"major": "电驱动系统", "minor": "主发电机", "keywords": ["主发电机", "主发", "励磁"], "all_keywords": [["主发", "轴承"], ["主发电机", "轴承"]], "priority": 30},
    {"major": "电驱动系统", "minor": "轮马达/电动轮", "keywords": ["牵引电机", "轮马达", "电动轮"], "all_keywords": [["轮马达", "轴承"], ["电动轮", "轴承"], ["轮马达", "螺栓"], ["电动轮", "螺栓"]], "priority": 30},
    {"major": "电驱动系统", "minor": "逆变/功率模块", "keywords": ["功率模块", "逆变器", "整流器", "igbt"], "priority": 30},
    {"major": "电驱动系统", "minor": "接触器/电阻栅", "keywords": ["高压接触器", "rp接触器", "制动电阻", "电阻栅"], "priority": 30},
    {"major": "电驱动系统", "minor": "高压电缆/驱动控制", "keywords": ["ac驱动故障", "ac驱动", "轮马达电缆", "电驱动报警", "高压电缆", "牵引控制"], "priority": 30},
    {"major": "制动系统", "minor": "行车制动", "keywords": ["行车制动", "制动无力", "刹车无力", "制动器", "脚刹"], "all_keywords": [["制动", "底座"], ["刹车", "底座"], ["制动", "支架"], ["刹车", "支架"], ["制动", "螺栓"], ["刹车", "螺栓"]]},
    {"major": "制动系统", "minor": "驻车制动", "keywords": ["驻车制动", "停车制动", "手刹"]},
    {"major": "制动系统", "minor": "管路/泵阀/压力", "keywords": ["制动蓄能器", "制动压力", "制动管", "刹车管", "制动泵", "制动阀"]},
    {"major": "制动系统", "minor": "摩擦件/盘片", "keywords": ["刹车片", "制动片", "刹车盘", "制动盘", "制动蹄"]},
    {"major": "制动系统", "minor": "报警/控制", "keywords": ["制动控制器", "制动报警", "刹车报警", "abs报警"]},
    {"major": "转向系统", "minor": "转向泵/阀", "keywords": ["转向泵", "转向阀", "优先阀"]},
    {"major": "转向系统", "minor": "连杆/球头/转向节", "keywords": ["转向拉杆", "转向球头", "转向节"]},
    {"major": "转向系统", "minor": "方向机/操纵", "keywords": ["转向操纵", "方向机", "方向盘"]},
    {"major": "转向系统", "minor": "功能/报警", "keywords": ["转向沉重", "不能转向", "转向报警", "转向卡", "无转向"]},
    {"major": "悬挂与车架", "minor": "悬挂蓄能器/压力", "keywords": ["悬挂蓄能器", "悬挂压力"]},
    {"major": "悬挂与车架", "minor": "A型架/平衡梁/关节", "keywords": ["平衡梁", "中心销", "a型架", "a型桥", "铰接"]},
    {"major": "悬挂与车架", "minor": "车架/大梁/底盘", "keywords": ["副车架", "车架", "大梁", "底盘"]},
    {"major": "悬挂与车架", "minor": "减震/弹簧/悬挂通用", "keywords": ["悬挂上座", "悬挂报警", "悬挂系统", "减震器", "弹簧"], "all_keywords": [["悬挂", "轴承"], ["悬挂", "支架"], ["悬挂", "螺栓"]]},
    {"major": "轮胎与车轮", "minor": "轮胎损伤/磨损", "keywords": ["轮胎损坏", "轮胎磨损", "轮胎发热", "轮胎脱空", "轮胎脱", "轮胎破", "轮胎爆", "轮胎裂"]},
    {"major": "轮胎与车轮", "minor": "漏气/胎压", "keywords": ["轮胎漏气", "轮胎补气", "充氮气", "气门嘴", "胎压"]},
    {"major": "轮胎与车轮", "minor": "拆装/换位/更换", "keywords": ["待前轮", "更换轮胎", "轮胎换位", "倒换轮胎", "换轮胎", "拆轮胎", "装轮胎"]},
    {"major": "轮胎与车轮", "minor": "轮辋/轮毂/轴承", "keywords": ["车轮轴承", "轮胎轴承", "轮毂油封", "轮辋", "轮毂"]},
    {"major": "轮胎与车轮", "minor": "防护链/防滑链", "keywords": ["轮胎防护链", "轮胎保护链", "防滑链", "锁齿"]},
    {"major": "履带与行走机构", "minor": "履带/链轨/履带板", "keywords": ["履带板", "履带", "链轨", "链节"]},
    {"major": "履带与行走机构", "minor": "支重轮/托链轮", "keywords": ["支重轮", "托链轮", "托轮"]},
    {"major": "履带与行走机构", "minor": "引导轮/张紧机构", "keywords": ["张紧油缸", "涨紧油缸", "张紧装置", "涨紧装置", "引导轮"]},
    {"major": "履带与行走机构", "minor": "行走马达/减速器", "keywords": ["维修行走", "行走减速器", "行走马达", "左侧行走", "右侧行走", "行走卡"]},
    {"major": "工作装置", "minor": "铲斗/斗齿/刀板", "keywords": ["工作装置牙齿", "松土齿", "推土铲", "铲斗", "斗齿", "齿座", "牙齿", "刀板"]},
    {"major": "工作装置", "minor": "动臂/斗杆/连杆", "keywords": ["工作装置连杆", "动臂", "大臂", "小臂", "斗杆"]},
    {"major": "工作装置", "minor": "销轴/衬套", "keywords": ["工作装置衬套", "工作装置销", "铲斗销", "大臂销", "小臂销", "斗杆销"]},
    {"major": "工作装置", "minor": "举升/翻斗机构", "keywords": ["车厢举升机构", "料斗提升机构", "翻斗机构"]},
    {"major": "工作装置", "minor": "属具/护板/通用", "keywords": ["工作装置护板", "铲斗护板", "松土器", "破碎锤", "抓斗"]},
    {"major": "驾驶室与车身", "minor": "车门/玻璃/后视镜", "keywords": ["驾驶室门", "后视镜", "车门", "玻璃", "门锁"]},
    {"major": "驾驶室与车身", "minor": "座椅/操纵踏板", "keywords": ["油门踏板", "刹车踏板", "操纵杆", "座椅", "踏板"]},
    {"major": "驾驶室与车身", "minor": "雨刷/清洗器", "keywords": ["玻璃水", "喷水器", "雨刷", "雨刮"]},
    {"major": "驾驶室与车身", "minor": "梯子/护栏/平台", "keywords": ["扶梯", "梯子", "护栏", "平台", "走台"]},
    {"major": "驾驶室与车身", "minor": "机罩/盖板/护板", "keywords": ["挡泥板", "底护板", "防护罩", "机罩", "盖板", "护网"]},
    {"major": "空调与暖风", "minor": "制冷回路", "keywords": ["空调压缩机", "制冷剂漏", "氟利昂漏", "冷凝器", "蒸发器", "膨胀阀"], "all_keywords": [["空调", "皮带"], ["空调", "轴承"], ["压缩机", "皮带"], ["压缩机", "轴承"]]},
    {"major": "空调与暖风", "minor": "风机/风道/滤芯", "keywords": ["空调滤芯", "空调风机", "暖风风机", "鼓风机", "风道"]},
    {"major": "空调与暖风", "minor": "暖风/加热", "keywords": ["暖风水管", "加热器", "暖风"]},
    {"major": "空调与暖风", "minor": "控制/性能异常", "keywords": ["空调不工作", "空调不制冷", "空调异响", "空调控制", "不制冷", "无冷气"]},
    {"major": "集中润滑系统", "minor": "润滑泵/控制", "keywords": ["自动润滑泵", "润滑控制器", "润滑泵", "黄油泵"]},
    {"major": "集中润滑系统", "minor": "管路/分配器/注油器", "keywords": ["润滑分配器", "润滑接头", "润滑管", "黄油管", "注油器"]},
    {"major": "集中润滑系统", "minor": "油箱/补脂作业", "keywords": ["加注黄油", "补加黄油", "充黄油", "黄油箱", "打点油", "打黄油"]},
    {"major": "集中润滑系统", "minor": "报警/不工作", "keywords": ["润滑不好", "润滑不良", "润滑不工作", "润滑无压力", "润滑报警", "润滑堵"]},
    {"major": "结构件与通用机械", "minor": "裂纹/断裂/焊修", "keywords": ["焊接维修", "结构裂纹", "支架裂", "开焊", "脱焊", "焊补"], "priority": -20},
    {"major": "结构件与通用机械", "minor": "支架/底座/护架", "keywords": ["固定架", "支架", "底座", "托架"], "priority": -20},
    {"major": "结构件与通用机械", "minor": "螺栓/紧固件", "keywords": ["紧固件", "螺栓", "螺母", "卡子"], "priority": -20},
    {"major": "结构件与通用机械", "minor": "通用轴承/皮带/联接件", "keywords": ["轴承", "皮带", "平键", "花键", "销子"], "priority": -20},
    {"major": "电缆与卷筒系统", "minor": "供电电缆", "keywords": ["供电电缆", "拖曳电缆", "电缆接头", "主电缆"]},
    {"major": "电缆与卷筒系统", "minor": "卷筒/滚筒", "keywords": ["电缆卷筒", "电缆滚筒", "卷筒", "滚筒"]},
    {"major": "电缆与卷筒系统", "minor": "集电环/收放控制", "keywords": ["卷筒控制", "集电环", "滑环", "收缆", "放缆"]},
    {"major": "安全消防与事故", "minor": "灭火/消防系统", "keywords": ["自动灭火", "防火系统", "灭火系统", "灭火器", "消防管"]},
    {"major": "安全消防与事故", "minor": "火灾/烧损", "keywords": ["烧车", "起火", "着火", "火灾", "烧损"], "exclude_keywords": ["不着火"]},
    {"major": "安全消防与事故", "minor": "碰撞/倾覆/外力损坏", "keywords": ["外力损坏", "碰撞", "翻车", "倾覆", "撞坏", "被撞", "落物"], "exclude_keywords": ["撞击飞轮"]},
    {"major": "安全消防与事故", "minor": "安全装置", "keywords": ["倒车报警器", "倒车影像", "安全带", "摄像头", "急停"]},
    {"major": "安全消防与事故", "minor": "资产报废/重大损失", "keywords": ["资产损失", "财产损失", "整机报废", "重大损失"]},
    {"major": "计划保养与非故障作业", "minor": "周期保养", "keywords": ["小时保养", "定期保养", "周期保养", "pm保养"], "regex_keywords": [r"\b\d+\s*小时保养"]},
    {"major": "计划保养与非故障作业", "minor": "滤芯保养", "keywords": ["吹清发动机空滤", "吹清空滤", "吹空滤", "吹滤芯", "更换滤芯", "换滤芯", "空滤吹风"]},
    {"major": "计划保养与非故障作业", "minor": "油液补加/更换", "keywords": ["发动机已补加", "补加防冻液", "补加液压油", "补加机油", "更换机油", "升防冻液", "升液压油", "升机油"]},
    {"major": "计划保养与非故障作业", "minor": "润滑/补脂", "keywords": ["加注黄油", "充黄油", "打点润滑", "打点油", "打黄油", "定期润滑"], "regex_keywords": [r"加注?\s*\d+\s*公斤黄油", r"充\s*\d+\s*公斤黄油"]},
    {"major": "计划保养与非故障作业", "minor": "轮胎充气/换位", "keywords": ["轮胎扭力矩", "轮胎效力矩", "给轮胎补气", "调节胎压", "轮胎换位", "轮胎倒换", "充氮气"], "priority": 10},
    {"major": "计划保养与非故障作业", "minor": "点检正常", "keywords": ["点检正常", "已点检,正常", "已点检，正常", "检查正常", "均为正常", "正常出车"]},
    {"major": "计划保养与非故障作业", "minor": "清洁/整理/交接", "keywords": ["清理场地", "清洁设备", "冲洗设备", "交接班", "转场", "清灰"]},
]

_DEFAULT_NOISE_EXACT: set[str] = {
    "(无注释)",
    "交接班",
    "停车",
    "出车",
    "出车。",
    "出车了",
    "启动检查",
    "均为正常",
    "夜班",
    "已点检",
    "正常",
    "点检",
    "点检时",
    "白班",
    "着车",
    "着车，出车",
    "着车，出车。",
    "计划点检",
}

_DEFAULT_NOISE_PATTERNS: list[str] = [
    r"^已?点检[，,/\s]*正常[。]?\s*$",
    r"^Author:.*",
    r"^MTC Translator:",
    r"^[已点检正常，,/\s\.。]+$",
    r"^(已点检|点检)[，,/\s]+.*正常.*$",
    r"^(白班|夜班)[：:]?\s*(已?点检)?[，,/\s]*正常[。]?\s*$",
    r"^(着车|搭接着车|外接电源启动)[，,\s]*出车[。]?\s*$",
    r"^(夜班|白班)[：:]?\s*$",
    r"^\d+PM\s*$",
    r"^计划点检[：:]?\s*(已?点检)?[，,/\s]*(正常)?[。]?\s*$",
    r"^(已?点检|检查|查看|试车)?[，,/\s]*(均为|一切)?正常[。]?\s*$",
    r"^(已点检|点检)[，,/\s]*(左侧|右侧)?[。]?\s*$",
    r"^(对设备)?进行检查[，,\s]*出车[。]?\s*$",
    r"^打着\s*$",
    r"^(由)?ETT验车[，,\s]*停车[。]?\s*$",
]

_DEFAULT_REASON_RULES: dict[str, str] = {
    "检修": "fault",
    "点检": "check_content",
    "保养": "non_fault",
    "待机": "skip",
}

_ROUTINE_MAINTENANCE_TERMS = (
    "小时保养", "定期保养", "周期保养", "pm保养",
    "吹清空滤", "吹空滤", "吹滤芯",
    "补加机油", "升机油", "补加防冻液", "升防冻液",
    "补加液压油", "升液压油", "打黄油", "充黄油",
    "加注黄油", "打点油", "调节胎压", "给轮胎补气", "充氮气",
)
_FAULT_MARKERS = (
    "故障", "报警", "损坏", "坏了", "坏", "漏油", "漏水", "漏气", "漏", "泄漏",
    "渗油", "渗", "异响", "磨损", "破裂", "断裂", "断", "裂纹", "无力", "高温", "压力异常",
    "不工作", "不能", "无法", "卡住", "卡死", "烧坏", "报废",
    "启动不了", "打不着", "启不动",
)


# ── Excel 模板样式 ────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_header(cell):
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER


def _set_cell(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    cell.border = _THIN_BORDER
    return cell


# ── 公开 API ──────────────────────────────────────────────────

def get_default_classifications() -> dict:
    """返回硬编码默认分类配置（20 大类，其他类由兜底逻辑生成）。"""
    return {
        "schema_version": MAINTENANCE_CLASSIFICATION_SCHEMA_VERSION,
        "classifications": [dict(c) for c in _DEFAULT_CLASSIFICATIONS],
        "noise_exact": set(_DEFAULT_NOISE_EXACT),
        "noise_patterns": list(_DEFAULT_NOISE_PATTERNS),
        "reason_rules": dict(_DEFAULT_REASON_RULES),
    }


def compile_noise_patterns(patterns: list[str]) -> list[re.Pattern]:
    """编译正则模式列表，跳过无效模式并记录警告。

    供调用方预编译一次后传入 is_fault_record / classify，避免每条记录重复编译。

    Args:
        patterns: 正则模式字符串列表。

    Returns:
        编译后的 re.Pattern 列表。
    """
    compiled = []
    for pat in patterns:
        try:
            compiled.append(re.compile(pat, re.IGNORECASE))
        except re.error:
            logger.warning("无效的噪声正则模式，已跳过: %s", pat)
    return compiled


@lru_cache(maxsize=131072)
def normalize_maintenance_content(content: str) -> str:
    """规范化维修内容并移除不应参与分类的元数据。"""
    if content is None:
        return ""
    text = unicodedata.normalize("NFKC", str(content)).casefold()
    text = re.sub(
        r"\b(author|mtc translator|site translator|lei\.gen)\s*:\s*[^;；,\n]*[;；,]?",
        " ",
        text,
    )
    text = re.sub(
        r"(发动机|发电机|设备|工作)?\s*小时(数|计)?\s*[:：;；]?\s*[\d.,]+",
        " ",
        text,
    )
    text = re.sub(r"\b\d+(?:\.\d+)?\s*(小时|分钟)\b", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,，。;；")


def _is_noise(
    content: str,
    noise_exact: set[str],
    compiled_noise: list[re.Pattern],
) -> bool:
    if not content:
        return True
    normalized_exact = {normalize_maintenance_content(value) for value in noise_exact}
    if content in normalized_exact:
        return True
    return any(pattern.match(content) for pattern in compiled_noise)


def _is_routine_maintenance_only(content: str) -> bool:
    """识别只描述例行保养、且不含故障信号的内容。"""
    return (
        any(term in content for term in _ROUTINE_MAINTENANCE_TERMS)
        and not any(marker in content for marker in _FAULT_MARKERS)
    )


def is_fault_record(
    reason: str,
    content: str,
    *,
    noise_exact: set[str] | None = None,
    noise_patterns: list[str] | None = None,
    compiled_noise: list[re.Pattern] | None = None,
    reason_rules: dict[str, str] | None = None,
) -> bool:
    """根据原因和维修内容判断是否为故障记录。

    Args:
        reason: 原因类型（检修/点检/保养/待机）。
        content: 维修内容文本。
        noise_exact: 精确匹配噪声集合，None 时使用默认值。
        noise_patterns: 正则噪声模式列表（字符串），None 时使用默认值。
            如果提供了 compiled_noise，则忽略此参数。
        compiled_noise: 预编译的噪声正则列表，优先于 noise_patterns 使用。
        reason_rules: 原因判定规则，None 时使用默认值。

    Returns:
        True 表示故障记录。
    """
    if noise_exact is None:
        noise_exact = _DEFAULT_NOISE_EXACT
    if reason_rules is None:
        reason_rules = _DEFAULT_REASON_RULES
    if compiled_noise is None:
        patterns = noise_patterns if noise_patterns is not None else _DEFAULT_NOISE_PATTERNS
        compiled_noise = compile_noise_patterns(patterns)

    normalized = normalize_maintenance_content(content)
    rule = reason_rules.get(str(reason).strip(), "check_content")

    if rule == "skip":
        return False
    if rule == "non_fault":
        return False
    if _is_noise(normalized, noise_exact, compiled_noise):
        return False
    if _is_routine_maintenance_only(normalized):
        return False
    if rule == "fault":
        return True

    # check_content: 检查内容是否有实质故障描述
    # 含"正常"且无其他实质描述 → 非故障
    if "正常" in normalized:
        if not any(marker in normalized for marker in _FAULT_MARKERS):
            return False
        cleaned = re.sub(r"[已点检，,/\s正常。\.]+", "", normalized)
        if len(cleaned.strip()) < 3:
            return False
    return True


def _group_by_major(classifications: list[dict]) -> dict[str, list[dict]]:
    """按大类分组，保留各组内原始顺序。

    Args:
        classifications: 分类规则列表。

    Returns:
        按大类分组的 OrderedDict，值为该大类下的规则列表（保持原顺序）。
    """
    grouped: dict[str, list[dict]] = {}
    for entry in classifications:
        major = entry["major"]
        grouped.setdefault(major, []).append(entry)
    return grouped


def _best_major(content: str, grouped: dict[str, list[dict]]) -> str | None:
    """从所有大选中选出最佳大类。

    评分规则（元组比较）：
      主指标 = 该大类下命中关键词的小类数量（entry_count）
      次指标 = 该大类下所有命中关键词中的最长字符数（max_keyword_len）
      两个指标均更高者胜出；平局时优先返回列表靠前的大类。

    Args:
        content: 维修内容文本。
        grouped: 按大类的分组数据（顺序保留）。

    Returns:
        最佳大类名称，无任何关键字匹配时返回 None。
    """
    best = None
    best_score = (0, 0)
    for major, entries in grouped.items():
        entry_count = 0
        max_len = 0
        for entry in entries:
            matched_any = False
            for kw in entry["keywords"]:
                if kw in content:
                    kw_len = len(kw)
                    if kw_len > max_len:
                        max_len = kw_len
                    matched_any = True
            if matched_any:
                entry_count += 1
        if entry_count > 0 and (entry_count, max_len) > best_score:
            best_score = (entry_count, max_len)
            best = major
    return best


def _entry_match_score(content: str, entry: dict, index: int) -> tuple | None:
    """返回规则命中的可比较分数；未命中时返回 None。"""
    excludes = [
        normalize_maintenance_content(value)
        for value in entry.get("exclude_keywords", [])
        if str(value).strip()
    ]
    if any(value and value in content for value in excludes):
        return None

    if entry.get("major") == "计划保养与非故障作业" and any(marker in content for marker in _FAULT_MARKERS):
        return None

    candidates: list[tuple[int, int, int]] = []
    matched_terms: set[str] = set()
    for keyword in entry.get("keywords", []):
        normalized = normalize_maintenance_content(keyword)
        if normalized and normalized in content:
            candidates.append((len(normalized), content.find(normalized), 1))
            matched_terms.add(normalized)

    for group in entry.get("all_keywords", []):
        normalized_group = [
            normalize_maintenance_content(keyword)
            for keyword in group
            if str(keyword).strip()
        ]
        if normalized_group and all(keyword in content for keyword in normalized_group):
            positions = [content.find(keyword) for keyword in normalized_group]
            specificity = sum(len(keyword) for keyword in normalized_group) + 2
            candidates.append((specificity, min(positions), len(normalized_group)))
            matched_terms.update(normalized_group)

    for pattern in entry.get("regex_keywords", []):
        try:
            match = re.search(pattern, content, re.IGNORECASE)
        except re.error:
            logger.warning("无效的分类正则模式，已跳过: %s", pattern)
            continue
        if match:
            matched = match.group(0)
            candidates.append((len(matched), match.start(), 2))
            matched_terms.add(matched)

    if not candidates:
        return None

    specificity, first_pos, group_strength = max(
        candidates,
        key=lambda item: (item[0], item[2], -item[1]),
    )
    priority = int(entry.get("priority", 0) or 0)
    return (
        priority,
        specificity,
        group_strength,
        len(matched_terms),
        -first_pos,
        -index,
    )


def _best_entry_and_ambiguity(
    content: str,
    classifications: list[dict],
) -> tuple[dict | None, bool]:
    matches: list[tuple[tuple, dict]] = []
    for index, entry in enumerate(classifications):
        score = _entry_match_score(content, entry, index)
        if score is not None:
            matches.append((score, entry))
    if not matches:
        return None, False

    best_score, best_entry = max(matches, key=lambda item: item[0])
    semantic_score = best_score[:-1]  # 忽略规则行序，仅判断业务评分是否并列
    tied_majors = {
        entry["major"]
        for score, entry in matches
        if score[:-1] == semantic_score
    }
    return best_entry, len(tied_majors) > 1


def _best_entry(content: str, classifications: list[dict]) -> dict | None:
    """返回最高分规则；同分歧义由 classify() 统一进入待确认。"""
    return _best_entry_and_ambiguity(content, classifications)[0]


def classify(
    content: str,
    *,
    classifications: list[dict] | None = None,
    noise_exact: set[str] | None = None,
    noise_patterns: list[str] | None = None,
    compiled_noise: list[re.Pattern] | None = None,
) -> tuple[str | None, str | None]:
    """对维修内容进行大类+小类分类。

    全局规则匹配：
      1. 先规范化文本并移除作者、翻译器、发动机/发电机小时数等元数据；
      2. 支持普通关键词、组合关键词、排除关键词和正则关键词；
      3. 显式优先级先处理已确认边界，再按具体命中长度、组合强度和位置判定。

    Args:
        content: 维修内容文本。
        classifications: 分类规则列表，None 时使用默认值。
        noise_exact: 精确噪声集合，None 时使用默认值。
        noise_patterns: 正则噪声列表（字符串），None 时使用默认值。
            如果提供了 compiled_noise，则忽略此参数。
        compiled_noise: 预编译的噪声正则列表，优先于 noise_patterns 使用。

    Returns:
        (大类, 小类)，无实质内容时返回 (None, None)。
    """
    if classifications is None:
        classifications = _DEFAULT_CLASSIFICATIONS
    if noise_exact is None:
        noise_exact = _DEFAULT_NOISE_EXACT
    if compiled_noise is None:
        patterns = noise_patterns if noise_patterns is not None else _DEFAULT_NOISE_PATTERNS
        compiled_noise = compile_noise_patterns(patterns)

    normalized = normalize_maintenance_content(content)
    if _is_noise(normalized, noise_exact, compiled_noise):
        return None, None

    # 纯例行作业应优先落入计划保养，避免“充黄油、补加油液”等通用动作
    # 被集中润滑或具体系统规则抢占；一旦出现故障信号则仍由故障规则优先。
    if _is_routine_maintenance_only(normalized):
        maintenance_entries = [
            entry
            for entry in classifications
            if entry.get("major") == "计划保养与非故障作业"
        ]
        maintenance_entry, maintenance_ambiguous = _best_entry_and_ambiguity(
            normalized,
            maintenance_entries,
        )
        if maintenance_entry is not None and not maintenance_ambiguous:
            return maintenance_entry["major"], maintenance_entry["minor"]

    best_entry, is_ambiguous = _best_entry_and_ambiguity(normalized, classifications)
    if best_entry is None:
        if any(marker in normalized for marker in _FAULT_MARKERS):
            return "其他/待确认", "仅现象未定位"
        return "其他/待确认", "信息不足"
    if is_ambiguous:
        return "其他/待确认", "多系统/需拆分"
    return best_entry["major"], best_entry["minor"]


def import_classifications_from_excel(path: str) -> dict:
    """从 Excel 配置模板导入分类规则。

    Sheet 结构：
    - "分类规则": 大类 | 小类 | 关键词 | 组合关键词 | 排除关键词 | 正则关键词 | 优先级
    - "噪声过滤": 类型（精确匹配/正则） | 值
    - "原因规则": 原因 | 处理方式（故障/检查内容/非故障/跳过）

    Args:
        path: Excel 文件路径。

    Returns:
        完整分类配置 dict，结构同 get_default_classifications()。

    Raises:
        ValueError: 文件格式错误或缺少必要 sheet。
    """
    with pd.ExcelFile(path) as xl:
        sheet_names = set(xl.sheet_names)

        # ── Sheet 1: 分类规则 ──
        if "分类规则" not in sheet_names:
            raise ValueError("Excel 文件缺少 '分类规则' sheet")

        classifications = []
        df = xl.parse("分类规则").fillna("")
        for _, row in df.iterrows():
            major = str(row.get("大类", "")).strip()
            minor = str(row.get("小类", "")).strip()
            kw_raw = str(row.get("关键词", "")).strip()
            if not major or not minor:
                continue
            keywords = [k.strip() for k in kw_raw.split("、") if k.strip()]
            all_raw = str(row.get("组合关键词", "")).strip()
            all_keywords = []
            for group in re.split(r"[；;]", all_raw):
                values = [value.strip() for value in re.split(r"[&＋+]", group) if value.strip()]
                if values:
                    all_keywords.append(values)
            exclude_raw = str(row.get("排除关键词", "")).strip()
            exclude_keywords = [k.strip() for k in exclude_raw.split("、") if k.strip()]
            regex_raw = str(row.get("正则关键词", "")).strip()
            regex_keywords = [k.strip() for k in regex_raw.split("、") if k.strip()]
            priority_raw = row.get("优先级", "")
            try:
                priority = int(float(priority_raw)) if str(priority_raw).strip() else 0
            except (TypeError, ValueError):
                priority = 0

            if keywords or all_keywords or regex_keywords:
                entry = {"major": major, "minor": minor, "keywords": keywords}
                if all_keywords:
                    entry["all_keywords"] = all_keywords
                if exclude_keywords:
                    entry["exclude_keywords"] = exclude_keywords
                if regex_keywords:
                    entry["regex_keywords"] = regex_keywords
                if priority:
                    entry["priority"] = priority
                classifications.append(entry)

        # ── Sheet 2: 噪声过滤 ──
        noise_exact: set[str] = set()
        noise_patterns: list[str] = []
        if "噪声过滤" in sheet_names:
            df = xl.parse("噪声过滤")
            for row in df.itertuples(index=False):
                filter_type = str(row[0]).strip() if row[0] is not None and str(row[0]).strip() != "nan" else ""
                value = str(row[1]).strip() if len(row) > 1 and row[1] is not None and str(row[1]).strip() != "nan" else ""
                if not value:
                    continue
                if filter_type == "精确匹配":
                    noise_exact.add(value)
                elif filter_type == "正则":
                    noise_patterns.append(value)

        # ── Sheet 3: 原因规则 ──
        reason_rules: dict[str, str] = {}
        _REASON_MAP = {"故障": "fault", "检查内容": "check_content", "非故障": "non_fault", "跳过": "skip"}
        if "原因规则" in sheet_names:
            df = xl.parse("原因规则")
            for row in df.itertuples(index=False):
                reason = str(row[0]).strip() if row[0] is not None and str(row[0]).strip() != "nan" else ""
                method = str(row[1]).strip() if len(row) > 1 and row[1] is not None and str(row[1]).strip() != "nan" else ""
                mapped = _REASON_MAP.get(method)
                if reason and mapped:
                    reason_rules[reason] = mapped

    # 空值兜底
    if not classifications:
        logger.warning("导入的分类规则为空，使用默认值")
        classifications = [dict(c) for c in _DEFAULT_CLASSIFICATIONS]
    if not noise_exact:
        noise_exact = set(_DEFAULT_NOISE_EXACT)
    if not noise_patterns:
        noise_patterns = list(_DEFAULT_NOISE_PATTERNS)
    if not reason_rules:
        reason_rules = dict(_DEFAULT_REASON_RULES)

    logger.info("从 Excel 导入分类配置: %d 条规则, %d 个精确噪声, %d 个正则噪声",
                len(classifications), len(noise_exact), len(noise_patterns))
    return {
        "classifications": classifications,
        "noise_exact": noise_exact,
        "noise_patterns": noise_patterns,
        "reason_rules": reason_rules,
    }


def export_classification_template(path: str, *, with_defaults: bool = False) -> str:
    """导出维修分类 Excel 配置模板。

    Args:
        path: 输出文件路径。
        with_defaults: True 时填充默认数据，False 时只输出表头和示例行。

    Returns:
        输出文件路径。
    """
    data = get_default_classifications() if with_defaults else None
    classifications = data["classifications"] if data else _DEFAULT_CLASSIFICATIONS[:3]
    noise_exact = data["noise_exact"] if data else {"出车", "已点检"}
    noise_patterns = data["noise_patterns"] if data else [r"^已?点检[，,/\s]*正常[。]?\s*$"]
    reason_rules = data["reason_rules"] if data else _DEFAULT_REASON_RULES

    wb = openpyxl.Workbook()

    # ── Sheet 1: 分类规则 ──
    ws1 = wb.active
    ws1.title = "分类规则"
    headers1 = ["大类", "小类", "关键词", "组合关键词", "排除关键词", "正则关键词", "优先级"]
    for col, h in enumerate(headers1, 1):
        _apply_header(ws1.cell(row=1, column=col, value=h))
    for row_idx, entry in enumerate(classifications, 2):
        _set_cell(ws1, row_idx, 1, entry["major"])
        _set_cell(ws1, row_idx, 2, entry["minor"])
        _set_cell(ws1, row_idx, 3, "、".join(entry.get("keywords", [])))
        _set_cell(
            ws1, row_idx, 4,
            "；".join("&".join(group) for group in entry.get("all_keywords", [])),
        )
        _set_cell(ws1, row_idx, 5, "、".join(entry.get("exclude_keywords", [])))
        _set_cell(ws1, row_idx, 6, "、".join(entry.get("regex_keywords", [])))
        _set_cell(ws1, row_idx, 7, entry.get("priority", 0))
        for col in range(3, 7):
            ws1.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 60
    ws1.column_dimensions["D"].width = 36
    ws1.column_dimensions["E"].width = 36
    ws1.column_dimensions["F"].width = 36
    ws1.column_dimensions["G"].width = 10
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:G{max(len(classifications), 1) + 1}"

    # ── Sheet 2: 噪声过滤 ──
    ws2 = wb.create_sheet("噪声过滤")
    headers2 = ["类型", "值"]
    for col, h in enumerate(headers2, 1):
        _apply_header(ws2.cell(row=1, column=col, value=h))
    row_idx = 2
    for val in sorted(noise_exact):
        _set_cell(ws2, row_idx, 1, "精确匹配")
        _set_cell(ws2, row_idx, 2, val)
        row_idx += 1
    for pat in noise_patterns:
        _set_cell(ws2, row_idx, 1, "正则")
        _set_cell(ws2, row_idx, 2, pat)
        row_idx += 1
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 60
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 原因规则 ──
    ws3 = wb.create_sheet("原因规则")
    headers3 = ["原因", "处理方式", "说明"]
    for col, h in enumerate(headers3, 1):
        _apply_header(ws3.cell(row=1, column=col, value=h))
    _REASON_DESC = {
        "fault": "故障", "check_content": "检查内容",
        "non_fault": "非故障", "skip": "跳过",
    }
    _REASON_EXPLAIN = {
        "fault": "视为故障记录", "check_content": "需检查维修内容是否为噪声",
        "non_fault": "不视为故障", "skip": "不进入明细",
    }
    for row_idx, (reason, method) in enumerate(reason_rules.items(), 2):
        _set_cell(ws3, row_idx, 1, reason)
        _set_cell(ws3, row_idx, 2, _REASON_DESC.get(method, method))
        _set_cell(ws3, row_idx, 3, _REASON_EXPLAIN.get(method, ""))
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 30
    ws3.freeze_panes = "A2"

    wb.save(path)
    logger.info("分类配置模板已导出: %s", path)
    return path
