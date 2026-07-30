"""处理宽格式柴油消耗报表（单文件脚本，无项目内部依赖）。

自动识别两种格式：
  - 带设备编号列（Парк дугаар）：A=序号, B=设备名称, C=设备编号, 之后=消耗值
  - 不带设备编号列：A=序号, B=设备名称, 之后=消耗值

用法:
    python excel_fuel_wide.py <输入文件> --year 2020 --month 4
"""

import argparse
import logging
import os
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(asctime)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# 用于识别表头行的关键字
_HEADER_KEYWORDS = {"Атомашин төрөл", "Парк дугаар", "Техник", "设备名称", "设备编号"}


# ── 自动检测工具 ──────────────────────────────────────────────────────────


def _find_header_row(ws) -> int | None:
    """扫描前 10 行，找到包含任意 _HEADER_KEYWORDS 的行。"""
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, min(6, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            if val and str(val).strip() in _HEADER_KEYWORDS:
                return row_idx
    return None


def _find_col_by_header(ws, header_row: int, *keywords: str) -> int | None:
    """在 header_row 中查找包含任一关键字的列（1-based），未找到返回 None。"""
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col_idx).value
        if val:
            text = str(val).strip()
            if text in keywords:
                return col_idx
    return None


def _find_day_header_row(ws, header_row: int) -> int | None:
    """从 header_row 之后扫描，找到包含连续日期数字 (1-31) 的行。"""
    for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
        nums = []
        started = False
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                if started:
                    break
                continue
            try:
                n = int(val)
                if 1 <= n <= 31:
                    started = True
                    nums.append(n)
                else:
                    if started:
                        break
            except (ValueError, TypeError):
                if started:
                    break
        if len(nums) >= 5 and nums == list(range(nums[0], nums[0] + len(nums))):
            return row_idx
    return None


def _find_day_col_start(ws, day_header_row: int) -> int | None:
    """返回日期编号起始列（1-based），即第一个 1-31 数字所在的列。"""
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(day_header_row, col_idx).value
        if val is None:
            continue
        try:
            n = int(val)
            if 1 <= n <= 31:
                return col_idx
        except (ValueError, TypeError):
            pass
    return None


def _detect_data_start_row(ws, day_header_row: int) -> int:
    """从 day_header_row 之后扫描，找到 A 列首个为数字的行（序号列）。"""
    for row_idx in range(day_header_row + 1, ws.max_row + 1):
        val = ws.cell(row_idx, 1).value
        if val is not None:
            try:
                int(val)
                return row_idx
            except (ValueError, TypeError):
                pass
    return day_header_row + 1


# ── 格式化输出 ────────────────────────────────────────────────────────────


def _write_formatted_excel(output_file: str, sheets: dict[str, pd.DataFrame]) -> str:
    """写入带格式的 Excel：表头蓝底白字、列宽自适应、日期列 yyyy-mm-dd、冻结首行。"""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    wb = openpyxl.load_workbook(output_file)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for name, df in sheets.items():
        ws = wb[name]
        if ws.max_row is None or ws.max_row < 1:
            continue

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_w = len(str(col_name)) + 2
            for val in df[col_name].dropna().head(500):
                max_w = max(max_w, len(str(val)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_w, 8), 50)

        for col_idx, col_name in enumerate(df.columns, start=1):
            sample = df[col_name].dropna().head(20)
            if sample.apply(lambda v: isinstance(v, (date,))).any():
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = "yyyy-mm-dd"

        ws.freeze_panes = "A2"
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    wb.save(output_file)
    wb.close()
    return output_file


# ── 核心处理 ──────────────────────────────────────────────────────────────


def process_fuel_wide(
    file_path: str,
    year: int,
    month: int,
    *,
    return_df: bool = False,
) -> str | pd.DataFrame:
    """处理宽格式柴油消耗 Excel，输出长格式结果。

    自动识别表头行、日期编号行、设备编号列和数据起始行。

    Args:
        file_path: 输入 Excel 文件路径。
        year: 年份（如 2020）。
        month: 月份（如 4）。
        return_df: 若为 True 返回 DataFrame 而非写入文件。

    Returns:
        当 return_df=False 时返回输出文件路径；
        当 return_df=True 时返回 DataFrame。

    Raises:
        ValueError: 文件结构无法识别或未找到有效数据。
    """
    logger.info("正在处理宽格式柴油报表: %s (年=%d, 月=%d)", file_path, year, month)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 1. 识别表头行
    header_row = _find_header_row(ws)
    if header_row is None:
        wb.close()
        raise ValueError("未找到包含设备名称关键字的表头行")
    logger.info("表头行: %d", header_row)

    # 2. 识别设备名称列和设备编号列
    name_col = _find_col_by_header(ws, header_row, "Атомашин төрөл", "设备名称")
    id_col = _find_col_by_header(ws, header_row, "Парк дугаар", "设备编号")
    if name_col is None:
        wb.close()
        raise ValueError("未找到设备名称列（Атомашин төрөл / 设备名称）")
    logger.info("设备名称列: %d, 设备编号列: %s", name_col, id_col if id_col else "无")

    # 3. 识别日期编号行
    day_header_row = _find_day_header_row(ws, header_row)
    if day_header_row is None:
        wb.close()
        raise ValueError("未找到日期编号行")
    logger.info("日期编号行: %d", day_header_row)

    # 4. 构建列→日映射
    day_col_start = _find_day_col_start(ws, day_header_row)
    if day_col_start is None:
        wb.close()
        raise ValueError("日期编号行中未找到起始日期 (1)")

    col_to_day: dict[int, int] = {}
    for col_idx in range(day_col_start, ws.max_column + 1):
        val = ws.cell(day_header_row, col_idx).value
        if val is not None:
            try:
                day_num = int(val)
                if 1 <= day_num <= 31:
                    col_to_day[col_idx] = day_num
            except (ValueError, TypeError):
                break

    if not col_to_day:
        wb.close()
        raise ValueError(f"在第 {day_header_row} 行未找到有效的日期编号列")
    logger.info("检测到 %d 个日期列（从第 %d 列起）", len(col_to_day), day_col_start)

    # 5. 自动检测数据起始行
    data_start_row = _detect_data_start_row(ws, day_header_row)
    logger.info("数据起始行: %d", data_start_row)

    # 6. 遍历数据行，融化为长格式
    has_id = id_col is not None
    records: list[dict] = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        raw_name = ws.cell(row_idx, name_col).value
        if raw_name is None:
            continue
        device_name = str(raw_name).strip()
        if not device_name:
            continue

        device_id = ""
        if has_id:
            raw_id = ws.cell(row_idx, id_col).value
            device_id = str(raw_id).strip() if raw_id is not None else ""

        for col_idx, day in col_to_day.items():
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                continue
            try:
                fuel = float(val)
            except (ValueError, TypeError):
                continue
            if fuel == 0:
                continue

            record: dict = {"日期": date(year, month, day), "设备名称": device_name}
            if has_id:
                record["设备编号"] = device_id
            record["油品消耗"] = fuel
            records.append(record)

    wb.close()

    if not records:
        raise ValueError("柴油消耗表中未找到有效数据（所有消耗值为 0 或为空）")

    df = pd.DataFrame(records)
    sort_cols = ["日期", "设备名称"] + (["设备编号"] if has_id else [])
    df.sort_values(by=sort_cols, inplace=True, ignore_index=True)

    n_devices = df.groupby(["设备名称"] + (["设备编号"] if has_id else [])).ngroups
    logger.info("提取 %d 条记录，%d 台设备", len(df), n_devices)

    if return_df:
        return df

    output_file = os.path.join(os.path.dirname(file_path) or ".", "Fuel_wide.xlsx")
    _write_formatted_excel(output_file, {"油耗信息": df})
    logger.info("处理完成！文件已保存: %s", output_file)
    return output_file


def main():
    parser = argparse.ArgumentParser(description="处理宽格式柴油消耗报表")
    parser.add_argument("input_file", help="输入 Excel 文件路径")
    parser.add_argument("--year", type=int, required=True, help="年份（如 2020）")
    parser.add_argument("--month", type=int, required=True, help="月份（如 4）")
    args = parser.parse_args()

    process_fuel_wide(args.input_file, year=args.year, month=args.month)


if __name__ == "__main__":
    main()
