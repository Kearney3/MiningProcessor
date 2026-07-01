"""
Excel 统一格式化输出工具

提供所有处理器共用的 Excel 写入与格式化能力：
- 表头加粗 + 蓝底白字
- 列宽自适应（支持中日韩宽字符）
- 日期列格式化（yyyy-mm-dd）
- 冻结首行 + 自动筛选
"""

import datetime
import logging
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

# ── 样式常量 ──────────────────────────────────────────────────────────────

HEADER_FILL = "4472C4"
HEADER_FONT_COLOR = "FFFFFF"
DATE_NUM_FORMAT = "yyyy-mm-dd"
MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 50
WIDTH_PADDING = 2


# ── 内部工具 ──────────────────────────────────────────────────────────────


def _display_width(text: str) -> int:
    """估算字符串显示宽度，CJK 字符计为 2。"""
    width = 0
    for ch in text:
        cp = ord(ch)
        if (
            0x4E00 <= cp <= 0x9FFF       # CJK Unified
            or 0x3400 <= cp <= 0x4DBF    # CJK Extension A
            or 0xFF00 <= cp <= 0xFFEF    # Fullwidth Forms
            or 0x3000 <= cp <= 0x303F    # CJK Symbols
            or 0x2E80 <= cp <= 0x2EFF    # CJK Radicals
            or 0xF900 <= cp <= 0xFAFF    # CJK Compatibility
            or 0xFE30 <= cp <= 0xFE4F    # CJK Compatibility Forms
        ):
            width += 2
        else:
            width += 1
    return width


def _auto_column_widths(
    df: pd.DataFrame,
    min_width: int = MIN_COL_WIDTH,
    max_width: int = MAX_COL_WIDTH,
    padding: int = WIDTH_PADDING,
) -> list[int]:
    """计算每列的自适应宽度（字符数）。"""
    widths = []
    for col in df.columns:
        header_w = _display_width(str(col)) + padding
        max_w = header_w
        for value in df[col].dropna().head(500):
            if isinstance(value, (pd.Timestamp, datetime.datetime, datetime.date)):
                cell_w = 12  # yyyy-mm-dd
            else:
                cell_w = _display_width(str(value)) + padding
            max_w = max(max_w, cell_w)
        widths.append(min(max(max_w, min_width), max_width))
    return widths


def _is_date_column(series: pd.Series) -> bool:
    """检测 Series 是否为日期列。"""
    if pd.api.types.is_datetime64_any_dtype(series):
        return True
    sample = series.dropna().head(20)
    if sample.empty:
        return False
    return sample.apply(
        lambda v: isinstance(v, (datetime.date, datetime.datetime, pd.Timestamp))
    ).any()


# ── 公开 API ──────────────────────────────────────────────────────────────


def write_formatted_excel(
    output_file: str,
    sheets: dict[str, pd.DataFrame],
    *,
    header_fill: str = HEADER_FILL,
    header_font_color: str = HEADER_FONT_COLOR,
    date_format: str = DATE_NUM_FORMAT,
    freeze_header: bool = True,
    auto_filter: bool = True,
    min_col_width: int = MIN_COL_WIDTH,
    max_col_width: int = MAX_COL_WIDTH,
) -> str:
    """写入带格式的 Excel 文件。

    对每个 sheet 应用统一格式：
    - 表头：加粗、蓝底白字
    - 列宽：按内容自适应（支持 CJK 宽字符）
    - 日期列：格式化为 yyyy-mm-dd
    - 冻结首行 + 自动筛选

    Args:
        output_file: 输出文件路径。
        sheets: {sheet_name: DataFrame} 映射。
        header_fill: 表头背景色（hex）。
        header_font_color: 表头字体颜色（hex）。
        date_format: 日期列的数字格式。
        freeze_header: 是否冻结首行。
        auto_filter: 是否启用自动筛选。
        min_col_width: 最小列宽（字符数）。
        max_col_width: 最大列宽（字符数）。

    Returns:
        输出文件路径。
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_file = str(output_file)

    # Step 1: 用 pandas 写入数据
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Step 2: 用 openpyxl 打开并格式化
    wb = load_workbook(output_file)

    header_style = Font(bold=True, color=header_font_color)
    fill_style = PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for sheet_name, df in sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        if ws.max_row is None or ws.max_row < 1:
            continue

        # 表头样式
        for cell in ws[1]:
            cell.font = header_style
            cell.fill = fill_style
            cell.alignment = center_align

        # 列宽自适应
        col_widths = _auto_column_widths(df, min_col_width, max_col_width)
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # 日期列格式化
        for col_idx, col_name in enumerate(df.columns):
            if _is_date_column(df[col_name]):
                excel_col = col_idx + 1  # openpyxl 1-based
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=excel_col)
                    if cell.value is not None:
                        cell.number_format = date_format

        # 冻结首行
        if freeze_header:
            ws.freeze_panes = "A2"

        # 自动筛选
        if auto_filter and ws.max_column:
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    wb.save(output_file)
    wb.close()

    logger.info("格式化输出完成: %s", output_file)
    return output_file
