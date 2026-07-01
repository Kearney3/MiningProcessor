"""
台账匹配数据处理模块
提供 Excel 导入、设备/油品批量匹配、结果导出等纯数据逻辑，
由 gui/components/ledger_match.py 薄封装调用。
"""
import datetime
import threading
from pathlib import Path
from typing import Any, Callable, Optional

import pandas as pd

from func.logger import get_logger
from func.excel_utils import strip_date_only_times

logger = get_logger(__name__)

DEVICE_RESULT_KEYS = ["标准设备名称", "标准设备编号", "标准公司名称"]
OIL_RESULT_KEY = "标准油品名称"
BATCH_SIZE = 100
EXPORT_BATCH = 1000


# ---------------------------------------------------------------------------
# 共享状态（不含任何 GUI 控件引用）
# ---------------------------------------------------------------------------

class MatchState:
    """Holds mutable data state shared between import/match/export steps.
    Free of Flet controls so it can be tested without GUI."""

    def __init__(self) -> None:
        self.all_sheets: dict[str, pd.DataFrame] = {}
        self.matched_all_sheets: dict[str, pd.DataFrame] = {}
        self.matched_sheets: dict[str, pd.DataFrame] = {}
        self.unmatched_sheets: dict[str, pd.DataFrame] = {}
        self.filtered_df: Optional[pd.DataFrame] = None
        self.current_sheet: str = ""
        self.page: int = 0
        self.columns: list[str] = []
        self.sort_column: Optional[str] = None
        self.sort_ascending: bool = True
        self.view_mode: str = "all"
        self.import_cancelled = threading.Event()

    def clear(self) -> None:
        self.all_sheets.clear()
        self.matched_all_sheets.clear()
        self.matched_sheets.clear()
        self.unmatched_sheets.clear()
        self.filtered_df = None
        self.current_sheet = ""
        self.page = 0
        self.columns.clear()
        self.sort_column = None
        self.sort_ascending = True
        self.view_mode = "all"
        self.import_cancelled.clear()


ProgressCallback = Callable[[float, str], None]  # (progress 0-1, message)


# ---------------------------------------------------------------------------
# Excel 导入
# ---------------------------------------------------------------------------


def import_excel(
    file_path: str,
    progress_cb: Optional[ProgressCallback] = None,
    cancel_event: Optional[threading.Event] = None,
) -> tuple[dict[str, pd.DataFrame], list[str]]:
    """Import all sheets from *file_path*.

    Returns ``(parsed_sheets, sheet_names)``.  *progress_cb* and
    *cancel_event* are optional hooks so the GUI can show a progress bar and
    honour a cancel button.
    """
    from openpyxl import load_workbook

    parsed_sheets: dict[str, pd.DataFrame] = {}
    cancelled = cancel_event or threading.Event()

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    total_sheets = len(sheet_names)

    if total_sheets == 0:
        wb.close()
        return parsed_sheets, sheet_names

    try:
        for sname in sheet_names:
            if cancelled.is_set():
                break

            ws = wb[sname]
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0

            if total_rows == 0 or total_cols == 0:
                parsed_sheets[sname] = pd.DataFrame()
                continue

            headers: list[str] = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = cell.value
                headers.append(str(val).strip() if val is not None else f"Col{cell.column - 1}")

            rows_data: list[list[Any]] = []
            rows_read = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if cancelled.is_set():
                    break
                rows_data.append(list(row))
                rows_read += 1
                if rows_read % 500 == 0 and progress_cb:
                    progress_cb(rows_read / total_rows if total_rows > 0 else 0,
                                f"正在导入 {sname}: {rows_read}/{total_rows} 行")

            if rows_data:
                header_len = len(headers)
                row_lens = set(len(r) for r in rows_data)
                if len(row_lens) > 1 or (row_lens and header_len not in row_lens):
                    logger.debug("Sheet %r: header_len=%d, row_lengths=%s", sname, header_len, row_lens)
                df = pd.DataFrame(rows_data, columns=headers)
                df = strip_date_only_times(df)
            else:
                df = pd.DataFrame(columns=headers)

            parsed_sheets[sname] = df
            logger.info("已导入 %s: %d 行, %d 列", sname, len(df), len(df.columns))
    finally:
        wb.close()

    return parsed_sheets, sheet_names


# ---------------------------------------------------------------------------
# 批量匹配（纯数据）
# ---------------------------------------------------------------------------

def _batch_match(
    df: pd.DataFrame, source_col: str, match_fn: Callable,
    result_keys: list[str], *, cancel_event: Optional[threading.Event] = None,
    id_col: Optional[str] = None, progress_cb: Optional[ProgressCallback] = None,
    progress_label: str = "", total_work: int = 0,
) -> dict[str, list]:
    """Apply *match_fn* row-by-row and return ``{key: [values...]}``."""
    cancelled = cancel_event or threading.Event()
    results: dict[str, list] = {k: [] for k in result_keys}
    total_rows = len(df)

    for start in range(0, total_rows, BATCH_SIZE):
        if cancelled.is_set():
            logger.info("匹配已取消")
            break
        batch = df.iloc[start:start + BATCH_SIZE]
        for _, row in batch.iterrows():
            n = (str(row[source_col])
                 if source_col and source_col in df.columns and not pd.isna(row.get(source_col))
                 else None)
            i_val = (str(row[id_col])
                     if id_col and id_col in df.columns and not pd.isna(row.get(id_col))
                     else None)
            r = match_fn(name=n, device_id=i_val) if id_col else (match_fn(n) if n else None)
            for k in result_keys:
                results[k].append(r.get(k, "") if r else "")

        if progress_cb and total_work > 0:
            processed = min(start + BATCH_SIZE, total_rows)
            progress_cb(processed / total_work, f"{progress_label}: {processed}/{total_work}")

    return results


# ---------------------------------------------------------------------------
# 匹配主流程（纯数据）
# ---------------------------------------------------------------------------

def match_sheet(
    df: pd.DataFrame, eq_ledger, oil_ledger,
    name_col: Optional[str], id_col: Optional[str], oil_col: Optional[str],
    cancel_event: Optional[threading.Event] = None,
    progress_cb: Optional[ProgressCallback] = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, int]:
    """Run equipment + oil matching on *df*.

    Returns ``(result_df, matched_df, unmatched_df, matched_count)``."""
    result_df = df.copy()
    total_rows = len(df)
    matched_count = 0

    has_truck = "矿卡名称" in result_df.columns
    has_excav = "挖机名称" in result_df.columns

    if eq_ledger and (name_col or id_col):
        if has_truck and has_excav:
            total_work = total_rows * 2
            truck_r = _batch_match(
                df, "矿卡名称", eq_ledger.match_device, DEVICE_RESULT_KEYS,
                cancel_event=cancel_event, id_col=id_col,
                progress_cb=progress_cb, progress_label="正在匹配矿卡", total_work=total_work,
            )
            for k, s in zip(DEVICE_RESULT_KEYS, ["（矿卡）"] * 3):
                result_df[k + s] = truck_r[k]
            matched_count += sum(1 for v in truck_r[DEVICE_RESULT_KEYS[0]] if v)

            excav_r = _batch_match(
                df, "挖机名称",
                lambda name, device_id=None: eq_ledger.match_device(name=name, device_id=device_id),
                DEVICE_RESULT_KEYS, cancel_event=cancel_event,
                progress_cb=progress_cb, progress_label="正在匹配挖机", total_work=total_work,
            )
            for k, s in zip(DEVICE_RESULT_KEYS, ["（挖机）"] * 3):
                result_df[k + s] = excav_r[k]
        else:
            equip_r = _batch_match(
                df, name_col, eq_ledger.match_device, DEVICE_RESULT_KEYS,
                cancel_event=cancel_event, id_col=id_col,
                progress_cb=progress_cb, progress_label="正在匹配设备", total_work=total_rows,
            )
            for k in DEVICE_RESULT_KEYS:
                result_df[k] = equip_r[k]
            matched_count += sum(1 for v in equip_r[DEVICE_RESULT_KEYS[0]] if v)

    if oil_ledger and oil_col and oil_col in result_df.columns:
        oil_r = _batch_match(
            df, oil_col, lambda v: oil_ledger.match(v), [OIL_RESULT_KEY],
            cancel_event=cancel_event,
            progress_cb=progress_cb, progress_label="正在匹配油品", total_work=total_rows,
        )
        result_df[OIL_RESULT_KEY] = oil_r[OIL_RESULT_KEY]

    # Split matched / unmatched
    if eq_ledger and (name_col or id_col):
        if has_truck and has_excav:
            mask = (result_df["标准设备名称（矿卡）"].astype(str).str.len() > 0) | \
                   (result_df["标准设备名称（挖机）"].astype(str).str.len() > 0)
        else:
            mask = result_df["标准设备名称"].astype(str).str.len() > 0
    elif oil_ledger and oil_col:
        mask = result_df[OIL_RESULT_KEY].astype(str).str.len() > 0
    else:
        mask = pd.Series(False, index=result_df.index)

    return result_df, result_df[mask].copy(), result_df[~mask].copy(), matched_count


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------

def export_to_excel(
    sheets: dict[str, pd.DataFrame], output_path: str,
    sheet_name: str = "全部", cancel_event: Optional[threading.Event] = None,
    progress_cb: Optional[ProgressCallback] = None,
    delete_on_cancel: bool = True,
) -> bool:
    """Write concatenated *sheets* to *output_path*.  Returns True on success."""
    import xlsxwriter
    cancelled = cancel_event or threading.Event()
    if not sheets:
        return False

    df_to_export = pd.concat(sheets.values(), ignore_index=True)
    workbook = xlsxwriter.Workbook(output_path)
    try:
        ws = workbook.add_worksheet(sheet_name)
        date_fmt = workbook.add_format({"num_format": "yyyy-mm-dd"})
        headers = list(df_to_export.columns)
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        date_cols: set[str] = set()
        for col in headers:
            if pd.api.types.is_datetime64_any_dtype(df_to_export[col]):
                date_cols.add(col)
            else:
                sample = df_to_export[col].dropna().head(10)
                if not sample.empty and sample.apply(
                    lambda v: isinstance(v, (datetime.date, datetime.datetime))
                ).any():
                    date_cols.add(col)

        total_rows = len(df_to_export)
        for i in range(0, total_rows, EXPORT_BATCH):
            if cancelled.is_set():
                logger.info("导出已取消")
                break
            batch = df_to_export.iloc[i:i + EXPORT_BATCH]
            for row_idx, (_, row) in enumerate(batch.iterrows(), start=i + 1):
                for col_idx, col_name in enumerate(headers):
                    value = row[col_name]
                    if pd.isna(value):
                        ws.write(row_idx, col_idx, "")
                    elif col_name in date_cols:
                        try:
                            if isinstance(value, pd.Timestamp):
                                ws.write_datetime(row_idx, col_idx, value.to_pydatetime(), date_fmt)
                            elif isinstance(value, (datetime.date, datetime.datetime)):
                                ws.write_datetime(row_idx, col_idx, value, date_fmt)
                            else:
                                ws.write(row_idx, col_idx, value)
                        except Exception:
                            ws.write(row_idx, col_idx, str(value))
                    else:
                        ws.write(row_idx, col_idx, value)
            if progress_cb and total_rows > 0:
                processed = min(i + EXPORT_BATCH, total_rows)
                progress_cb(processed / total_rows, f"正在导出第 {processed}/{total_rows} 行...")
    finally:
        workbook.close()

    if cancelled.is_set() and delete_on_cancel:
        try:
            Path(output_path).unlink(missing_ok=True)
            logger.info("已删除不完整的导出文件")
        except OSError:
            pass
        return False
    return True


# ---------------------------------------------------------------------------
# 视图 / 状态辅助
# ---------------------------------------------------------------------------

def get_view_df(state: MatchState) -> Optional[pd.DataFrame]:
    """Return the DataFrame for the current view mode."""
    if state.view_mode == "matched":
        return state.matched_sheets.get(state.current_sheet)
    if state.view_mode == "unmatched":
        return state.unmatched_sheets.get(state.current_sheet)
    return state.filtered_df


def get_current_df(state: MatchState) -> Optional[pd.DataFrame]:
    """Return the best available DataFrame for the current sheet."""
    name = state.current_sheet
    if not name:
        return None
    if name in state.matched_all_sheets:
        return state.matched_all_sheets[name]
    return state.all_sheets.get(name)


def apply_sort(state: MatchState) -> None:
    """Apply sort settings and update ``state.filtered_df``."""
    df = get_current_df(state)
    if df is None:
        state.filtered_df = None
        return
    result = df.copy()
    col = state.sort_column
    if col and col in result.columns:
        try:
            result = result.sort_values(by=col, ascending=state.sort_ascending, kind="stable")
        except Exception:
            logger.debug("排序失败: col=%s", col)
    state.filtered_df = result


def build_match_count_text(matched: dict, unmatched: dict, sheet: str) -> str:
    """Return ``'已匹配: N  |  未匹配: M'`` or empty string."""
    m, u = matched.get(sheet), unmatched.get(sheet)
    if m is not None and u is not None:
        return f"已匹配: {len(m)}  |  未匹配: {len(u)}"
    return ""
