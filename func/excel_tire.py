"""轮胎寿命统计表处理器。

轮胎寿命源表通常把多个安装/拆卸周期横向展开在同一行。本模块按
“开始日期”到下一个“开始日期”（或“寿命合计”）的范围识别周期，
将每个周期规范化为一行，并复用项目统一的异常检测和 Excel 输出模块。
"""
from __future__ import annotations

import datetime as dt
import logging
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.datetime import from_excel

from func.anomaly import detect_and_filter
from func.excel_utils import get_hidden_indices, load_workbook_safely
from func.number_utils import decimal_divide, decimal_multiply, decimal_subtract

logger = logging.getLogger(__name__)


TIRE_OUTPUT_SHEET = "轮胎安装拆卸明细"
TIRE_OUTPUT_COLUMNS = [
    "胎号",
    "品牌",
    "型号",
    "当前状态",
    "安装车辆",
    "安装位置",
    "安装日期",
    "安装时使用时间",
    "拆卸日期",
    "拆卸时使用时间",
    "安装时公里数",
    "拆卸时公里数",
    "寿命（时间）",
    "寿命（里程）",
    "备注",
    "周期状态",
    "异常原因",
    "安装次数",
    "标准寿命",
    "磨损程度",
    "源表",
    "源行号",
]

# 用户确认的去重组合。去重完成后必须重新计算安装次数及派生寿命字段。
DEDUP_FIELDS = [
    "胎号",
    "品牌",
    "型号",
    "安装车辆",
    "安装位置",
    "安装日期",
    "安装时使用时间",
    "拆卸日期",
    "拆卸时使用时间",
    "安装时公里数",
    "拆卸时公里数",
]


def normalize_header(value: object) -> str:
    """规范化表头，兼容换行、空格、中英文括号及英文表头。"""
    if value is None:
        return ""
    text = str(value).replace("\n", "").replace("\r", "")
    return re.sub(r"[\s_/\\:：()（）]+", "", text).lower()


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text or None


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def is_error(value: object) -> bool:
    return isinstance(value, str) and value.strip().startswith("#")


def parse_number(value: object) -> float | int | None:
    """解析带逗号、文本或 Excel 数值的读数，避免把 8,000 当成字符串。"""
    if is_blank(value) or is_error(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return int(value) if isinstance(value, float) and value.is_integer() else value
    text = str(value).strip().replace(",", "")
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def as_iso_date(value: dt.date | dt.datetime) -> str:
    if isinstance(value, dt.datetime):
        value = value.date()
    return value.isoformat()


def parse_low_quality_date_text(text: str) -> tuple[str | None, str | None]:
    """识别常见的人工录入日期错误并返回修正说明。

    支持例如：
    2025.020.1、2024.04.013、2024.05..14、
    2026..03.20 和 20230.08.15。
    """
    match = re.fullmatch(r"(\d{4,})\D+(\d{1,3})\D*(\d{1,3})\D*", text)
    if not match:
        return None, None

    raw_year, raw_month, raw_day = match.groups()
    year_text = raw_year[:4] if len(raw_year) > 4 else raw_year
    try:
        year = int(year_text)
    except ValueError:
        return None, None
    if not 1900 <= year <= 2099:
        return None, None

    if len(raw_month) <= 2:
        month = int(raw_month)
    else:
        # 020 这类月份优先按去除前导零后的值处理；仍不合法时，
        # 再尝试前两位/后两位，覆盖 Excel 人工录入中的多余数字。
        stripped_month = raw_month.lstrip("0") or "0"
        month = int(stripped_month)
        if not 1 <= month <= 12:
            first_two = int(raw_month[:2])
            last_two = int(raw_month[-2:])
            if 1 <= first_two <= 12:
                month = first_two
            elif 1 <= last_two <= 12:
                month = last_two

    day = int(raw_day.lstrip("0") or "0")
    try:
        parsed = dt.date(year, month, day)
    except ValueError:
        return None, None

    standard_shape = bool(re.fullmatch(r"\d{4}([./-])\d{2}\1\d{2}", text))
    low_quality = (
        len(raw_year) != 4
        or len(raw_month) != 2
        or len(raw_day) != 2
        or not standard_shape
    )
    if not low_quality:
        return parsed.isoformat(), None
    return parsed.isoformat(), f"{text}→{parsed.isoformat()}"


def parse_date(
    value: object,
    epoch,
) -> tuple[str | None, str | None, str | None, str | None]:
    """统一识别日期，月份和年份精度统一落到当月/当年 1 日。"""
    if is_blank(value):
        return None, None, None, None
    if is_error(value):
        return None, None, f"日期错误值 {value}", None
    if isinstance(value, dt.datetime):
        return as_iso_date(value), "日", None, None
    if isinstance(value, dt.date):
        return as_iso_date(value), "日", None, None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        # 2025.03 这类数值是年-月，而不是 Excel 日期序列号。
        if 1900 <= number < 10000:
            year = int(number)
            fraction = decimal_subtract(number, year)
            month = round(decimal_multiply(fraction, 100)) if fraction else 1
            if 1 <= month <= 12:
                return (
                    f"{year:04d}-{month:02d}-01",
                    "月" if fraction else "年",
                    None,
                    None,
                )
            return None, None, f"无法识别日期 {value}", None
        if 1 <= number <= 100000:
            try:
                converted = from_excel(number, epoch=epoch)
                return as_iso_date(converted), "日", None, None
            except Exception:
                return None, None, f"无法识别日期 {value}", None
        return None, None, f"无法识别日期 {value}", None

    text = str(value).strip().lstrip("'")
    text = re.sub(r"\s+", "", text)
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    low_quality_date, correction = parse_low_quality_date_text(text)
    if low_quality_date:
        return low_quality_date, "日", None, correction

    text = text.replace("/", "-").replace(".", "-")
    match = re.fullmatch(r"(\d{4})-(\d{1,2})(?:-(\d{1,2}))?", text)
    if match:
        year, month = int(match.group(1)), int(match.group(2))
        day = int(match.group(3)) if match.group(3) else 1
        precision = "日" if match.group(3) else "月"
        try:
            return dt.date(year, month, day).isoformat(), precision, None, None
        except ValueError:
            return None, None, f"无法识别日期 {value}", None
    if re.fullmatch(r"\d{4}", text):
        year = int(text)
        try:
            return dt.date(year, 1, 1).isoformat(), "年", None, None
        except ValueError:
            pass
    return None, None, f"无法识别日期 {value}", None


def role_predicate(role: str, normalized: str) -> bool:
    if role == "vehicle":
        return "所属车辆" in normalized or "ownedvehicle" in normalized
    if role == "position":
        return "安装位置" in normalized or "installationlocation" in normalized
    if role == "start_date":
        return "开始日期" in normalized or "startingdate" in normalized
    if role == "start_time":
        return "开始时间" in normalized or "startingenginetime" in normalized
    if role == "start_km":
        return "startingmachinekm" in normalized or (
            "开始" in normalized and "公里" in normalized
        )
    if role == "end_date":
        return "结束日期" in normalized or "enddate" in normalized
    if role == "end_time":
        return (
            "结束时间" in normalized
            or "endenginetime" in normalized
            or normalized.endswith("endtime")
        )
    if role == "end_km":
        return "endmachin" in normalized or (
            "结束" in normalized and "公里" in normalized
        )
    return False


def named_header_predicate(role: str, normalized: str) -> bool:
    if role == "serial":
        return "胎号" in normalized or "tireserialnumber" in normalized
    if role == "brand":
        return "品牌" in normalized or normalized == "brand"
    if role == "model":
        return "型号" in normalized or "modelspecification" in normalized
    if role == "standard_life":
        return (
            (
                "寿命" in normalized
                and any(token in normalized for token in ("norom", "norm", "标准"))
            )
            or "standardlife" in normalized
        )
    if role == "status":
        return "当前状态" in normalized or normalized == "status"
    if role == "desc":
        return normalized in {"desc", "description", "备注"} or "描述" in normalized
    return False


def _visible_headers(ws, row: int, hidden_col_indices: set[int]) -> list[str]:
    return [
        ""
        if col in hidden_col_indices
        else normalize_header(ws.cell(row, col).value)
        for col in range(1, ws.max_column + 1)
    ]


def find_header_row(
    ws,
    hidden_rows: set[int] | None = None,
    hidden_col_indices: set[int] | None = None,
) -> int:
    hidden_rows = hidden_rows or set()
    hidden_col_indices = hidden_col_indices or set()
    for row in range(1, min(ws.max_row, 20) + 1):
        if row in hidden_rows:
            continue
        normalized = _visible_headers(ws, row, hidden_col_indices)
        has_serial = any("胎号" in value for value in normalized)
        starts = sum(
            "开始日期" in value or "startingdate" in value
            for value in normalized
        )
        has_total = any(
            "寿命合计" in value or "totallife" in value for value in normalized
        )
        if has_serial and starts and has_total:
            return row
    raise ValueError(f"未找到表头行: {ws.title}")


def build_periods(
    ws,
    header_row: int,
    hidden_col_indices: set[int] | None = None,
) -> tuple[list[dict], int]:
    hidden_col_indices = hidden_col_indices or set()
    headers = _visible_headers(ws, header_row, hidden_col_indices)
    starts = [
        col
        for col, value in enumerate(headers, start=1)
        if "开始日期" in value or "startingdate" in value
    ]
    summary_candidates = [
        col
        for col, value in enumerate(headers, start=1)
        if "寿命合计" in value or "totallife" in value
    ]
    if not starts or not summary_candidates:
        raise ValueError(f"未找到安装周期边界: {ws.title}")

    summary_col = min(summary_candidates)
    starts = [col for col in starts if col < summary_col]
    if not starts:
        raise ValueError(f"寿命合计之前没有开始日期: {ws.title}")

    vehicles = [
        col
        for col, value in enumerate(headers, start=1)
        if role_predicate("vehicle", value)
    ]
    positions = [
        col
        for col, value in enumerate(headers, start=1)
        if role_predicate("position", value)
    ]

    blocks: list[dict] = []
    for start_date_col in starts:
        vehicle_candidates = [col for col in vehicles if col <= start_date_col]
        vehicle_col = max(vehicle_candidates) if vehicle_candidates else start_date_col
        position_candidates = [
            col
            for col in positions
            if vehicle_col <= col < start_date_col + 2
        ]
        position_col = max(position_candidates) if position_candidates else None
        blocks.append(
            {
                "start": vehicle_col,
                "start_date": start_date_col,
                "position": position_col,
            }
        )

    for index, block in enumerate(blocks):
        next_start = (
            blocks[index + 1]["start"]
            if index + 1 < len(blocks)
            else summary_col
        )
        block["end"] = next_start - 1
        role_cols: dict[str, int] = {}
        for role in (
            "vehicle",
            "position",
            "start_date",
            "start_time",
            "start_km",
            "end_date",
            "end_time",
            "end_km",
        ):
            candidates = [
                col
                for col in range(block["start"], block["end"] + 1)
                if role_predicate(role, headers[col - 1])
            ]
            if candidates:
                role_cols[role] = candidates[0]
        block["cols"] = role_cols
        block["header_range"] = (
            f"{get_column_letter(block['start'])}{header_row}:"
            f"{get_column_letter(block['end'])}{header_row}"
        )
    return blocks, summary_col


def find_named_column(
    ws,
    role: str,
    headers: list[str],
    *,
    required: bool = True,
) -> int | None:
    candidates = [
        col
        for col, value in enumerate(headers, start=1)
        if named_header_predicate(role, value)
    ]
    if not candidates:
        if required:
            raise ValueError(f"未找到字段表头 {role}: {ws.title}")
        return None
    return candidates[0]


def read_value(
    wsf,
    wsv,
    row: int,
    col: int | None,
) -> tuple[object, str | None]:
    if col is None:
        return None, None
    formula_or_value = wsf.cell(row, col).value
    displayed = wsv.cell(row, col).value
    if is_error(displayed):
        return None, f"{get_column_letter(col)}{row}={displayed}"
    if (
        isinstance(formula_or_value, str)
        and formula_or_value.startswith("=")
        and displayed is None
    ):
        return None, f"{get_column_letter(col)}{row}=公式未计算"
    return displayed, None


def combine_texts(values: list[object]) -> str | None:
    cleaned: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.add(text)
            cleaned.append(text)
    return "；".join(cleaned) if cleaned else None


def deduplicate_rows(rows: list[dict]) -> tuple[list[dict], dict]:
    """按用户指定的 11 字段组合去重，保留首次出现的记录。"""
    seen: set[tuple] = set()
    unique_rows: list[dict] = []
    removed = 0
    for row in rows:
        key = tuple(row.get(field) for field in DEDUP_FIELDS)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows, {
        "fields": DEDUP_FIELDS,
        "before": len(rows),
        "after": len(unique_rows),
        "removed": removed,
    }


def recalculate_derived_fields(rows: list[dict]) -> None:
    """去重后重算寿命、磨损程度和同一胎号的安装次数。"""
    for row in rows:
        period_status = row.get("周期状态")
        if period_status == "运行中":
            # 运行中没有可靠的拆卸读数，相关字段保持为空。
            row["拆卸时使用时间"] = None
            row["拆卸时公里数"] = None
            row["寿命（时间）"] = None
            row["寿命（里程）"] = None
            row["磨损程度"] = None
            continue

        if period_status == "待核对":
            row["寿命（时间）"] = None
            row["寿命（里程）"] = None
            row["磨损程度"] = None
            continue

        start_time = parse_number(row.get("安装时使用时间"))
        end_time = parse_number(row.get("拆卸时使用时间"))
        start_km = parse_number(row.get("安装时公里数"))
        end_km = parse_number(row.get("拆卸时公里数"))

        row["寿命（时间）"] = (
            decimal_subtract(end_time, start_time)
            if start_time is not None
            and end_time is not None
            and end_time >= start_time
            else None
        )
        row["寿命（里程）"] = (
            decimal_subtract(end_km, start_km)
            if start_km is not None
            and end_km is not None
            and end_km >= start_km
            else None
        )

        standard_life = parse_number(row.get("标准寿命"))
        life_time = row.get("寿命（时间）")
        row["磨损程度"] = (
            decimal_divide(life_time, standard_life)
            if life_time is not None
            and standard_life is not None
            and standard_life > 0
            else None
        )

    grouped_indexes: defaultdict[str, list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        serial = clean_text(row.get("胎号"))
        if serial:
            grouped_indexes[serial].append(index)
        else:
            row["安装次数"] = None

    for indexes in grouped_indexes.values():
        ordered_indexes = sorted(
            indexes,
            key=lambda index: (
                rows[index].get("安装日期") is None,
                rows[index].get("安装日期") or "",
                index,
            ),
        )
        for count, index in enumerate(ordered_indexes, start=1):
            rows[index]["安装次数"] = count


def summarize_rows(rows: list[dict]) -> dict:
    stats: Counter = Counter()
    for row in rows:
        status = row.get("周期状态")
        if status:
            stats[status] += 1
        stats["异常记录"] += bool(row.get("异常原因"))
    return dict(stats)


def parse_sheet(
    ws_formula,
    ws_values,
    epoch,
    *,
    hidden_rows: set[int] | None = None,
    hidden_col_indices: set[int] | None = None,
) -> dict:
    """解析一个可识别的轮胎寿命 sheet。"""
    hidden_rows = hidden_rows or set()
    hidden_col_indices = hidden_col_indices or set()
    header_row = find_header_row(ws_formula, hidden_rows, hidden_col_indices)
    periods, summary_col = build_periods(
        ws_formula, header_row, hidden_col_indices
    )
    headers = _visible_headers(ws_formula, header_row, hidden_col_indices)
    columns = {
        "serial": find_named_column(ws_formula, "serial", headers),
        "brand": find_named_column(ws_formula, "brand", headers, required=False),
        "model": find_named_column(ws_formula, "model", headers, required=False),
        "standard_life": find_named_column(
            ws_formula, "standard_life", headers, required=False
        ),
        "status": find_named_column(ws_formula, "status", headers, required=False),
        "desc": find_named_column(ws_formula, "desc", headers, required=False),
    }

    serials_by_row: dict[int, str] = {}
    serial_errors: dict[int, str] = {}
    row_has_core: dict[int, bool] = {}
    data_rows = [
        row
        for row in range(header_row + 1, ws_formula.max_row + 1)
        if row not in hidden_rows
    ]

    for row in data_rows:
        serial_value, serial_error = read_value(
            ws_formula, ws_values, row, columns["serial"]
        )
        serial = clean_text(serial_value)
        if serial:
            serials_by_row[row] = serial
        if serial_error:
            serial_errors[row] = serial_error

        period_has_data = False
        for block in periods:
            for role in (
                "vehicle",
                "position",
                "start_date",
                "start_time",
                "start_km",
                "end_date",
                "end_time",
                "end_km",
            ):
                col = block["cols"].get(role)
                if col:
                    value, _ = read_value(ws_formula, ws_values, row, col)
                    if not is_blank(value):
                        period_has_data = True
                        break
            if period_has_data:
                break
        row_has_core[row] = bool(serial or period_has_data)

    serial_counts = Counter(serials_by_row.values())
    rows: list[dict] = []
    stats: Counter = Counter()

    for row in data_rows:
        if not row_has_core.get(row):
            continue
        serial = serials_by_row.get(row)
        base_errors: list[str] = []
        if serial_errors.get(row):
            base_errors.append(f"胎号字段错误({serial_errors[row]})")
        if not serial:
            base_errors.append("胎号为空或公式错误")
        elif serial_counts[serial] > 1:
            base_errors.append("胎号重复")

        brand_value, _ = read_value(
            ws_formula, ws_values, row, columns["brand"]
        )
        model_value, _ = read_value(
            ws_formula, ws_values, row, columns["model"]
        )
        standard_life_value, standard_life_error = read_value(
            ws_formula, ws_values, row, columns["standard_life"]
        )
        standard_life = parse_number(standard_life_value)
        status_value, _ = read_value(
            ws_formula, ws_values, row, columns["status"]
        )
        desc_value, _ = read_value(
            ws_formula, ws_values, row, columns["desc"]
        )

        comments: list[str] = []
        for col in range(1, ws_formula.max_column + 1):
            if col in hidden_col_indices:
                continue
            comment = ws_formula.cell(row, col).comment
            if comment and comment.text:
                comments.append(
                    f"批注({get_column_letter(col)}{row})：{comment.text.strip()}"
                )
        remark = combine_texts([clean_text(desc_value), *comments])

        emitted_order = 0
        for block in periods:
            raw: dict[str, object] = {}
            raw_errors: dict[str, str | None] = {}
            for role, col in block["cols"].items():
                raw[role], raw_errors[role] = read_value(
                    ws_formula, ws_values, row, col
                )
            if not any(not is_blank(raw.get(role)) for role in raw):
                continue
            emitted_order += 1

            exceptions = list(base_errors)
            for role, error in raw_errors.items():
                if error:
                    exceptions.append(f"{role}字段错误({error})")

            install_date, _, install_date_error, install_date_correction = parse_date(
                raw.get("start_date"), epoch
            )
            remove_date, _, remove_date_error, remove_date_correction = parse_date(
                raw.get("end_date"), epoch
            )
            if install_date_error:
                exceptions.append(f"安装日期{install_date_error}")
            if install_date_correction:
                exceptions.append(
                    f"安装日期格式已自动修正：{install_date_correction}"
                )
            if remove_date_error:
                exceptions.append(f"拆卸日期{remove_date_error}")
            if remove_date_correction:
                exceptions.append(
                    f"拆卸日期格式已自动修正：{remove_date_correction}"
                )

            start_time = parse_number(raw.get("start_time"))
            start_km = parse_number(raw.get("start_km"))
            end_time = parse_number(raw.get("end_time"))
            end_km = parse_number(raw.get("end_km"))
            if install_date is None:
                exceptions.append("安装日期缺失")
            if start_time is None:
                exceptions.append("安装时使用时间缺失")
            if start_km is None:
                exceptions.append("安装时公里数缺失")

            if remove_date is None and remove_date_error is None:
                period_status = "运行中"
                if end_time is not None or end_km is not None:
                    exceptions.append("无拆卸日期，拆卸读数未采用")
                output_end_time = None
                output_end_km = None
                life_time = None
                life_km = None
            elif remove_date_error:
                period_status = "待核对"
                output_end_time = end_time
                output_end_km = end_km
                life_time = None
                life_km = None
            else:
                period_status = "已结束"
                output_end_time = end_time
                output_end_km = end_km
                if end_time is None:
                    exceptions.append("拆卸时使用时间缺失")
                if end_km is None:
                    exceptions.append("拆卸时公里数缺失")
                life_time = None
                life_km = None
                if start_time is not None and end_time is not None:
                    candidate = decimal_subtract(end_time, start_time)
                    if candidate >= 0:
                        life_time = candidate
                    else:
                        exceptions.append("寿命时间为负，已置空")
                if start_km is not None and end_km is not None:
                    candidate = decimal_subtract(end_km, start_km)
                    if candidate >= 0:
                        life_km = candidate
                    else:
                        exceptions.append("寿命里程为负，已置空")

            wear = None
            if life_time is not None:
                if standard_life is None or standard_life <= 0:
                    exceptions.append("标准寿命缺失或为0，磨损程度未计算")
                else:
                    wear = decimal_divide(life_time, standard_life)
            if standard_life_error:
                exceptions.append(f"标准寿命字段错误({standard_life_error})")

            position = parse_number(raw.get("position"))
            values = {
                "胎号": serial,
                "品牌": clean_text(brand_value),
                "型号": clean_text(model_value),
                "当前状态": clean_text(status_value),
                "安装车辆": clean_text(raw.get("vehicle")),
                "安装位置": position
                if position is not None
                else clean_text(raw.get("position")),
                "安装日期": install_date,
                "安装时使用时间": start_time,
                "拆卸日期": remove_date,
                "拆卸时使用时间": output_end_time,
                "安装时公里数": start_km,
                "拆卸时公里数": output_end_km,
                "寿命（时间）": life_time,
                "寿命（里程）": life_km,
                "备注": remark,
                "周期状态": period_status,
                "异常原因": combine_texts(exceptions),
                "安装次数": emitted_order,
                "标准寿命": standard_life,
                "磨损程度": wear,
                "源表": ws_formula.title,
                "源行号": row,
            }
            rows.append(values)
            stats[period_status] += 1
            stats["异常记录"] += bool(exceptions)

    return {
        "sheet": ws_formula.title,
        "header_row": header_row,
        "summary_column": get_column_letter(summary_col),
        "periods": periods,
        "row_count": len(rows),
        "stats": dict(stats),
        "rows": rows,
    }


def parse_tire_workbook(
    file_path: str | Path,
    *,
    skip_hidden: bool = False,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
) -> tuple[list[dict], dict]:
    """解析全部 sheet；无法识别的 sheet 会被跳过并记录原因。"""
    source = Path(file_path)
    if not source.is_file():
        raise FileNotFoundError(f"轮胎源文件不存在: {source}")
    if source.suffix.lower() == ".xls":
        raise ValueError("轮胎模块需要 .xlsx 文件，暂不支持旧版 .xls 格式")

    if skip_hidden:
        skip_hidden_rows = True
        skip_hidden_cols = True

    wb_formula = load_workbook_safely(source, data_only=False, read_only=False)
    wb_values = load_workbook_safely(source, data_only=True, read_only=False)
    rows: list[dict] = []
    recognized_sheets: list[dict] = []
    skipped_sheets: list[dict] = []

    try:
        for ws_formula in wb_formula.worksheets:
            ws_values = wb_values[ws_formula.title]
            hidden_rows: set[int] = set()
            hidden_cols: set[str] = set()
            if skip_hidden_rows or skip_hidden_cols:
                hidden_rows, hidden_cols = get_hidden_indices(
                    str(source),
                    ws_formula.title,
                    _workbook=wb_formula,
                )
            hidden_col_indices = (
                {
                    column_index_from_string(letter)
                    for letter in hidden_cols
                    if isinstance(letter, str)
                }
                if skip_hidden_cols
                else set()
            )
            effective_hidden_rows = hidden_rows if skip_hidden_rows else set()
            try:
                parsed = parse_sheet(
                    ws_formula,
                    ws_values,
                    wb_values.epoch,
                    hidden_rows=effective_hidden_rows,
                    hidden_col_indices=hidden_col_indices,
                )
            except Exception as exc:
                skipped_sheets.append(
                    {"sheet": ws_formula.title, "reason": str(exc)}
                )
                logger.info(
                    "跳过无法识别的轮胎 sheet: %s (%s)",
                    ws_formula.title,
                    exc,
                )
                continue

            recognized_sheets.append(
                {
                    key: parsed[key]
                    for key in (
                        "sheet",
                        "header_row",
                        "summary_column",
                        "periods",
                        "row_count",
                        "stats",
                    )
                }
            )
            rows.extend(parsed["rows"])
    finally:
        wb_formula.close()
        wb_values.close()

    if not recognized_sheets:
        raise ValueError("未识别到轮胎寿命统计表 sheet")

    rows, deduplication = deduplicate_rows(rows)
    recalculate_derived_fields(rows)
    for sheet_info in recognized_sheets:
        sheet_rows = [
            row for row in rows if row.get("源表") == sheet_info["sheet"]
        ]
        sheet_info["row_count"] = len(sheet_rows)
        sheet_info["stats"] = summarize_rows(sheet_rows)

    metadata = {
        "source": str(source),
        "sheets": recognized_sheets,
        "skipped_sheets": skipped_sheets,
        "deduplication": deduplication,
        "row_count": len(rows),
        "stats": summarize_rows(rows),
    }
    return rows, metadata


def _rows_to_dataframe(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=TIRE_OUTPUT_COLUMNS)
    for column in ("安装日期", "拆卸日期"):
        if column in df:
            parsed = pd.to_datetime(df[column], errors="coerce")
            df[column] = parsed.dt.date
    return df


def process_tire_data(
    file_path: str | Path,
    output_file: str | Path | None = None,
    *,
    return_sheets: bool = False,
    skip_hidden: bool = False,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
    anomaly_config=None,
) -> str | dict[str, pd.DataFrame]:
    """处理轮胎寿命表并写出统一格式的 Excel。

    return_sheets=True 供批处理/测试场景获取 DataFrame；默认返回输出文件路径。
    异常检测在去重和派生字段重算之后执行，目标列为寿命（时间）和寿命（里程）。
    """
    source = Path(file_path)
    rows, metadata = parse_tire_workbook(
        source,
        skip_hidden=skip_hidden,
        skip_hidden_rows=skip_hidden_rows,
        skip_hidden_cols=skip_hidden_cols,
    )
    df = _rows_to_dataframe(rows)

    if anomaly_config is not None and anomaly_config.enabled:
        df, _ = detect_and_filter(
            df,
            "tire",
            config=anomaly_config,
            output_dir=str(source.parent),
        )

    logger.info(
        "轮胎寿命处理完成: 识别 sheet=%d, 跳过 sheet=%d, 输出行数=%d, 去重=%d",
        len(metadata["sheets"]),
        len(metadata["skipped_sheets"]),
        len(df),
        metadata["deduplication"]["removed"],
    )
    for skipped in metadata["skipped_sheets"]:
        logger.warning(
            "轮胎 sheet 已跳过: %s，原因：%s",
            skipped["sheet"],
            skipped["reason"],
        )

    sheets = {TIRE_OUTPUT_SHEET: df}
    if return_sheets:
        return sheets

    output_path = source.with_name("轮胎寿命统计.xlsx") if output_file is None else Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    from func.excel_formatter import write_formatted_excel

    # 日期列已在 _rows_to_dataframe 中转换为 date 对象，交给统一格式化器
    # 做列级 yyyy-mm-dd 格式设置；不启用 date_only 的字符串猜测，避免
    # 胎号等业务编码被 pandas 误判为日期。
    write_formatted_excel(str(output_path), sheets)
    return str(output_path)


__all__ = [
    "DEDUP_FIELDS",
    "TIRE_OUTPUT_COLUMNS",
    "TIRE_OUTPUT_SHEET",
    "deduplicate_rows",
    "parse_date",
    "parse_low_quality_date_text",
    "parse_tire_workbook",
    "process_tire_data",
    "recalculate_derived_fields",
]
