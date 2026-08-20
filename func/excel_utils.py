"""
Excel 处理共享工具函数

提供各处理器共用的 DataFrame 后处理逻辑：
- 日期列标准化（去时间、可选覆盖年份）
- 按日期+班次排序
- 工时报表 Day/Night 班次分割与清洗
- 最终 DataFrame 全列去重
- 隐藏行/列检测与过滤
"""

import logging
import re
import warnings

import pandas as pd

from func.string_utils import clean_equipment_name, clean_string

logger = logging.getLogger(__name__)


# openpyxl can read the cell values from some Excel extension objects, but it
# emits one UserWarning per extension while parsing them.  Those extensions
# are not used by the processors and are discarded only if a workbook is
# later saved by openpyxl.  Keep the suppression scoped to the load call so
# unrelated warnings remain visible to the caller.
_OPENPYXL_EXTENSION_WARNING = (
    r"^(Unknown|Conditional Formatting) extension is not supported "
    r"and will be removed$"
)


def load_workbook_safely(*args, **kwargs):
    """Load an Excel workbook without noisy unsupported-extension warnings.

    The warning is emitted by ``openpyxl.worksheet._reader`` for extension
    objects such as vendor conditional-formatting metadata.  Values and
    comments needed by the processors are still read normally.
    """
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=_OPENPYXL_EXTENSION_WARNING,
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        return load_workbook(*args, **kwargs)


# ---------------------------------------------------------------------------
# Filename sanitization (path-traversal prevention)
# ---------------------------------------------------------------------------


def sanitize_filename(name: str) -> str:
    """Remove path separators and ``..`` sequences from *name*.

    This prevents path-traversal attacks when user-supplied strings are
    embedded in file paths (e.g. the ``keyword`` argument of
    :func:`~func.excel_merger.merge_excel_files`).

    Specifically:
    - Strips ``/`` and ``\\`` (path separators on Unix and Windows).
    - Removes ``..`` sequences to prevent directory escalation.

    Args:
        name: The raw user-supplied string to sanitize.

    Returns:
        A sanitised string safe for use as part of a filename.
    """
    # Use a regex to replace sequences containing / or \ with nothing.
    # We do NOT want to remove dots used in legitimate names, only ".."
    sanitized = re.sub(r'[/\\]', '', name)
    # Remove ".." sequences (they may appear after separator removal too)
    sanitized = re.sub(r'\.{2,}', '', sanitized)
    return sanitized


# ---------------------------------------------------------------------------
# Hidden rows / columns detection
# ---------------------------------------------------------------------------

def _is_xls(file_path: str) -> bool:
    """Return ``True`` if *file_path* is an old-format ``.xls`` workbook."""
    return str(file_path).lower().endswith(".xls")


def get_hidden_indices(
    file_path: str,
    sheet_name: str | int = 0,
    _workbook=None,
) -> tuple[set[int], set[str]]:
    """Return hidden row numbers and column letters for a given sheet.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Sheet name (str) or index (int, default 0).
        _workbook: Optional pre-loaded openpyxl workbook. When provided the
            file is **not** re-read, which avoids repeated ``load_workbook``
            calls in loops that process many sheets.

    Returns:
        ``(hidden_rows, hidden_cols)`` where *hidden_rows* is a set of
        1-based row indices and *hidden_cols* is a set of column letters
        (e.g. ``{'B', 'F'}``).  Both may be empty.

    Note:
        Old-format ``.xls`` files (BIFF) are opened with ``xlrd`` which
        does not expose row/column dimension metadata.  For ``.xls``
        files this function always returns two empty sets — hidden
        row/column filtering is silently skipped.
    """
    if _is_xls(file_path):
        return set(), set()

    should_close = _workbook is None
    wb = _workbook if _workbook is not None else load_workbook_safely(
        file_path, read_only=False, data_only=True,
    )
    try:
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
        hidden_cols = {col for col, dim in ws.column_dimensions.items() if dim.hidden}
    finally:
        if should_close:
            wb.close()
    return hidden_rows, hidden_cols


def open_workbook(file_path: str):
    """Open an openpyxl workbook for reuse across multiple ``get_hidden_indices`` calls.

    For old-format ``.xls`` files, returns ``None`` — callers should
    treat a ``None`` return the same as "hidden-row detection unavailable".

    Usage::

        wb = open_workbook(path)
        try:
            h_rows, h_cols = get_hidden_indices(path, "Sheet1", _workbook=wb)
            ...
        finally:
            if wb is not None:
                wb.close()
    """
    if _is_xls(file_path):
        return None
    return load_workbook_safely(file_path, read_only=False, data_only=True)


def get_column_letter(col_idx: int) -> str:
    """Convert a 1-based column number to a letter (1→A, 27→AA, …)."""
    from openpyxl.utils import get_column_letter as _get_column_letter

    return _get_column_letter(col_idx)


def filter_hidden_from_df(
    df: pd.DataFrame,
    hidden_rows: set[int],
    hidden_cols: set[str],
    *,
    has_header: bool = False,
) -> pd.DataFrame:
    """Remove hidden rows and columns from a pandas DataFrame.

    *hidden_rows* uses **1-based** indices (matching openpyxl).
    The DataFrame's own 0-based positional index is used for filtering.

    *hidden_cols* is a set of column **letters** (e.g. ``{'B', 'D'}``).
    This function assumes the DataFrame was read with ``header=None`` so
    that positional column indices map directly to Excel columns.
    When *has_header* is ``True`` the header row was consumed by pandas
    and hidden-row indices must be offset by 1 (Excel row 2 → df index 0).

    Args:
        df: Input DataFrame (typically from ``pd.read_excel(…, header=None)``).
        hidden_rows: Set of 1-based Excel row numbers to drop.
        hidden_cols: Set of Excel column letters to drop.
        has_header: Set ``True`` when the DataFrame was read with a header
            row (the default ``pd.read_excel`` mode).  Defaults to ``False``.

    Returns:
        A new DataFrame with hidden rows/columns removed.
    """
    if not hidden_rows and not hidden_cols:
        return df

    # Build row mask (keep non-hidden rows).
    # When has_header=False (header=None): Excel row 1 → df index 0.
    # When has_header=True: pandas consumed the header row, so Excel row 2 → df index 0.
    header_offset = 1 if has_header else 0
    row_mask = pd.Series(True, index=df.index)
    if hidden_rows:
        for r in hidden_rows:
            idx = r - 1 - header_offset
            if 0 <= idx < len(df):
                row_mask.iat[idx] = False

    # Build column mask (keep non-hidden columns).
    if hidden_cols:
        keep_cols = [
            df.columns[i]
            for i in range(df.shape[1])
            if get_column_letter(i + 1) not in hidden_cols
        ]
    else:
        keep_cols = list(df.columns)

    result = df.loc[row_mask, keep_cols]
    # Re-index columns to maintain sequential integer positions
    result.columns = range(len(result.columns))

    # IMPORTANT: Do NOT reset the row index when has_header=False.
    # Preserving the original integer index lets all processors that use
    # .iloc[] (positional access) work correctly after hidden rows are
    # removed, while code that reads .index (e.g. fuel's
    # ``df[df.iloc[:,0] == 1].index[0]``) continues to return meaningful
    # original row numbers.
    # For has_header=True (merger path) we do reset because pandas already
    # consumed the header row, making index 0 correspond to Excel row 2;
    # a reset keeps the offset arithmetic consistent.
    if has_header:
        result = result.reset_index(drop=True)

    return result


def adjust_index_for_hidden(
    original_index: int,
    hidden_set: set[int],
    *,
    one_based: bool = False,
) -> int:
    """Translate an original positional index to its new position after hidden removal.

    After filtering hidden rows or columns, positional indices shift.
    This function maps the original index to the corresponding index in
    the filtered sequence.

    Args:
        original_index: The 0-based positional index in the unfiltered data.
        hidden_set: Set of indices to remove.  Interpretation depends on
            *one_based*: when ``False`` (default) the set contains 0-based
            indices; when ``True`` it contains 1-based indices (Excel rows).
        one_based: If ``True``, *hidden_set* uses 1-based numbering
            (e.g. Excel row numbers).  *original_index* is always 0-based.

    Returns:
        The adjusted 0-based index in the filtered sequence.
    """
    offset = 0
    for h in sorted(hidden_set):
        h0 = h - 1 if one_based else h
        if h0 <= original_index:
            offset += 1
        else:
            break
    return original_index - offset


def letters_to_col_indices(letters: set[str]) -> set[int]:
    """Convert a set of Excel column letters to 0-based positional indices.

    >>> letters_to_col_indices({'A', 'C'})
    {0, 2}
    """
    from openpyxl.utils import column_index_from_string

    return {column_index_from_string(letter) - 1 for letter in letters}


# All known shift patterns for text-based detection
_SHIFT_PATTERNS: dict[str, list[str]] = {
    "Day": ["白班", "өдөр", "day", "日班"],
    "Night": ["夜班", "шөнө", "night", "夜"],
}


def detect_shift(text: str) -> str | None:
    """Identify shift from text content. Returns 'Day', 'Night', or None."""
    if not text:
        return None
    lower = str(text).lower()
    for shift, keywords in _SHIFT_PATTERNS.items():
        if any(kw in lower for kw in keywords):
            return shift
    return None


# 模块类型到默认输出文件名的映射（不含 worktime，因其含动态年月）
MODULE_OUTPUT_FILES: dict[str, str] = {
    "fuel": "Fuel.xlsx",
    "electrical": "电力消耗统计.xlsx",
    "production": "合并产量.xlsx",
    "maint": "维修记录统计.xlsx",
    "maintenance": "维修记录统计.xlsx",
}


def get_output_filename(module_type: str, year: int = 2025, month: int = 1) -> str | None:
    """根据模块类型返回默认输出文件名。

    Args:
        module_type: 模块类型（fuel/electrical/production/worktime）。
        year: 工时模块的年份。
        month: 工时模块的月份。

    Returns:
        输出文件名，未知类型返回 None。
    """
    if module_type == "worktime":
        return f"{year}{month:02d}_工作效率表.xlsx"
    if module_type == "tire":
        return "轮胎寿命统计.xlsx"
    return MODULE_OUTPUT_FILES.get(module_type)


def resolve_shift(
    col_to_shift: dict[int, str],
    target_idx: int,
    max_lookahead: int = 3,
    num_cols: int | None = None,
) -> str | None:
    """在 col_to_shift 映射中查找 target_idx 对应的班次。

    查找策略：
    1. 精确匹配 target_idx
    2. 向后搜索 (target_idx+1 .. target_idx+max_lookahead)
    3. 向前搜索 (target_idx-1 .. 3)

    Args:
        col_to_shift: {列索引: "Day"/"Night"} 映射。
        target_idx: 要查找班次的列索引。
        max_lookahead: 向后搜索的最大列数，默认 3。
        num_cols: 总列数，用于限制向后搜索范围。None 时不限制。

    Returns:
        "Day"、"Night" 或 None（未找到）。
    """
    # 精确匹配
    if target_idx in col_to_shift:
        return col_to_shift[target_idx]

    # 向后搜索
    end = target_idx + max_lookahead + 1
    if num_cols is not None:
        end = min(end, num_cols)
    for search_idx in range(target_idx + 1, end):
        if search_idx in col_to_shift:
            return col_to_shift[search_idx]

    # 向前搜索
    for search_idx in range(target_idx - 1, 2, -1):
        if search_idx in col_to_shift:
            return col_to_shift[search_idx]

    return None


def strip_date_column(
    df: pd.DataFrame,
    date_column: str = "日期",
    target_year: int | None = None,
    date_format: str | None = None,
) -> pd.DataFrame:
    """将 DataFrame 的日期列标准化为 date 对象（去除时间部分）。

    Args:
        df: 待处理的 DataFrame（返回新对象，不修改原 df）。
        date_column: 日期列名。
        target_year: 若指定，覆盖所有日期的年份。
        date_format: pd.to_datetime 的 format 参数，None 时自动推断。

    Returns:
        处理后的新 DataFrame。
    """
    if date_column not in df.columns or df.empty:
        return df

    result = df.copy()
    result[date_column] = pd.to_datetime(result[date_column], format=date_format, errors="coerce")
    if target_year is not None:
        result[date_column] = result[date_column].apply(
            lambda d: d.replace(year=target_year) if pd.notna(d) else d
        )
    result[date_column] = result[date_column].dt.date
    return result


def sort_by_date_shift(
    df: pd.DataFrame,
    sort_columns: list[str] | None = None,
    kind: str = "stable",
) -> pd.DataFrame:
    """按日期和班次排序。

    Args:
        df: 待排序的 DataFrame（返回新对象，不修改原 df）。
        sort_columns: 排序列，默认 ["日期", "班次"]。
        kind: 排序算法，默认 "stable"。

    Returns:
        排序后的新 DataFrame。
    """
    if sort_columns is None:
        sort_columns = ["日期", "班次"]

    existing = [c for c in sort_columns if c in df.columns]
    if existing:
        return df.sort_values(by=existing, kind=kind)
    return df


def split_day_night_shifts(
    df_raw: pd.DataFrame,
    header_row_index: int = 1,
    data_start_index: int = 2,
    day_end_offset: int = -1,
) -> pd.DataFrame:
    """将工时报表按 Day/Night 班次分割。

    检测 header_row 中的有效列，然后在数据行中查找与 header 首列
    相同的行作为 Day/Night 分割点。分割点之前为 Day 数据，之后为 Night 数据。

    Args:
        df_raw: 原始 DataFrame（header=None 读入）。
        header_row_index: 表头行索引，默认 1。
        data_start_index: 数据起始行索引，默认 2。
        day_end_offset: Day 数据结束位置相对 split_idx 的偏移量。
            默认 -1 表示 ``df_raw.iloc[data_start:split_idx - 1]``（excel_worktime.py 行为）。
            设为 0 表示 ``df_raw.iloc[data_start:split_idx]``（excel_worktime_multifile.py 行为）。

    Returns:
        合并后的 DataFrame，包含 '班次' 列（'Day' 或 'Night'）。
    """
    header_row = df_raw.iloc[header_row_index]
    valid_mask = header_row.notna() & (header_row.apply(lambda x: clean_string(x)) != "")
    valid_cols = valid_mask[valid_mask].index.tolist()
    valid_headers = header_row[valid_cols].apply(clean_string).tolist()

    if not valid_cols:
        day_data = df_raw.iloc[data_start_index:].copy()
        day_data.columns = header_row
        day_data["班次"] = "Day"
        return day_data

    split_idx = -1
    for idx in range(data_start_index, len(df_raw)):
        current_row_vals = df_raw.iloc[idx][valid_cols].apply(clean_string).tolist()
        if current_row_vals[0] == valid_headers[0]:
            split_idx = idx
            break

    if split_idx == -1:
        day_data = df_raw.iloc[data_start_index:].copy()
        day_data.columns = header_row
        day_data["班次"] = "Day"
        return day_data
    else:
        day_end = split_idx + day_end_offset
        day_data = df_raw.iloc[data_start_index:day_end].copy()
        day_data.columns = header_row
        day_data["班次"] = "Day"
        night_data = df_raw.iloc[split_idx + 1 :].copy()
        night_data.columns = header_row
        night_data["班次"] = "Night"
        return pd.concat([day_data, night_data], axis=0, ignore_index=True)


def clean_split_dataframe(
    df: pd.DataFrame,
    skip_columns: list[str] | None = None,
    check_keyword: str = "Техникийн",
) -> pd.DataFrame:
    """清洗 Day/Night 分割后的 DataFrame。

    - 移除 NaN 列
    - 移除空列名列
    - 按关键字列去空行
    - 按非元数据列全空去行

    Args:
        df: 分割后的 DataFrame（返回新对象，不修改原 df）。
        skip_columns: 不参与全空检查的列，默认 ["日期", "班次"]。
        check_keyword: 用于定位检查列的关键字。

    Returns:
        清洗后的新 DataFrame。
    """
    if skip_columns is None:
        skip_columns = ["日期", "班次"]

    result = df.copy()

    # 列名统一转为 clean string（split_day_night_shifts 可能产生 float/int/NaN 列名）
    # clean_string 将 NaN → ""，后续会自动移除这些空列名
    result.columns = [clean_string(c) for c in result.columns]

    # 移除空列名列（包括原始 NaN 列名被转为 "" 的情况）
    if "" in result.columns:
        result = result.drop(columns=[""])

    # 注意：不在 per-sheet 阶段移除全 NaN 数据列，因为不同 sheet 的空列集合不同，
    # 过早移除会导致 pd.concat 时列顺序错乱。全 NaN 列的移除应延迟到 concat 之后。
    # 参见 drop_all_nan_columns()。

    # 按关键字列去空行
    if len(result.columns) > 1:
        check_idx = -1
        for idx, col in enumerate(result.columns):
            if check_keyword in col:
                check_idx = idx
                break
        if check_idx != -1:
            check_col = result.columns[check_idx]
            result = result.dropna(subset=[check_col])

    # 按非元数据列全空去行
    subset_cols = [c for c in result.columns if c not in skip_columns]
    result = result.dropna(how="all", subset=subset_cols)

    return result


def drop_all_nan_columns(
    df: pd.DataFrame,
    preserve_columns: set[str] | None = None,
) -> pd.DataFrame:
    """移除 DataFrame 中所有值均为 NaN 的列，可保留指定列。

    应在 pd.concat 合并多个 sheet 之后调用，而非在单个 sheet 处理阶段。
    这样可以避免因不同 sheet 的空列集合不同而导致 concat 时列顺序错乱。

    Args:
        df: 合并后的 DataFrame（返回新对象，不修改原 df）。
        preserve_columns: 即使全为 NaN 也要保留的已有列名集合。
            不会凭空创建 DataFrame 中不存在的列。

    Returns:
        移除全 NaN 列后的新 DataFrame。
    """
    if df.empty:
        return df
    preserve = preserve_columns or set()
    nan_cols = [
        c for c in df.columns
        if c not in preserve and df[c].isna().all()
    ]
    if nan_cols:
        return df.drop(columns=nan_cols)
    return df


def drop_empty_device_name(
    df: pd.DataFrame,
    device_column: str = "设备名称",
) -> pd.DataFrame:
    """移除设备名称为空（NaN 或纯空白）的行。

    应在 pd.concat + drop_all_nan_columns 之后、排序之前调用。

    当指定列不存在时直接返回原 DataFrame，不做任何过滤。

    Args:
        df: 待过滤的 DataFrame（返回新对象，不修改原 df）。
        device_column: 设备名称列名，默认 "设备名称"。

    Returns:
        过滤后的新 DataFrame。
    """
    if df.empty or device_column not in df.columns:
        return df
    mask = df[device_column].isna() | (df[device_column].astype(str).str.strip() == "")
    removed = mask.sum()
    if removed > 0:
        logger.info(f"过滤了 {removed} 行设备名称为空的数据")
        return df[~mask].reset_index(drop=True)
    return df


def normalize_equipment_name_column(
    df: pd.DataFrame,
    device_column: str = "设备名称",
) -> pd.DataFrame:
    """对指定的设备名称列进行设备文本标准化。

    只处理设备名称列，其他文本字段保持原样（仅由各自的清洗流程处理）。
    当列不存在时返回原 DataFrame，便于兼容未启用表头映射的工效表。
    """
    if df.empty or device_column not in df.columns:
        return df

    result = df.copy()
    result[device_column] = result[device_column].map(clean_equipment_name)
    return result


def dedup_dataframe(
    df: pd.DataFrame,
    label: str = "",
) -> pd.DataFrame:
    """对 DataFrame 进行全列去重，保留首次出现的行。

    在各处理器最终写出前调用，消除源数据中的完全重复行。

    Args:
        df: 待去重的 DataFrame。
        label: 可选标签，用于日志中标识数据来源。

    Returns:
        去重后的 DataFrame（新对象）。
    """
    if df.empty:
        return df
    before = len(df)
    df = df.drop_duplicates()
    removed = before - len(df)
    if removed > 0:
        tag = f"[{label}] " if label else ""
        logger.info("%s去重: %d → %d 行，移除 %d 条重复记录", tag, before, len(df), removed)
    return df


# ---------------------------------------------------------------------------
# Header mapping
# ---------------------------------------------------------------------------


def apply_header_mapping(
    df: pd.DataFrame,
    mapping_config: dict,
    fuzzy_threshold: int = 70,
    mode_filter: str | None = None,
) -> pd.DataFrame:
    """Rename DataFrame columns according to a mapping configuration.

    Supports two modes (controlled by ``mapping_config["mode"]`` and
    the ``mode_filter`` parameter):

    - ``"position"``: match columns by 1-based index in the **original**
      DataFrame (before ``clean_split_dataframe``).  The ``index`` field
      in each entry corresponds to the column's position in the raw
      Excel data.

    - ``"name"``: match columns by keyword search.  Each entry provides
      a ``keywords`` list.  A column matches if **any** keyword appears
      in the cleaned column name (OR logic, case-sensitive substring).
      The first matching entry wins; once a column is renamed it is
      excluded from subsequent entries.

      Example entry::

          {"keywords": ["设备种类", "төрөл"], "new": "设备种类"}

    Args:
        df: The source DataFrame (returned as a new object).
        mapping_config: A dict with ``mode``, ``fuzzy``, and ``entries`` keys.
            ``entries`` is a list of dicts with ``index``, ``keywords``, and
            ``new`` keys.
        fuzzy_threshold: Deprecated, kept for API compatibility.
        mode_filter: If set, only apply entries that match this mode.
            ``"position"`` — only position-based entries (index != None).
            ``"name"`` — only keyword-based entries (keywords/非空).
            ``None`` — apply all entries (default).

    Returns:
        A new DataFrame with matched columns renamed.
    """
    if not mapping_config or not mapping_config.get("entries"):
        return df

    entries = mapping_config["entries"]
    cols = list(df.columns)
    rename_map: dict[str, str] = {}

    # 按 entry 自身特征和 mode 分类
    # - 当 mode_filter 指定时，以 mode_filter 为准（有对应字段即收录）
    # - 否则检查 config 的 mode 字段
    # - 都没有时按 entry 自身特征分类（向后兼容：有 keywords 优先当 name）
    effective_mode = mode_filter or mapping_config.get("mode")
    pos_entries = []
    kw_entries = []
    for entry in entries:
        has_keywords = bool(entry.get("keywords"))
        has_index = entry.get("index") is not None
        if effective_mode == "position":
            # position 模式：只按 index 分类，忽略 keywords
            if has_index:
                pos_entries.append(entry)
        elif effective_mode == "name":
            # name 模式：只按 keywords 分类，忽略 index
            if has_keywords:
                kw_entries.append(entry)
        else:
            # 无明确 mode：按 entry 自身特征分类（向后兼容）
            if has_keywords:
                kw_entries.append(entry)
            elif has_index:
                pos_entries.append(entry)

    # 位置映射（应在 clean_split_dataframe 之前调用，index 对应原始列位置）
    for entry in pos_entries:
        idx = entry.get("index")
        new_name = clean_string(entry.get("new"))
        if idx is None or not new_name:
            continue
        try:
            idx = int(idx)
        except (TypeError, ValueError):
            continue
        # Config uses 1-based indices; convert to 0-based.
        if 1 <= idx <= len(cols):
            old_name = cols[idx - 1]
            rename_map[old_name] = new_name

    # 关键字映射（应在 clean_split_dataframe 之前调用，关键字匹配原始列名）
    if kw_entries:
        kw_map = _match_name_by_keyword(kw_entries, cols)
        rename_map.update(kw_map)

    if rename_map:
        pos_count = len(pos_entries) if not mode_filter or mode_filter == "position" else 0
        kw_count = len(kw_entries) if not mode_filter or mode_filter == "name" else 0
        mode_desc = mode_filter or ("position+name" if pos_count and kw_count else "position" if pos_count else "name")
        logger.info(
            "表头映射生效（模式: %s），重命名 %d 列: %s",
            mode_desc,
            len(rename_map),
            rename_map,
        )
    return df.rename(columns=rename_map)


def _match_name_by_keyword(
    entries: list[dict],
    cols: list[str],
) -> dict[str, str]:
    """Match mapping entries to DataFrame columns using keyword OR logic.

    Each entry provides a ``keywords`` list.  A column matches if **any**
    keyword is a substring of the cleaned column name.  First match
    wins — once a column is renamed it is excluded from later entries.

    Entry format::

        {"keywords": ["设备种类", "төрөл"], "new": "设备种类"}

    Returns:
        ``{original_col_name: new_name}`` rename mapping.
    """
    rename_map: dict[str, str] = {}

    # Prepare entries: (keywords_list, new_name)
    prepared: list[tuple[list[str], str]] = []
    for entry in entries:
        new_name = clean_string(entry.get("new"))
        if not new_name:
            continue
        raw_keywords = entry.get("keywords")
        if not raw_keywords or not isinstance(raw_keywords, list):
            continue
        keywords = [clean_string(kw) for kw in raw_keywords if clean_string(kw)]
        if keywords:
            prepared.append((keywords, new_name))

    # Build cleaned column list: (clean_name, original_name)
    col_pairs: list[tuple[str, str]] = [(clean_string(c), c) for c in cols]

    matched: set[str] = set()  # original column names already matched
    for keywords, new_name in prepared:
        for col_clean, col_orig in col_pairs:
            if col_orig in matched:
                continue
            if any(kw in col_clean for kw in keywords):
                rename_map[col_orig] = new_name
                matched.add(col_orig)
                break

    return rename_map


def strip_date_only_times(df: pd.DataFrame) -> pd.DataFrame:
    """对 datetime 列检测：若所有非空值的时间部分均为 00:00:00，
    则转换为 date 对象，避免 Excel 导出时出现多余的 ' 00:00:00'。

    Args:
        df: 待处理的 DataFrame（返回新对象，不修改原 df）。

    Returns:
        处理后的新 DataFrame。
    """
    import datetime as _dt

    midnight = _dt.time(0, 0, 0)
    result = df.copy()
    for col in result.columns:
        if not pd.api.types.is_datetime64_any_dtype(result[col]):
            continue
        times = result[col].dropna().dt.time
        if times.empty:
            continue
        if (times == midnight).all():
            result[col] = result[col].apply(
                lambda v: v.date() if pd.notna(v) else v
            )
    return result
