"""维修记录提取逻辑

从设备出勤统计表的单元格批注中提取维修记录，支持隐藏行/列过滤和公式解析。
"""
import ast as _ast
import glob
import operator
import os
import re
from datetime import date

import openpyxl

from func.logger import get_logger
from func.maintenance_utils import (
    detect_header_layout,
    parse_comment,
    parse_month_from_sheetname,
    parse_year_month_from_filename,
)

logger = get_logger(__name__)


# ── 安全公式解析 ──────────────────────────────────────────────

_AST_OPS = {
    _ast.Add: operator.add,
    _ast.Sub: operator.sub,
    _ast.Mult: operator.mul,
    _ast.Div: operator.truediv,
}


def _eval_ast_node(node):
    """递归求值 AST 节点（仅允许数值和四则运算）。"""
    if isinstance(node, _ast.Constant) and isinstance(node.value, (int, float)):
        return node.value
    if isinstance(node, _ast.UnaryOp) and isinstance(node.op, _ast.USub):
        return -_eval_ast_node(node.operand)
    if isinstance(node, _ast.BinOp) and type(node.op) in _AST_OPS:
        return _AST_OPS[type(node.op)](_eval_ast_node(node.left), _eval_ast_node(node.right))
    raise ValueError(f"不支持的表达式: {node}")


def _safe_eval_formula(value) -> int | None:
    """安全解析简单四则运算公式字符串为整数。

    支持 "=25+30"、"=100-20"、"=5*6"、"=(10+5)*2" 等四则运算公式。
    不使用 eval()，通过 ast 模块安全解析。

    Args:
        value: 单元格值（可能是公式字符串）。

    Returns:
        计算结果的整数值，失败返回 None。
    """
    s = str(value).strip()
    if not s.startswith("="):
        return None
    expr = s[1:]  # 去掉 "="
    if not re.match(r'^[\d\s+\-*/().]+$', expr):
        return None
    try:
        tree = _ast.parse(expr, mode='eval')
        result = _eval_ast_node(tree.body)
        return int(float(result))
    except (ZeroDivisionError, SyntaxError, ValueError, TypeError):
        return None


# ── 记录提取 ──────────────────────────────────────────────────

def extract_sheet_records(
    ws, year: int, month: int,
    *,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
    ws_values=None,
) -> list[dict]:
    """从单个工作表提取维修记录。

    Args:
        ws: openpyxl Worksheet（用于读取批注）。
        year: 年份。
        month: 月份。
        skip_hidden_rows: 跳过 Excel 中的隐藏行。
        skip_hidden_cols: 跳过 Excel 中的隐藏列对应的日期数据。
        ws_values: data_only=True 模式的 Worksheet（用于读取公式计算结果）。
                   为 None 时回退到 ws.cell.value。

    Returns:
        记录列表，每条包含 原始设备名称、原因、班次、维修内容、工时_分钟、日期。
    """
    reason_col, day_col_map = detect_header_layout(ws)
    records: list[dict] = []
    current_vehicle: str | None = None

    # 预计算隐藏行集合（1-based 行号）
    hidden_rows: set[int] = set()
    if skip_hidden_rows:
        hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}

    # 预计算隐藏列集合（日期号），用于过滤日期列
    hidden_cols: set[int] = set()
    if skip_hidden_cols:
        hidden_col_letters = {col for col, dim in ws.column_dimensions.items() if dim.hidden}
        for day_num, col in day_col_map.items():
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter in hidden_col_letters:
                hidden_cols.add(day_num)

    for row_num in range(2, ws.max_row + 1):
        if row_num in hidden_rows:
            continue

        cell_a = ws.cell(row=row_num, column=1)
        cell_b = ws.cell(row=row_num, column=reason_col)

        vehicle = cell_a.value
        reason_type = cell_b.value

        if vehicle:
            current_vehicle = str(vehicle).strip()
        if not reason_type:
            continue
        reason_type = str(reason_type).strip()

        for day_num, col in day_col_map.items():
            if day_num in hidden_cols:
                continue
            cell = ws.cell(row=row_num, column=col)

            has_comment = cell.comment is not None
            # 从 data_only 工作表读取数值（正确处理公式结果）
            if ws_values is not None:
                val_cell = ws_values.cell(row=row_num, column=col)
                raw_value = val_cell.value
            else:
                raw_value = cell.value
            has_value = raw_value is not None and str(raw_value).strip() != ""
            if not has_comment and not has_value:
                continue

            comment_text = cell.comment.text if cell.comment else ""
            minutes = raw_value
            if minutes is not None:
                try:
                    minutes = int(float(str(minutes)))
                except (ValueError, TypeError):
                    # 回退：尝试解析简单公式字符串（如 "=25+30"）
                    minutes = _safe_eval_formula(minutes)

            # 有批注但无数值时，工时记为 0
            if minutes is None:
                minutes = 0

            # 构造日期
            try:
                dt = date(year, month, day_num)
            except ValueError:
                continue  # 无效日期（如2月30日）

            entries = parse_comment(comment_text)
            if entries:
                # 多班次时平分工时
                per_shift_minutes = round(minutes / len(entries)) if minutes and len(entries) > 1 else minutes
                for shift, content in entries:
                    records.append({
                        "日期": dt,
                        "原始设备名称": current_vehicle or "",
                        "原因": reason_type,
                        "班次": shift,
                        "维修内容": content,
                        "工时_分钟": per_shift_minutes,
                    })
            elif minutes is not None:
                records.append({
                    "日期": dt,
                    "原始设备名称": current_vehicle or "",
                    "原因": reason_type,
                    "班次": "未标注",
                    "维修内容": "",
                    "工时_分钟": minutes,
                })

    return records


# ── 文件发现 ──────────────────────────────────────────────────

def discover_files(file_path: str, file_keywords: list[str]) -> list[str]:
    """发现待处理的 Excel 文件。

    单文件：直接返回 [file_path]。
    文件夹：遍历 .xlsx 文件，按关键字过滤，排除 ~$ 临时文件。

    Args:
        file_path: 文件或文件夹路径。
        file_keywords: 文件名关键字列表（任一命中即保留）。

    Returns:
        排序后的文件路径列表。
    """
    if os.path.isfile(file_path):
        return [file_path]

    files = []
    for f in sorted(glob.glob(os.path.join(file_path, "*.xlsx"))):
        basename = os.path.basename(f)
        if basename.startswith("~$"):
            continue
        if any(kw in basename for kw in file_keywords):
            files.append(f)
    return files


def extract_all_records(
    file_path: str,
    file_keywords: list[str],
    *,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
) -> list[dict]:
    """从文件或文件夹提取全部维修记录（去重）。

    Args:
        file_path: 文件或文件夹路径。
        file_keywords: 文件名关键字。
        skip_hidden_rows: 跳过隐藏行。
        skip_hidden_cols: 跳过隐藏列。

    Returns:
        合并后的记录列表。
    """
    files = discover_files(file_path, file_keywords)
    if not files:
        logger.warning("未找到匹配的文件 (关键字: %s)", file_keywords)
        return []

    all_records: list[dict] = []
    processed_months: set[tuple[int, int]] = set()

    for filepath in files:
        filename = os.path.basename(filepath)
        file_year, file_month = parse_year_month_from_filename(filename)
        logger.info("处理: %s (年=%s, 月=%s)", filename, file_year, file_month)

        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception as e:
            logger.error("无法打开文件 %s: %s", filepath, e)
            continue

        # 第二个 workbook 用于读取公式单元格的计算结果
        try:
            wb_values = openpyxl.load_workbook(filepath, data_only=True)
        except Exception:
            wb_values = None

        try:
            multi_sheet = len(wb.sheetnames) > 1

            for sheetname in wb.sheetnames:
                sheet_month = parse_month_from_sheetname(sheetname)
                year = file_year
                month = sheet_month if (multi_sheet and sheet_month) else (file_month or sheet_month)

                if not year or not month:
                    logger.debug("跳过 sheet '%s': 无法确定年月", sheetname)
                    continue

                key = (year, month)
                if key in processed_months:
                    logger.debug("跳过 sheet '%s' -> %d年%d月 (已处理)", sheetname, year, month)
                    continue

                ws = wb[sheetname]
                ws_values = wb_values[sheetname] if wb_values and sheetname in wb_values.sheetnames else None
                records = extract_sheet_records(
                    ws, year, month,
                    skip_hidden_rows=skip_hidden_rows,
                    skip_hidden_cols=skip_hidden_cols,
                    ws_values=ws_values,
                )
                all_records.extend(records)
                processed_months.add(key)
                logger.info("  sheet '%s' -> %d年%d月: %d 条记录", sheetname, year, month, len(records))
        finally:
            wb.close()
            if wb_values:
                wb_values.close()

    logger.info("共提取 %d 条维修记录 (%d 个月)", len(all_records), len(processed_months))
    return all_records
