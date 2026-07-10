"""维修记录处理主模块

从设备出勤统计表的单元格批注中提取维修记录，经台账匹配和故障分类后，
生成包含 10 个统计 sheet 的 Excel 报告。

支持单文件和文件夹批量处理，年月从文件名/表名自动解析。
"""
import argparse
import glob
import os
import re
from collections import Counter, defaultdict
from datetime import date

import openpyxl
import pandas as pd
import xlsxwriter

from func.logger import get_logger, setup_logging
from func.maintenance_utils import (
    detect_header_layout,
    extract_device_model,
    parse_comment,
    parse_date,
    parse_month_from_sheetname,
    parse_year_month_from_filename,
    preprocess_device_name,
)
from func.maintenance_classification import (
    classify,
    get_default_classifications,
    is_fault_record,
)
from func.string_utils import clean_string


# ── 公式解析辅助 ──────────────────────────────────────────────

def _try_eval_formula(value) -> int | None:
    """尝试解析简单公式字符串为整数。

    支持 "=25+30"、"=100-20"、"=5*6" 等四则运算公式。
    复杂公式或解析失败时返回 None。

    Args:
        value: 单元格值（可能是公式字符串）。

    Returns:
        计算结果的整数值，失败返回 None。
    """
    s = str(value).strip()
    if not s.startswith("="):
        return None
    expr = s[1:]  # 去掉 "="
    # 仅允许数字、运算符、空格、括号
    import re
    if not re.match(r'^[\d\s+\-*/().]+$', expr):
        return None
    try:
        result = eval(expr)  # noqa: S307 — 已严格校验输入
        return int(float(result))
    except (ZeroDivisionError, SyntaxError, ValueError, TypeError):
        return None

logger = get_logger(__name__)


# ── 样式常量 ──────────────────────────────────────────────────

_DATE_FMT = "yyyy-mm-dd"
_PCT_FMT = "0.00%"
_HOUR_FMT = "0.0"
_HEADER_BG = "#4472C4"
_HEADER_FG = "#FFFFFF"
_ALT_ROW_BG = "#F2F7FB"


# ── xlsxwriter 样式工厂 ───────────────────────────────────────

def _make_formats(wb: xlsxwriter.Workbook) -> dict:
    """创建统一的格式集，避免重复定义。"""
    return {
        "header": wb.add_format({
            "bold": True, "font_color": _HEADER_FG, "bg_color": _HEADER_BG,
            "border": 1, "align": "center", "valign": "vcenter",
            "text_wrap": True,
        }),
        "date": wb.add_format({"num_format": _DATE_FMT, "border": 1, "align": "center"}),
        "pct": wb.add_format({"num_format": _PCT_FMT, "border": 1, "align": "center"}),
        "hour": wb.add_format({"num_format": _HOUR_FMT, "border": 1, "align": "center"}),
        "int": wb.add_format({"border": 1, "align": "center"}),
        "text": wb.add_format({"border": 1, "text_wrap": True, "valign": "top"}),
        "text_center": wb.add_format({"border": 1, "align": "center", "valign": "vcenter"}),
        "bold_int": wb.add_format({"bold": True, "border": 1, "align": "center"}),
        "bold_text": wb.add_format({"bold": True, "border": 1}),
        "alt_row": wb.add_format({"border": 1, "bg_color": _ALT_ROW_BG}),
        "alt_date": wb.add_format({
            "num_format": _DATE_FMT, "border": 1, "align": "center",
            "bg_color": _ALT_ROW_BG,
        }),
        "alt_pct": wb.add_format({
            "num_format": _PCT_FMT, "border": 1, "align": "center",
            "bg_color": _ALT_ROW_BG,
        }),
        "alt_hour": wb.add_format({
            "num_format": _HOUR_FMT, "border": 1, "align": "center",
            "bg_color": _ALT_ROW_BG,
        }),
        "alt_int": wb.add_format({
            "border": 1, "align": "center", "bg_color": _ALT_ROW_BG,
        }),
        "alt_text": wb.add_format({
            "border": 1, "text_wrap": True, "valign": "top", "bg_color": _ALT_ROW_BG,
        }),
        "alt_text_center": wb.add_format({
            "border": 1, "align": "center", "valign": "vcenter",
            "bg_color": _ALT_ROW_BG,
        }),
    }


def _auto_col_width(headers: list[str], rows: list[tuple], min_w: int = 8, max_w: int = 50) -> list[int]:
    """计算自适应列宽（支持 CJK 宽字符）。"""
    def _width(text: str) -> int:
        w = 0
        for ch in str(text):
            cp = ord(ch)
            if 0x4E00 <= cp <= 0x9FFF or 0xFF00 <= cp <= 0xFFEF or 0x3000 <= cp <= 0x303F:
                w += 2
            else:
                w += 1
        return w

    widths = [_width(h) + 2 for h in headers]
    for row in rows[:500]:
        for i, val in enumerate(row):
            if val is None:
                continue
            cell_w = 12 if isinstance(val, (date, pd.Timestamp)) else _width(str(val)) + 2
            if i < len(widths):
                widths[i] = max(widths[i], cell_w)
    return [min(max(w, min_w), max_w) for w in widths]


def _write_sheet(
    wb: xlsxwriter.Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[tuple],
    fmts: dict,
    col_formats: list[str] | None = None,
    date_cols: set[int] | None = None,
    pct_cols: set[int] | None = None,
    hour_cols: set[int] | None = None,
    wrap_cols: set[int] | None = None,
):
    """通用 sheet 写入：表头 + 数据行 + 自适应列宽 + 冻结 + 筛选。"""
    ws = wb.add_worksheet(sheet_name)
    date_cols = date_cols or set()
    pct_cols = pct_cols or set()
    hour_cols = hour_cols or set()
    wrap_cols = wrap_cols or set()

    # 表头
    for col, h in enumerate(headers):
        ws.write(0, col, h, fmts["header"])

    # 数据行（交替底色）
    for row_idx, row in enumerate(rows):
        is_alt = row_idx % 2 == 1
        for col_idx, val in enumerate(row):
            if val is None:
                fmt = fmts["alt_text"] if is_alt else fmts["text"]
                ws.write(row_idx + 1, col_idx, "", fmt)
            elif col_idx in date_cols:
                fmt = fmts["alt_date"] if is_alt else fmts["date"]
                if isinstance(val, (date, pd.Timestamp)):
                    ws.write_datetime(row_idx + 1, col_idx, pd.Timestamp(val).to_pydatetime(), fmt)
                else:
                    ws.write(row_idx + 1, col_idx, str(val), fmt)
            elif col_idx in pct_cols:
                fmt = fmts["alt_pct"] if is_alt else fmts["pct"]
                ws.write_number(row_idx + 1, col_idx, float(val) if val else 0, fmt)
            elif col_idx in hour_cols:
                fmt = fmts["alt_hour"] if is_alt else fmts["hour"]
                ws.write_number(row_idx + 1, col_idx, float(val) if val else 0, fmt)
            elif col_idx in wrap_cols:
                fmt = fmts["alt_text"] if is_alt else fmts["text"]
                ws.write(row_idx + 1, col_idx, str(val), fmt)
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                fmt = fmts["alt_int"] if is_alt else fmts["int"]
                ws.write_number(row_idx + 1, col_idx, float(val), fmt)
            else:
                fmt = fmts["alt_text_center"] if is_alt else fmts["text_center"]
                ws.write(row_idx + 1, col_idx, str(val) if val is not None else "", fmt)

    # 列宽
    widths = _auto_col_width(headers, rows)
    for i, w in enumerate(widths):
        ws.set_column(i, i, w)

    # 冻结首行 + 自动筛选
    ws.freeze_panes(1, 0)
    if rows:
        last_col = len(headers) - 1
        ws.autofilter(0, 0, len(rows), last_col)


# ── 记录提取 ──────────────────────────────────────────────────

def extract_sheet_records(
    ws, year: int, month: int,
    *,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
    ws_values=None,
) -> list[dict]:
    """从单个工作表提取维修记录。

    Args:
        ws: openpyxl Worksheet（用于读取批注）。
        year: 年份。
        month: 月份。
        skip_hidden_rows: 跳过 Excel 中的隐藏行。
        skip_hidden_cols: 跳过 Excel 中的隐藏列对应的日期数据。
        ws_values: data_only=True 模式的 Worksheet（用于读取公式计算结果）。
                   为 None 时回退到 ws.cell.value。

    Returns:
        记录列表，每条包含 原始设备名称、原因、班次、维修内容、工时_分钟、日期。
    """
    reason_col, day_col_map = detect_header_layout(ws)
    records: list[dict] = []
    current_vehicle: str | None = None

    # 预计算隐藏行集合（1-based 行号）
    hidden_rows: set[int] = set()
    if skip_hidden_rows:
        hidden_rows = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}

    # 预计算隐藏列集合（列字母），用于过滤日期列
    hidden_cols: set[int] = set()
    if skip_hidden_cols:
        hidden_col_letters = {col for col, dim in ws.column_dimensions.items() if dim.hidden}
        for day_num, col in day_col_map.items():
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter in hidden_col_letters:
                hidden_cols.add(day_num)

    for row_num in range(2, ws.max_row + 1):
        if row_num in hidden_rows:
            continue

    for row_num in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row_num, column=1)
        cell_b = ws.cell(row=row_num, column=reason_col)

        vehicle = cell_a.value
        reason_type = cell_b.value

        if vehicle:
            current_vehicle = str(vehicle).strip()
        if not reason_type:
            continue
        reason_type = str(reason_type).strip()

        for day_num, col in day_col_map.items():
            if day_num in hidden_cols:
                continue
            cell = ws.cell(row=row_num, column=col)

            has_comment = cell.comment is not None
            # 从 data_only 工作表读取数值（正确处理公式结果）
            if ws_values is not None:
                val_cell = ws_values.cell(row=row_num, column=col)
                raw_value = val_cell.value
            else:
                raw_value = cell.value
            has_value = raw_value is not None and str(raw_value).strip() != ""
            if not has_comment and not has_value:
                continue

            comment_text = cell.comment.text if cell.comment else ""
            minutes = raw_value
            if minutes is not None:
                try:
                    minutes = int(float(str(minutes)))
                except (ValueError, TypeError):
                    # 回退：尝试解析简单公式字符串（如 "=25+30"）
                    minutes = _try_eval_formula(minutes)

            # 构造日期
            try:
                dt = date(year, month, day_num)
            except ValueError:
                continue  # 无效日期（如2月30日）

            entries = parse_comment(comment_text)
            if entries:
                for shift, content in entries:
                    records.append({
                        "日期": dt,
                        "原始设备名称": current_vehicle or "",
                        "原因": reason_type,
                        "班次": shift,
                        "维修内容": content,
                        "工时_分钟": minutes,
                    })
            elif minutes is not None:
                records.append({
                    "日期": dt,
                    "原始设备名称": current_vehicle or "",
                    "原因": reason_type,
                    "班次": "未标注",
                    "维修内容": "",
                    "工时_分钟": minutes,
                })

    return records


def _discover_files(file_path: str, file_keywords: list[str]) -> list[str]:
    """发现待处理的 Excel 文件。

    单文件：直接返回 [file_path]。
    文件夹：遍历 .xlsx 文件，按关键字过滤，排除 ~$ 临时文件。

    Args:
        file_path: 文件或文件夹路径。
        file_keywords: 文件名关键字列表（任一命中即保留）。

    Returns:
        排序后的文件路径列表。
    """
    if os.path.isfile(file_path):
        return [file_path]

    files = []
    for f in sorted(glob.glob(os.path.join(file_path, "*.xlsx"))):
        basename = os.path.basename(f)
        if basename.startswith("~$"):
            continue
        if any(kw in basename for kw in file_keywords):
            files.append(f)
    return files


def extract_all_records(
    file_path: str,
    file_keywords: list[str],
    *,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
) -> list[dict]:
    """从文件或文件夹提取全部维修记录（去重）。

    Args:
        file_path: 文件或文件夹路径。
        file_keywords: 文件名关键字。
        skip_hidden_rows: 跳过隐藏行。
        skip_hidden_cols: 跳过隐藏列。

    Returns:
        合并后的记录列表。
    """
    files = _discover_files(file_path, file_keywords)
    if not files:
        logger.warning("未找到匹配的文件 (关键字: %s)", file_keywords)
        return []

    all_records: list[dict] = []
    processed_months: set[tuple[int, int]] = set()

    for filepath in files:
        filename = os.path.basename(filepath)
        file_year, file_month = parse_year_month_from_filename(filename)
        logger.info("处理: %s (年=%s, 月=%s)", filename, file_year, file_month)

        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception as e:
            logger.error("无法打开文件 %s: %s", filepath, e)
            continue

        # 第二个 workbook 用于读取公式单元格的计算结果
        try:
            wb_values = openpyxl.load_workbook(filepath, data_only=True)
        except Exception:
            wb_values = None

        multi_sheet = len(wb.sheetnames) > 1

        for sheetname in wb.sheetnames:
            sheet_month = parse_month_from_sheetname(sheetname)
            year = file_year
            month = sheet_month if (multi_sheet and sheet_month) else (file_month or sheet_month)

            if not year or not month:
                logger.debug("跳过 sheet '%s': 无法确定年月", sheetname)
                continue

            key = (year, month)
            if key in processed_months:
                logger.debug("跳过 sheet '%s' -> %d年%d月 (已处理)", sheetname, year, month)
                continue

            ws = wb[sheetname]
            ws_values = wb_values[sheetname] if wb_values and sheetname in wb_values.sheetnames else None
            records = extract_sheet_records(
                ws, year, month,
                skip_hidden_rows=skip_hidden_rows,
                skip_hidden_cols=skip_hidden_cols,
                ws_values=ws_values,
            )
            all_records.extend(records)
            processed_months.add(key)
            logger.info("  sheet '%s' -> %d年%d月: %d 条记录", sheetname, year, month, len(records))

        wb.close()
        if wb_values:
            wb_values.close()

    logger.info("共提取 %d 条维修记录 (%d 个月)", len(all_records), len(processed_months))
    return all_records


# ── 主处理入口 ────────────────────────────────────────────────

def process_maintenance_data(
    file_path: str,
    *,
    eq_ledger=None,
    classifications: dict | None = None,
    file_keywords: list[str] | None = None,
    skip_hidden_rows: bool = False,
    skip_hidden_cols: bool = False,
    return_sheets: bool = False,
    split_by_year: bool = False,
) -> str | list[str] | dict[str, pd.DataFrame]:
    """维修记录处理统一入口。

    流程: 提取 → 预处理 → 台账匹配 → 分类 → 构建 DataFrame → 输出 Excel。

    Args:
        file_path: 输入文件或文件夹路径。
        eq_ledger: EquipmentLedger 实例（可选）。
        classifications: 分类配置 dict（None 时从 config 加载或使用默认值）。
        file_keywords: 文件名关键字（None 时从 config 加载）。
        skip_hidden_rows: 跳过隐藏行。
        skip_hidden_cols: 跳过隐藏列。
        return_sheets: True 时返回 dict[str, DataFrame]，False 时写文件。
        split_by_year: True 时按年份拆分输出为多个文件。

    Returns:
        输出文件路径（str）、文件路径列表（split_by_year=True）或 sheets 字典。
    """
    if not file_path or not os.path.exists(file_path):
        raise ValueError(f"输入路径不存在: {file_path}")

    # 加载配置
    if classifications is None:
        classifications = get_default_classifications()
    if file_keywords is None:
        try:
            from func.config_loader import get_maintenance_file_keywords
            file_keywords = get_maintenance_file_keywords()
        except Exception:
            file_keywords = ["设备出勤统计表"]

    class_rules = classifications.get("classifications", [])
    noise_exact = classifications.get("noise_exact", set())
    noise_patterns = classifications.get("noise_patterns", [])
    reason_rules = classifications.get("reason_rules", {})

    # 1. 提取记录
    raw_records = extract_all_records(
        file_path, file_keywords,
        skip_hidden_rows=skip_hidden_rows,
        skip_hidden_cols=skip_hidden_cols,
    )
    if not raw_records:
        msg = "未提取到任何维修记录"
        logger.warning(msg)
        if return_sheets:
            return {}
        raise ValueError(msg)

    # 2. 预处理 + 台账匹配 + 分类
    classified: list[dict] = []
    fault_records: list[dict] = []
    matched_count = 0
    unmatched_names: set[str] = set()

    for rec in raw_records:
        # 预处理设备名
        raw_name = preprocess_device_name(rec["原始设备名称"])
        std_name = raw_name
        match_method = ""

        # 台账匹配
        if eq_ledger and raw_name:
            result = eq_ledger.match(raw_name)
            if result:
                std_name = result["标准名称"]
                match_method = result["匹配方式"]
                matched_count += 1
            else:
                unmatched_names.add(raw_name)

        model = extract_device_model(std_name) if std_name else ""

        # 故障判定
        is_fault = is_fault_record(
            rec["原因"], rec["维修内容"],
            noise_exact=noise_exact,
            noise_patterns=noise_patterns,
            reason_rules=reason_rules,
        )

        # 分类
        major, minor = (None, None)
        if is_fault:
            major, minor = classify(
                rec["维修内容"],
                classifications=class_rules,
                noise_exact=noise_exact,
                noise_patterns=noise_patterns,
            )

        classified_rec = {
            "日期": rec["日期"],
            "原始设备名称": raw_name,
            "标准设备名称": std_name,
            "设备型号": model,
            "原因": rec["原因"],
            "班次": rec["班次"],
            "大类": major,
            "小类": minor,
            "是否故障": "是" if is_fault else "否",
            "维修内容": rec["维修内容"],
            "工时_分钟": rec["工时_分钟"],
        }
        classified.append(classified_rec)
        if is_fault and major is not None:
            fault_records.append(classified_rec)

    logger.info("分类完成: 总 %d 条, 故障 %d 条, 台账匹配 %d 条",
                len(classified), len(fault_records), matched_count)
    if unmatched_names:
        logger.warning("未匹配设备名 %d 个: %s", len(unmatched_names),
                       ", ".join(sorted(unmatched_names)[:10]))

    # 3. 构建统计 DataFrame
    sheets = _build_sheets(classified, fault_records)

    # 4. 输出
    if return_sheets:
        return sheets

    output_dir = file_path if os.path.isdir(file_path) else os.path.dirname(file_path) or "."

    if split_by_year:
        output_files = _write_split_by_year(output_dir, classified, fault_records)
        return output_files

    output_file = os.path.join(output_dir, "维修记录统计.xlsx")
    _write_excel(output_file, sheets)
    logger.info("输出完成: %s", output_file)
    return output_file


# ── 统计构建 ──────────────────────────────────────────────────

def _write_split_by_year(
    output_dir: str,
    classified: list[dict],
    fault_records: list[dict],
) -> list[str]:
    """按年份拆分输出为多个 Excel 文件。

    每个年份生成一个独立文件，另外生成一个汇总文件。
    汇总文件额外包含大类×年月和设备名称×原因两个跨年统计 sheet。

    Returns:
        输出文件路径列表。
    """
    from collections import defaultdict as _dd

    # 按年分组
    year_classified: dict[int, list] = _dd(list)
    year_faults: dict[int, list] = _dd(list)
    for rec in classified:
        if rec["日期"]:
            year_classified[rec["日期"].year].append(rec)
    for rec in fault_records:
        if rec["日期"]:
            year_faults[rec["日期"].year].append(rec)

    output_files: list[str] = []

    # 每年一个文件
    for year in sorted(year_classified.keys()):
        sheets = _build_sheets(year_classified[year], year_faults.get(year, []))
        output_file = os.path.join(output_dir, f"维修记录统计_{year}年.xlsx")
        _write_excel(output_file, sheets)
        output_files.append(output_file)
        logger.info("输出完成 (%d年): %s (%d 条)", year, output_file, len(year_classified[year]))

    # 汇总文件
    all_sheets = _build_sheets(classified, fault_records)
    summary_file = os.path.join(output_dir, "维修记录统计_汇总.xlsx")
    _write_excel(summary_file, all_sheets)
    output_files.append(summary_file)
    logger.info("汇总输出完成: %s (%d 条)", summary_file, len(classified))

    return output_files

def _safe_minutes(val) -> int:
    """安全取工时分钟数，None/NaN → 0。"""
    return int(val) if val else 0


def _build_sheets(
    classified: list[dict],
    fault_records: list[dict],
) -> dict[str, pd.DataFrame]:
    """从分类记录构建 10 个统计 DataFrame。"""
    sheets: dict[str, pd.DataFrame] = {}

    # ── Sheet 1: 维修明细 ──
    detail_rows = []
    for rec in classified:
        detail_rows.append({
            "日期": rec["日期"],
            "原始设备名称": rec["原始设备名称"],
            "标准设备名称": rec["标准设备名称"],
            "设备型号": rec["设备型号"],
            "原因": rec["原因"],
            "班次": rec["班次"],
            "大类": rec["大类"] or "",
            "小类": rec["小类"] or "",
            "是否故障": rec["是否故障"],
            "维修内容": rec["维修内容"],
            "工时_分钟": rec["工时_分钟"],
        })
    sheets["维修明细"] = pd.DataFrame(detail_rows)

    # ── 设备日期范围（全量记录用于计算统计跨度）──
    device_dates: dict[str, dict] = defaultdict(
        lambda: {"min_date": None, "max_date": None}
    )
    for rec in classified:
        v = rec["标准设备名称"]
        d = rec["日期"]
        if v and d:
            info = device_dates[v]
            if info["min_date"] is None or d < info["min_date"]:
                info["min_date"] = d
            if info["max_date"] is None or d > info["max_date"]:
                info["max_date"] = d

    # ── Sheet 2: 大类汇总（仅故障记录）──
    major_stats: dict[str, dict] = defaultdict(lambda: {"count": 0, "minutes": 0})
    for rec in fault_records:
        major_stats[rec["大类"]]["count"] += 1
        major_stats[rec["大类"]]["minutes"] += _safe_minutes(rec["工时_分钟"])
    total_fault_count = len(fault_records)
    total_fault_minutes = sum(s["minutes"] for s in major_stats.values())

    major_rows = []
    for major, s in sorted(major_stats.items(), key=lambda x: -x[1]["count"]):
        major_rows.append({
            "大类": major,
            "记录数": s["count"],
            "工时_分钟": s["minutes"],
            "工时_小时": round(s["minutes"] / 60, 1),
            "占比(记录)": s["count"] / total_fault_count if total_fault_count else 0,
            "占比(工时)": s["minutes"] / total_fault_minutes if total_fault_minutes else 0,
        })
    # 合计行
    major_rows.append({
        "大类": "合计",
        "记录数": total_fault_count,
        "工时_分钟": total_fault_minutes,
        "工时_小时": round(total_fault_minutes / 60, 1),
        "占比(记录)": 1.0,
        "占比(工时)": 1.0,
    })
    sheets["大类汇总"] = pd.DataFrame(major_rows)

    # ── Sheet 3: 大类×小类 ──
    sub_stats: dict[tuple, dict] = defaultdict(lambda: {"count": 0, "minutes": 0})
    major_totals: dict[str, int] = defaultdict(int)
    for rec in fault_records:
        key = (rec["大类"], rec["小类"])
        sub_stats[key]["count"] += 1
        sub_stats[key]["minutes"] += _safe_minutes(rec["工时_分钟"])
        major_totals[rec["大类"]] += 1

    sub_rows = []
    for (major, minor), s in sorted(sub_stats.items(), key=lambda x: (x[0][0], -x[1]["count"])):
        sub_rows.append({
            "大类": major,
            "小类": minor,
            "记录数": s["count"],
            "工时_分钟": s["minutes"],
            "工时_小时": round(s["minutes"] / 60, 1),
            "占大类比": s["count"] / major_totals[major] if major_totals[major] else 0,
        })
    sheets["大类×小类"] = pd.DataFrame(sub_rows)

    # ── Sheet 4: 按设备统计 ──
    dev_stats: dict[str, dict] = defaultdict(
        lambda: {"count": 0, "minutes": 0, "days": set(), "model": ""}
    )
    for rec in fault_records:
        v = rec["标准设备名称"]
        ds = dev_stats[v]
        ds["model"] = rec["设备型号"]
        ds["count"] += 1
        ds["minutes"] += _safe_minutes(rec["工时_分钟"])
        if rec["日期"]:
            ds["days"].add(rec["日期"])

    dev_rows = []
    for v, ds in sorted(dev_stats.items()):
        dd = device_dates.get(v, {})
        min_d = dd.get("min_date")
        max_d = dd.get("max_date")
        total_days = (max_d - min_d).days + 1 if min_d and max_d else 0
        total_minutes = total_days * 24 * 60 if total_days else 0
        rate = ds["minutes"] / total_minutes if total_minutes else 0
        dev_rows.append({
            "设备型号": ds["model"],
            "标准设备名称": v,
            "统计开始日期": min_d,
            "统计结束日期": max_d,
            "总日数": total_days,
            "有故障日数": len(ds["days"]),
            "总故障分钟": ds["minutes"],
            "总故障小时": round(ds["minutes"] / 60, 1),
            "故障率": rate,
        })
    sheets["按设备统计"] = pd.DataFrame(dev_rows)

    # ── Sheet 5: 按设备型号统计 ──
    model_stats: dict[str, dict] = defaultdict(
        lambda: {"count": 0, "minutes": 0, "devices": set(), "fault_days": set()}
    )
    for rec in fault_records:
        model = rec["设备型号"]
        ms = model_stats[model]
        ms["count"] += 1
        ms["minutes"] += _safe_minutes(rec["工时_分钟"])
        ms["devices"].add(rec["标准设备名称"])
        if rec["日期"]:
            ms["fault_days"].add(rec["日期"])

    model_rows = []
    for model, ms in sorted(model_stats.items()):
        model_total_days = 0
        for v in ms["devices"]:
            dd = device_dates.get(v, {})
            min_d, max_d = dd.get("min_date"), dd.get("max_date")
            if min_d and max_d:
                model_total_days += (max_d - min_d).days + 1
        model_total_minutes = model_total_days * 24 * 60 if model_total_days else 0
        rate = ms["minutes"] / model_total_minutes if model_total_minutes else 0
        model_rows.append({
            "设备型号": model,
            "台数": len(ms["devices"]),
            "总统计日数": model_total_days,
            "总故障分钟": ms["minutes"],
            "总故障小时": round(ms["minutes"] / 60, 1),
            "有故障日数": len(ms["fault_days"]),
            "总故障次数": ms["count"],
            "故障率": rate,
        })
    sheets["按设备型号统计"] = pd.DataFrame(model_rows)

    # ── Sheet 6: 大类×年月 ──
    month_stats: dict[tuple, dict] = defaultdict(lambda: {"count": 0, "minutes": 0})
    for rec in fault_records:
        key = (rec["大类"], rec["日期"].year, rec["日期"].month) if rec["日期"] else None
        if key:
            month_stats[key]["count"] += 1
            month_stats[key]["minutes"] += _safe_minutes(rec["工时_分钟"])

    month_rows = []
    for (major, y, m), s in sorted(month_stats.items()):
        month_rows.append({
            "大类": major,
            "年月": f"{y}-{m:02d}",
            "记录数": s["count"],
            "工时_分钟": s["minutes"],
            "工时_小时": round(s["minutes"] / 60, 1),
        })
    sheets["大类×年月"] = pd.DataFrame(month_rows)

    # ── Sheet 7: 设备名称×大类小类 ──
    dev_sub_stats: dict[tuple, dict] = defaultdict(lambda: {"count": 0, "minutes": 0, "model": ""})
    for rec in fault_records:
        key = (rec["标准设备名称"], rec["大类"], rec["小类"])
        ds = dev_sub_stats[key]
        ds["model"] = rec["设备型号"]
        ds["count"] += 1
        ds["minutes"] += _safe_minutes(rec["工时_分钟"])

    dev_sub_rows = []
    for (v, major, minor), ds in sorted(dev_sub_stats.items(), key=lambda x: (x[0][0], -x[1]["count"])):
        dd = device_dates.get(v, {})
        min_d, max_d = dd.get("min_date"), dd.get("max_date")
        total_days = (max_d - min_d).days + 1 if min_d and max_d else 0
        total_minutes = total_days * 24 * 60 if total_days else 0
        rate = ds["minutes"] / total_minutes if total_minutes else 0
        dev_sub_rows.append({
            "设备型号": ds["model"],
            "标准设备名称": v,
            "大类": major,
            "小类": minor,
            "总故障分钟": ds["minutes"],
            "总故障小时": round(ds["minutes"] / 60, 1),
            "故障率": rate,
        })
    sheets["设备名称×大类小类"] = pd.DataFrame(dev_sub_rows)

    # ── Sheet 8: 型号×大类小类 ──
    model_sub_stats: dict[tuple, dict] = defaultdict(
        lambda: {"count": 0, "minutes": 0, "devices": set(), "fault_days": set()}
    )
    for rec in fault_records:
        key = (rec["设备型号"], rec["大类"], rec["小类"])
        ms = model_sub_stats[key]
        ms["count"] += 1
        ms["minutes"] += _safe_minutes(rec["工时_分钟"])
        ms["devices"].add(rec["标准设备名称"])
        if rec["日期"]:
            ms["fault_days"].add(rec["日期"])

    model_sub_rows = []
    for (model, major, minor), ms in sorted(model_sub_stats.items()):
        model_total_days = sum(
            (device_dates[v]["max_date"] - device_dates[v]["min_date"]).days + 1
            for v in ms["devices"]
            if v in device_dates and device_dates[v]["min_date"] and device_dates[v]["max_date"]
        )
        model_total_minutes = model_total_days * 24 * 60 if model_total_days else 0
        rate = ms["minutes"] / model_total_minutes if model_total_minutes else 0
        model_sub_rows.append({
            "设备型号": model,
            "大类": major,
            "小类": minor,
            "台数": len(ms["devices"]),
            "总统计日数": model_total_days,
            "总故障分钟": ms["minutes"],
            "总故障小时": round(ms["minutes"] / 60, 1),
            "故障率": rate,
        })
    sheets["型号×大类小类"] = pd.DataFrame(model_sub_rows)

    # ── Sheet 9: 设备名称×原因 ──
    reason_stats: dict[tuple, dict] = defaultdict(
        lambda: {"count": 0, "minutes": 0, "days": set(), "model": ""}
    )
    for rec in fault_records:
        key = (rec["标准设备名称"], rec["原因"])
        rs = reason_stats[key]
        rs["model"] = rec["设备型号"]
        rs["count"] += 1
        rs["minutes"] += _safe_minutes(rec["工时_分钟"])
        if rec["日期"]:
            rs["days"].add(rec["日期"])

    reason_rows = []
    for (v, reason), rs in sorted(reason_stats.items(), key=lambda x: (x[0][0], -x[1]["count"])):
        dd = device_dates.get(v, {})
        min_d, max_d = dd.get("min_date"), dd.get("max_date")
        total_days = (max_d - min_d).days + 1 if min_d and max_d else 0
        total_minutes = total_days * 24 * 60 if total_days else 0
        rate = rs["minutes"] / total_minutes if total_minutes else 0
        reason_rows.append({
            "设备型号": rs["model"],
            "标准设备名称": v,
            "原因": reason,
            "故障次数": rs["count"],
            "统计开始日期": min_d,
            "统计结束日期": max_d,
            "总日数": total_days,
            "有故障日数": len(rs["days"]),
            "总故障分钟": rs["minutes"],
            "总故障小时": round(rs["minutes"] / 60, 1),
            "故障率": rate,
        })
    sheets["设备名称×原因"] = pd.DataFrame(reason_rows)

    # ── Sheet 10: 发动机故障深挖 ──
    engine_rows = []
    for rec in classified:
        if rec["大类"] != "发动机" or rec["是否故障"] != "是":
            continue
        engine_rows.append({
            "设备型号": rec["设备型号"],
            "标准设备名称": rec["标准设备名称"],
            "日期": rec["日期"],
            "小类": rec["小类"] or "",
            "班次": rec["班次"],
            "维修内容": rec["维修内容"],
            "工时_分钟": rec["工时_分钟"],
        })
    engine_rows.sort(key=lambda r: (r["设备型号"], r["标准设备名称"], r["日期"]))
    sheets["发动机故障深挖"] = pd.DataFrame(engine_rows)

    return sheets


# ── Excel 写入 ────────────────────────────────────────────────

def _write_excel(output_file: str, sheets: dict[str, pd.DataFrame]) -> None:
    """用 xlsxwriter 将所有 sheet 写入 Excel。"""
    wb = xlsxwriter.Workbook(output_file, {"strings_to_urls": False, "nan_inf_to_errors": True})
    fmts = _make_formats(wb)

    # Sheet 1: 维修明细
    if "维修明细" in sheets:
        df = sheets["维修明细"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "维修明细", headers, rows, fmts,
                     date_cols={0}, wrap_cols={9})

    # Sheet 2: 大类汇总
    if "大类汇总" in sheets:
        df = sheets["大类汇总"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "大类汇总", headers, rows, fmts,
                     pct_cols={4, 5}, hour_cols={3})

    # Sheet 3: 大类×小类
    if "大类×小类" in sheets:
        df = sheets["大类×小类"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "大类×小类", headers, rows, fmts,
                     pct_cols={5}, hour_cols={4})

    # Sheet 4: 按设备统计
    if "按设备统计" in sheets:
        df = sheets["按设备统计"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "按设备统计", headers, rows, fmts,
                     date_cols={2, 3}, pct_cols={8}, hour_cols={7})

    # Sheet 5: 按设备型号统计
    if "按设备型号统计" in sheets:
        df = sheets["按设备型号统计"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "按设备型号统计", headers, rows, fmts,
                     pct_cols={7}, hour_cols={4})

    # Sheet 6: 大类×年月
    if "大类×年月" in sheets:
        df = sheets["大类×年月"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "大类×年月", headers, rows, fmts,
                     hour_cols={4})

    # Sheet 7: 设备名称×大类小类
    if "设备名称×大类小类" in sheets:
        df = sheets["设备名称×大类小类"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "设备名称×大类小类", headers, rows, fmts,
                     pct_cols={6}, hour_cols={5})

    # Sheet 8: 型号×大类小类
    if "型号×大类小类" in sheets:
        df = sheets["型号×大类小类"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "型号×大类小类", headers, rows, fmts,
                     pct_cols={7}, hour_cols={6})

    # Sheet 9: 设备名称×原因
    if "设备名称×原因" in sheets:
        df = sheets["设备名称×原因"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "设备名称×原因", headers, rows, fmts,
                     date_cols={4, 5}, pct_cols={10}, hour_cols={9})

    # Sheet 10: 发动机故障深挖
    if "发动机故障深挖" in sheets:
        df = sheets["发动机故障深挖"]
        headers = list(df.columns)
        rows = [tuple(row) for row in df.itertuples(index=False, name=None)]
        _write_sheet(wb, "发动机故障深挖", headers, rows, fmts,
                     date_cols={2}, wrap_cols={5})

    wb.close()


# ── CLI ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="维修记录处理：从出勤统计表提取维修记录并生成统计报告")
    parser.add_argument("input_path", help="输入文件或文件夹路径")
    parser.add_argument("--ledger", help="设备台账 Excel 文件路径", default=None)
    parser.add_argument("--config", help="维修分类配置 Excel 文件路径", default=None)
    parser.add_argument("--skip-hidden-rows", action="store_true", help="跳过隐藏行")
    args = parser.parse_args()

    setup_logging()

    # 加载分类配置
    classifications = None
    if args.config:
        from func.maintenance_classification import import_classifications_from_excel
        classifications = import_classifications_from_excel(args.config)
        logger.info("使用自定义分类配置: %s", args.config)

    # 加载台账
    eq_ledger = None
    if args.ledger:
        from func.equipment_ledger import EquipmentLedger
        eq_ledger = EquipmentLedger(args.ledger)
        logger.info("使用设备台账: %s", args.ledger)

    output = process_maintenance_data(
        args.input_path,
        eq_ledger=eq_ledger,
        classifications=classifications,
        skip_hidden_rows=args.skip_hidden_rows,
    )
    print(f"\n输出: {output}")


if __name__ == "__main__":
    main()
