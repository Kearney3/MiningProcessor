"""Tests for func.excel_formatter — unified Excel formatting output."""

import datetime
from pathlib import Path

import pandas as pd
import pytest


# ── _display_width ────────────────────────────────────────────────────────


class TestDisplayWidth:
    def test_ascii_string(self):
        from func.excel_formatter import _display_width

        assert _display_width("hello") == 5

    def test_cjk_characters_count_as_two(self):
        from func.excel_formatter import _display_width

        assert _display_width("日期") == 4

    def test_mixed_ascii_and_cjk(self):
        from func.excel_formatter import _display_width

        # 日期(2*2=4) + Date(4*1=4) = 8
        assert _display_width("日期Date") == 8

    def test_empty_string(self):
        from func.excel_formatter import _display_width

        assert _display_width("") == 0

    def test_fullwidth_forms(self):
        from func.excel_formatter import _display_width

        # Fullwidth A (U+FF21)
        assert _display_width("Ａ") == 2


# ── _auto_column_widths ──────────────────────────────────────────────────


class TestAutoColumnWidths:
    def test_basic_widths(self):
        from func.excel_formatter import _auto_column_widths

        df = pd.DataFrame({"Name": ["Alice", "Bob"], "Age": [30, 25]})
        widths = _auto_column_widths(df, min_width=5, max_width=30)
        assert len(widths) == 2
        # "Name" header = 4 + padding(2) = 6, "Alice" = 5+2 = 7
        assert widths[0] >= 6
        # "Age" header = 3+2 = 5
        assert widths[1] >= 5

    def test_respects_min_width(self):
        from func.excel_formatter import _auto_column_widths

        df = pd.DataFrame({"A": [1]})
        widths = _auto_column_widths(df, min_width=10, max_width=50)
        assert widths[0] == 10

    def test_respects_max_width(self):
        from func.excel_formatter import _auto_column_widths

        df = pd.DataFrame({"col": ["a" * 100]})
        widths = _auto_column_widths(df, min_width=5, max_width=20)
        assert widths[0] == 20

    def test_cjk_content_wider(self):
        from func.excel_formatter import _auto_column_widths

        df = pd.DataFrame({"设备名称": ["卡特彼勒320挖掘机"]})
        widths = _auto_column_widths(df, min_width=5, max_width=50)
        # header "设备名称" = 8 + 2 = 10, content "卡特彼勒320挖掘机" = 16 + 2 = 18
        assert widths[0] >= 10

    def test_date_columns_use_fixed_width(self):
        from func.excel_formatter import _auto_column_widths

        df = pd.DataFrame({"日期": [datetime.date(2025, 1, 1)]})
        widths = _auto_column_widths(df, min_width=5, max_width=50)
        # date should be detected and use width 12 + padding
        assert widths[0] >= 12


# ── _is_date_column ──────────────────────────────────────────────────────


class TestIsDateColumn:
    def test_datetime64_dtype(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series(pd.to_datetime(["2025-01-01", "2025-06-15"]))
        assert _is_date_column(s) is True

    def test_python_date_objects(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series([datetime.date(2025, 1, 1), datetime.date(2025, 6, 15)])
        assert _is_date_column(s) == True  # noqa: E712

    def test_python_datetime_objects(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series([datetime.datetime(2025, 1, 1, 12, 0)])
        assert _is_date_column(s) == True  # noqa: E712

    def test_string_column_not_date(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series(["2025-01-01", "2025-06-15"])
        assert _is_date_column(s) == False  # noqa: E712

    def test_numeric_column_not_date(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series([1, 2, 3])
        assert _is_date_column(s) == False  # noqa: E712

    def test_empty_series(self):
        from func.excel_formatter import _is_date_column

        s = pd.Series([], dtype=object)
        assert _is_date_column(s) == False  # noqa: E712


# ── write_formatted_excel ────────────────────────────────────────────────


class TestWriteFormattedExcel:
    def test_creates_file(self, tmp_path):
        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "test.xlsx"
        df = pd.DataFrame({"Name": ["Alice"], "Age": [30]})
        result = write_formatted_excel(str(out), {"Sheet1": df})
        assert Path(result).exists()

    def test_multi_sheet_output(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "multi.xlsx"
        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        write_formatted_excel(str(out), {"First": df1, "Second": df2})

        wb = load_workbook(str(out))
        assert wb.sheetnames == ["First", "Second"]
        wb.close()

    def test_header_has_bold_font(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "header.xlsx"
        df = pd.DataFrame({"Col1": [1]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        cell = wb["S"].cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        wb.close()

    def test_header_has_fill(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "fill.xlsx"
        df = pd.DataFrame({"Col": [1]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        cell = wb["S"].cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "004472C4"
        wb.close()

    def test_freeze_panes_set(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "freeze.xlsx"
        df = pd.DataFrame({"A": [1, 2]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        assert wb["S"].freeze_panes == "A2"
        wb.close()

    def test_auto_filter_set(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "filter.xlsx"
        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        assert wb["S"].auto_filter.ref is not None
        assert "A1:B3" in wb["S"].auto_filter.ref
        wb.close()

    def test_date_column_formatted(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "date.xlsx"
        df = pd.DataFrame({
            "日期": pd.to_datetime(["2025-01-15", "2025-06-30"]),
            "值": [10, 20],
        })
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        ws = wb["S"]
        # Date column (col 1) should have date format
        assert ws.cell(row=2, column=1).number_format == "yyyy-mm-dd"
        # Value column (col 2) should not
        assert ws.cell(row=2, column=2).number_format == "General"
        wb.close()

    def test_python_date_objects_formatted(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "pydate.xlsx"
        df = pd.DataFrame({"日期": [datetime.date(2025, 3, 15)]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        assert wb["S"].cell(row=2, column=1).number_format == "yyyy-mm-dd"
        wb.close()

    def test_column_widths_applied(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "width.xlsx"
        df = pd.DataFrame({"Name": ["Alice", "Bob"], "X": [1, 2]})
        write_formatted_excel(str(out), {"S": df}, min_col_width=5, max_col_width=50)

        wb = load_workbook(str(out))
        ws = wb["S"]
        # Widths should be > 0
        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > 0
        wb.close()

    def test_empty_dataframe(self, tmp_path):
        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "empty.xlsx"
        df = pd.DataFrame()
        write_formatted_excel(str(out), {"S": df})
        assert Path(out).exists()

    def test_returns_output_path(self, tmp_path):
        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "ret.xlsx"
        df = pd.DataFrame({"A": [1]})
        result = write_formatted_excel(str(out), {"S": df})
        assert result == str(out)

    def test_overwrites_existing_file(self, tmp_path):
        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "overwrite.xlsx"
        # Write first time
        write_formatted_excel(str(out), {"S": pd.DataFrame({"A": [1]})})
        # Write second time with different data
        write_formatted_excel(str(out), {"S": pd.DataFrame({"B": [2, 3]})})

        from openpyxl import load_workbook

        wb = load_workbook(str(out))
        assert wb["S"].cell(row=1, column=1).value == "B"
        wb.close()

    def test_custom_header_colors(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "custom.xlsx"
        df = pd.DataFrame({"A": [1]})
        write_formatted_excel(
            str(out), {"S": df},
            header_fill="FF0000",
            header_font_color="000000",
        )

        wb = load_workbook(str(out))
        cell = wb["S"].cell(row=1, column=1)
        # openpyxl returns 8-char ARGB (alpha prefix + 6-digit hex)
        assert cell.font.color.rgb == "00000000"
        wb.close()

    def test_nan_values_written(self, tmp_path):
        from openpyxl import load_workbook

        from func.excel_formatter import write_formatted_excel

        out = tmp_path / "nan.xlsx"
        df = pd.DataFrame({"A": [1.0, None, 3.0]})
        write_formatted_excel(str(out), {"S": df})

        wb = load_workbook(str(out))
        ws = wb["S"]
        assert ws.cell(row=2, column=1).value == 1.0
        # NaN -> None in openpyxl
        assert ws.cell(row=3, column=1).value is None
        wb.close()
