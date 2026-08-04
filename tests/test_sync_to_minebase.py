"""sync_to_minebase 模块测试"""
import json
import pathlib
import sys
import uuid
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from func.sync_to_minebase import (
    MineBaseAPIClient,
    _apply_defaults,
    _build_field_mappings,
    _df_to_mapped_rows,
    _filter_by_date_range,
    _map_row_to_db_columns,
    _resolve_fks_for_db,
    discover_files,
    load_column_mapping,
    read_and_map_excel,
    sync,
    sync_via_api,
)


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_mapping(tmp_path):
    """创建临时映射配置文件。"""
    mapping = {
        "fuel_consumption": {
            "日期": "date",
            "班次": "shiftType",
            "设备名称": "equipmentName",
            "油品种类": "fuelName",
            "油品消耗": "consumption",
        },
        "electricity_consumption": {
            "日期": "date",
            "班次": "shiftType",
            "设备名称": "equipmentName",
            "电力消耗": "consumption",
        },
        "equipment_operation": {
            "日期": "date",
            "班次": "shiftType",
            "设备名称": "equipmentName",
            "运行小时数": "runningHours",
            "趟次": "tripCount",
        },
        "production_record": {
            "日期": "date",
            "班次": "shiftType",
            "矿卡名称": "truckName",
            "挖机名称": "excavatorName",
            "矿石类型": "materialTypeName",
            "运次": "tripCount",
            "产量": "production",
        },
        "work_efficiency": {
            "设备名称": "equipmentName",
            "应运行分钟": "plannedMinutes",
            "停车/换班": "parkShift",
        },
    }
    path = tmp_path / "mapping.json"
    path.write_text(json.dumps(mapping, ensure_ascii=False), encoding="utf-8")
    return path, mapping


@pytest.fixture
def sample_fuel_excel(tmp_path_factory):
    """创建模拟的 Fuel.xlsx 油耗信息 sheet。"""
    tmp = tmp_path_factory.mktemp("fuel")
    df = pd.DataFrame({
        "日期": [date(2025, 6, 1), date(2025, 6, 1)],
        "班次": ["Day", "Night"],
        "设备名称": ["CAT785D-01", "NTE240-02"],
        "设备编号": ["CAT785D-01", "NTE240-02"],
        "油品种类": ["0#柴油", "0#柴油"],
        "油品消耗": [150.5, 200.0],
    })
    path = tmp / "Fuel.xlsx"
    df.to_excel(path, sheet_name="油耗信息", index=False)
    return path


@pytest.fixture
def sample_electrical_excel(tmp_path_factory):
    """创建模拟的电力消耗统计.xlsx。"""
    tmp = tmp_path_factory.mktemp("elec")
    df = pd.DataFrame({
        "日期": [date(2025, 6, 1), date(2025, 6, 2)],
        "班次": ["Day", "Day"],
        "设备名称": ["EX-001", "EX-002"],
        "电力消耗": [500.0, 600.0],
    })
    path = tmp / "电力消耗统计.xlsx"
    df.to_excel(path, index=False)
    return path


@pytest.fixture
def sample_production_excel(tmp_path_factory):
    """创建模拟的合并产量.xlsx（两个 sheet）。"""
    tmp = tmp_path_factory.mktemp("prod")
    running = pd.DataFrame({
        "日期": [date(2025, 6, 1)],
        "班次": ["Day"],
        "设备名称": ["CAT785D-01"],
        "公司": ["测试公司"],
        "小时数仪表开始": [100.0],
        "小时数仪表结束": [108.0],
        "运行小时数": [8.0],
        "公里数仪表开始": [1000.0],
        "公里数仪表结束": [1050.0],
        "运行里程": [50.0],
        "趟次": [10],
    })
    production = pd.DataFrame({
        "日期": [date(2025, 6, 1)],
        "班次": ["Day"],
        "矿卡名称": ["CAT785D-01"],
        "挖机名称": ["EX-001"],
        "矿石类型": ["WASTE"],
        "运次": [10],
        "产量": [850.0],
    })
    path = tmp / "合并产量.xlsx"
    with pd.ExcelWriter(path) as writer:
        running.to_excel(writer, sheet_name="运行数据", index=False)
        production.to_excel(writer, sheet_name="生产数据", index=False)
    return path


@pytest.fixture
def sample_worktime_excel(tmp_path_factory):
    """创建模拟的工作效率表。"""
    tmp = tmp_path_factory.mktemp("wt")
    df = pd.DataFrame({
        "日期": ["2025-06-01"],
        "班次": ["Day"],
        "设备名称": ["CAT785D-01"],
        "应运行分钟": [720],
        "停车/换班": [30],
    })
    path = tmp / "202506_工作效率表.xlsx"
    df.to_excel(path, index=False)
    return path


# ---------------------------------------------------------------------------
# load_column_mapping
# ---------------------------------------------------------------------------


class TestLoadColumnMapping:
    def test_load_existing_file(self, sample_mapping):
        path, expected = sample_mapping
        result = load_column_mapping(path)
        assert result == expected

    def test_load_nonexistent_file_returns_defaults(self, tmp_path):
        result = load_column_mapping(tmp_path / "nonexistent.json")
        assert "fuel_consumption" in result
        assert "work_efficiency" in result

    def test_load_default_file(self):
        """默认映射文件应该存在且包含所有 5 种数据类型。"""
        result = load_column_mapping()
        assert "fuel_consumption" in result
        assert "electricity_consumption" in result
        assert "equipment_operation" in result
        assert "production_record" in result
        assert "work_efficiency" in result

    def test_default_mapping_has_remark_fields(self):
        """默认映射应包含 remark 字段映射（有备注列的数据类型）。"""
        result = load_column_mapping()
        assert result["equipment_operation"].get("备注") == "remark"
        assert result["production_record"].get("备注") == "remark"
        assert result["work_efficiency"].get("注释") == "remark"

    def test_default_mapping_has_company_for_operation(self):
        """equipment_operation 映射应包含公司 → company。"""
        result = load_column_mapping()
        assert result["equipment_operation"].get("公司") == "company"


# ---------------------------------------------------------------------------
# read_and_map_excel
# ---------------------------------------------------------------------------


class TestReadAndMapExcel:
    def test_read_fuel(self, sample_fuel_excel, sample_mapping):
        _, mapping = sample_mapping
        rows = read_and_map_excel(sample_fuel_excel, "油耗信息", mapping["fuel_consumption"])
        assert len(rows) == 2
        assert rows[0]["date"] == "2025-06-01"
        assert rows[0]["shiftType"] == "Day"
        assert rows[0]["equipmentName"] == "CAT785D-01"
        assert rows[0]["fuelName"] == "0#柴油"
        assert rows[0]["consumption"] == pytest.approx(150.5)

    def test_read_electrical(self, sample_electrical_excel, sample_mapping):
        _, mapping = sample_mapping
        rows = read_and_map_excel(sample_electrical_excel, None, mapping["electricity_consumption"])
        assert len(rows) == 2
        assert rows[0]["equipmentName"] == "EX-001"
        assert rows[0]["consumption"] == pytest.approx(500.0)

    def test_read_production_running(self, sample_production_excel, sample_mapping):
        _, mapping = sample_mapping
        rows = read_and_map_excel(sample_production_excel, "运行数据", mapping["equipment_operation"])
        assert len(rows) == 1
        assert rows[0]["equipmentName"] == "CAT785D-01"
        assert rows[0]["runningHours"] == pytest.approx(8.0)
        assert rows[0]["tripCount"] == 10

    def test_read_production_production(self, sample_production_excel, sample_mapping):
        _, mapping = sample_mapping
        rows = read_and_map_excel(sample_production_excel, "生产数据", mapping["production_record"])
        assert len(rows) == 1
        assert rows[0]["truckName"] == "CAT785D-01"
        assert rows[0]["excavatorName"] == "EX-001"
        assert rows[0]["production"] == pytest.approx(850.0)

    def test_read_empty_excel(self, tmp_path, sample_mapping):
        _, mapping = sample_mapping
        df = pd.DataFrame(columns=["日期", "班次", "设备名称", "油品种类", "油品消耗"])
        path = tmp_path / "empty.xlsx"
        df.to_excel(path, index=False)
        rows = read_and_map_excel(path, None, mapping["fuel_consumption"])
        assert rows == []

    def test_nan_values_skipped(self, tmp_path, sample_mapping):
        _, mapping = sample_mapping
        df = pd.DataFrame({
            "日期": [date(2025, 6, 1)],
            "班次": ["Day"],
            "设备名称": ["TEST-01"],
            "油品种类": ["0#柴油"],
            "油品消耗": [float("nan")],
        })
        path = tmp_path / "nan.xlsx"
        df.to_excel(path, index=False)
        rows = read_and_map_excel(path, None, mapping["fuel_consumption"])
        assert len(rows) == 1
        assert "consumption" not in rows[0]


# ---------------------------------------------------------------------------
# _build_field_mappings
# ---------------------------------------------------------------------------


class TestBuildFieldMappings:
    def test_fuel_mappings(self, sample_mapping):
        _, mapping = sample_mapping
        result = _build_field_mappings(mapping["fuel_consumption"], "fuel_consumption")
        assert len(result) == 5
        # equipmentName 应有 fkResolve
        equip_mapping = next(m for m in result if m["systemField"] == "equipmentName")
        assert equip_mapping["fkResolve"]["relation"] == "equipment"
        # date 不应有 fkResolve
        date_mapping = next(m for m in result if m["systemField"] == "date")
        assert "fkResolve" not in date_mapping

    def test_production_mappings(self, sample_mapping):
        _, mapping = sample_mapping
        result = _build_field_mappings(mapping["production_record"], "production_record")
        truck_mapping = next(m for m in result if m["systemField"] == "truckName")
        assert truck_mapping["fkResolve"]["relation"] == "truck"
        excavator_mapping = next(m for m in result if m["systemField"] == "excavatorName")
        assert excavator_mapping["fkResolve"]["relation"] == "excavator"
        material_mapping = next(m for m in result if m["systemField"] == "materialTypeName")
        assert material_mapping["fkResolve"]["relation"] == "materialType"


# ---------------------------------------------------------------------------
# _map_row_to_db_columns
# ---------------------------------------------------------------------------


class TestMapRowToDbColumns:
    def test_basic_mapping(self):
        row = {
            "date": "2025-06-01",
            "shiftType": "Day",
            "equipmentName": "TEST-01",
            "equipmentId": "uuid-123",
            "consumption": 150.5,
        }
        columns, values = _map_row_to_db_columns(row)
        assert "id" in columns
        assert "updated_at" in columns
        assert "date" in columns
        assert "shift_type" in columns
        assert "equipment_name" in columns
        assert "equipment_id" in columns
        assert "consumption" in columns

    def test_id_is_valid_uuid(self):
        """id 列必须是合法 UUID 字符串（Prisma @default(uuid()) 仅在应用层生效）。"""
        row = {"date": "2026-06-18", "shiftType": "Night", "equipmentName": "EX-001", "equipmentId": "eq-001"}
        columns, values = _map_row_to_db_columns(row)
        idx = columns.index("id")
        parsed = uuid.UUID(values[idx])
        assert str(parsed) == values[idx]

    def test_updated_at_is_recent(self):
        """updated_at 列必须是接近当前时间的 datetime（Prisma @updatedAt 仅在应用层生效）。"""
        row = {"date": "2026-06-18", "shiftType": "Night", "equipmentName": "EX-001", "equipmentId": "eq-001"}
        before = datetime.now()
        columns, values = _map_row_to_db_columns(row)
        after = datetime.now()
        idx = columns.index("updated_at")
        ts = values[idx]
        assert before <= ts <= after

    def test_unknown_fields_ignored(self):
        row = {"date": "2025-06-01", "unknownField": "value"}
        columns, values = _map_row_to_db_columns(row)
        assert "unknownField" not in columns
        # id + updated_at + date = 3 columns
        assert len(columns) == 3


# ---------------------------------------------------------------------------
# _apply_defaults
# ---------------------------------------------------------------------------


class TestApplyDefaults:
    def test_electrical_no_shift_gets_night(self):
        rows = [
            {"date": "2025-06-01", "equipmentName": "EX-001", "consumption": 500.0},
            {"date": "2025-06-02", "equipmentName": "EX-002", "consumption": 600.0},
        ]
        result = _apply_defaults(rows, "electrical")
        assert all(r["shiftType"] == "Night" for r in result)
        assert len(result) == 2

    def test_electrical_with_shift_not_overridden(self):
        rows = [
            {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "EX-001", "consumption": 500.0},
        ]
        result = _apply_defaults(rows, "electrical")
        assert result[0]["shiftType"] == "Day"

    def test_non_electrical_not_affected(self):
        rows = [
            {"date": "2025-06-01", "equipmentName": "CAT785D-01", "consumption": 150.0},
        ]
        result = _apply_defaults(rows, "fuel")
        assert "shiftType" not in result[0]

    def test_empty_rows(self):
        result = _apply_defaults([], "electrical")
        assert result == []

    def test_returns_new_list(self):
        """不修改原始列表（不可变性）。"""
        rows = [{"date": "2025-06-01", "consumption": 500.0}]
        result = _apply_defaults(rows, "electrical")
        assert result is not rows
        assert "shiftType" not in rows[0]
        assert result[0]["shiftType"] == "Night"


# ---------------------------------------------------------------------------
# _filter_by_date_range
# ---------------------------------------------------------------------------


class TestFilterByDateRange:
    SAMPLE_ROWS = [
        {"date": "2025-06-01", "equipmentName": "A", "consumption": 100},
        {"date": "2025-06-02", "equipmentName": "B", "consumption": 200},
        {"date": "2025-06-03", "equipmentName": "C", "consumption": 300},
        {"date": "2025-06-04", "equipmentName": "D", "consumption": 400},
    ]

    def test_no_filter_returns_all(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, None, None)
        assert len(result) == 4

    def test_start_date_filter(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, "2025-06-02", None)
        assert len(result) == 3
        assert result[0]["date"] == "2025-06-02"

    def test_end_date_filter(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, None, "2025-06-03")
        assert len(result) == 3
        assert result[-1]["date"] == "2025-06-03"

    def test_both_dates_filter(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, "2025-06-02", "2025-06-03")
        assert len(result) == 2
        assert result[0]["date"] == "2025-06-02"
        assert result[1]["date"] == "2025-06-03"

    def test_exact_date_match(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, "2025-06-03", "2025-06-03")
        assert len(result) == 1
        assert result[0]["date"] == "2025-06-03"

    def test_no_match(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, "2025-07-01", "2025-07-31")
        assert len(result) == 0

    def test_rows_without_date_kept(self):
        rows = [{"equipmentName": "X"}, {"date": "2025-06-01"}]
        result = _filter_by_date_range(rows, "2025-06-02", None)
        assert len(result) == 1  # 无 date 的行保留

    def test_returns_new_list(self):
        result = _filter_by_date_range(self.SAMPLE_ROWS, "2025-06-02", None)
        assert result is not self.SAMPLE_ROWS

    def test_empty_rows(self):
        result = _filter_by_date_range([], "2025-06-01", "2025-06-30")
        assert result == []


# ---------------------------------------------------------------------------
# discover_files (keyword-based)
# ---------------------------------------------------------------------------


class TestDiscoverFiles:
    def test_discover_all(self, tmp_path, sample_fuel_excel, sample_electrical_excel, sample_production_excel, sample_worktime_excel):
        # Copy files to same directory
        import shutil
        shutil.copy(sample_fuel_excel, tmp_path / "Fuel.xlsx")
        shutil.copy(sample_electrical_excel, tmp_path / "电力消耗统计.xlsx")
        shutil.copy(sample_production_excel, tmp_path / "合并产量.xlsx")
        shutil.copy(sample_worktime_excel, tmp_path / "202506_工作效率表.xlsx")

        found = discover_files(tmp_path)
        assert "fuel" in found
        assert "electrical" in found
        assert "operation" in found
        assert "production" in found
        assert "work_efficiency" in found

    def test_discover_partial(self, tmp_path, sample_fuel_excel):
        import shutil
        shutil.copy(sample_fuel_excel, tmp_path / "Fuel.xlsx")
        found = discover_files(tmp_path)
        assert "fuel" in found
        assert "electrical" not in found

    def test_discover_with_keywords(self, tmp_path):
        """关键字匹配模式与 excel_batch 一致。"""
        import shutil
        # 创建带有关键字的文件
        (tmp_path / "Fuel report 2025.xlsx").write_bytes(b"")
        (tmp_path / "工作效率表_2025.xlsx").write_bytes(b"")

        keywords = {
            "fuel": ["Fuel report"],
            "worktime": ["工作效率表"],
        }
        found = discover_files(tmp_path, keywords=keywords)
        assert "fuel" in found
        assert found["fuel"][0].name == "Fuel report 2025.xlsx"
        assert "work_efficiency" in found

    def test_discover_with_year_month(self, tmp_path):
        """year/month 用于 work_efficiency 文件名匹配。"""
        (tmp_path / "202501_工作效率表.xlsx").write_bytes(b"")
        (tmp_path / "202506_工作效率表.xlsx").write_bytes(b"")

        found = discover_files(tmp_path, year=2025, month=6, keywords={})
        assert "work_efficiency" in found
        assert "202506" in found["work_efficiency"][0].name


# ---------------------------------------------------------------------------
# MineBaseAPIClient
# ---------------------------------------------------------------------------


class TestMineBaseAPIClient:
    def test_init(self):
        client = MineBaseAPIClient("http://localhost:3000", "admin", "pass")
        assert client.base_url == "http://localhost:3000"
        assert client.token is None

    def test_init_trailing_slash(self):
        client = MineBaseAPIClient("http://localhost:3000/", "admin", "pass")
        assert client.base_url == "http://localhost:3000"


# ---------------------------------------------------------------------------
# MineBaseAPIClient contract v2 错误处理
# ---------------------------------------------------------------------------


class TestAPIClientV2Errors:
    """测试 contractVersion 2 错误码映射。"""

    def test_409_session_limit_raises_session_limit_error(self):
        """HTTP 409 + IMPORT_SESSION_LIMIT_REACHED 应抛出 SessionLimitReachedError。"""
        from func.sync.api_client import SessionLimitReachedError

        client = MineBaseAPIClient("http://localhost:3000", "admin", "pass")
        client.token = "fake-token"

        import urllib.error
        from io import BytesIO

        error_body = json.dumps({"errorCode": "IMPORT_SESSION_LIMIT_REACHED", "message": "Too many sessions"}).encode()
        resp = urllib.error.HTTPError(url="", code=409, msg="", hdrs=None, fp=BytesIO(error_body))
        resp.read = BytesIO(error_body).read

        with patch("urllib.request.urlopen", side_effect=resp):
            with pytest.raises(SessionLimitReachedError) as exc_info:
                client.create_session("fuel_consumption", expected_batches=1, total_rows=10)
            assert exc_info.value.status_code == 409
            assert exc_info.value.error_code == "IMPORT_SESSION_LIMIT_REACHED"

    def test_410_session_expired_raises_expired_error(self):
        """HTTP 410 应抛出 SessionExpiredError。"""
        from func.sync.api_client import SessionExpiredError

        client = MineBaseAPIClient("http://localhost:3000", "admin", "pass")
        client.token = "fake-token"

        import urllib.error
        from io import BytesIO

        error_body = json.dumps({"errorCode": "SESSION_EXPIRED", "message": "Session expired"}).encode()
        resp = urllib.error.HTTPError(url="", code=410, msg="", hdrs=None, fp=BytesIO(error_body))
        resp.read = BytesIO(error_body).read

        with patch("urllib.request.urlopen", side_effect=resp):
            with pytest.raises(SessionExpiredError) as exc_info:
                client.confirm_batch("fuel_consumption", "session-1")
            assert exc_info.value.status_code == 410

    def test_generic_409_raises_api_error(self):
        """HTTP 409 (非限制/过期) 应抛出 MineBaseAPIError。"""
        from func.sync.api_client import MineBaseAPIError

        client = MineBaseAPIClient("http://localhost:3000", "admin", "pass")
        client.token = "fake-token"

        import urllib.error
        from io import BytesIO

        error_body = json.dumps({"errorCode": "SESSION_STATE_CONFLICT", "message": "Bad state"}).encode()
        resp = urllib.error.HTTPError(url="", code=409, msg="", hdrs=None, fp=BytesIO(error_body))
        resp.read = BytesIO(error_body).read

        with patch("urllib.request.urlopen", side_effect=resp):
            with pytest.raises(MineBaseAPIError) as exc_info:
                client.send_batch("fuel_consumption", "session-1", [], [], 0, 1)
            assert exc_info.value.status_code == 409
            assert exc_info.value.error_code == "SESSION_STATE_CONFLICT"


# ---------------------------------------------------------------------------
# sync_via_api contract v2 流程测试
# ---------------------------------------------------------------------------


class TestSyncViaApiV2:
    """测试 sync_via_api 的 contractVersion 2 完整流程。"""

    def _make_v2_client(self, inserted=1, updated=0, skipped=0, rejected=0):
        """创建符合 v2 合约的 mock 客户端。"""
        client = MagicMock()
        client.create_session.return_value = {
            "data": {"session": {"id": "session-v2", "version": 1}},
        }
        client.send_batch.return_value = {
            "data": {
                "receipt": {"result": {"success": 1, "skipped": 0, "failed": 0}},
                "session": {"version": 1},
            },
        }
        client.confirm_batch.return_value = {
            "data": {
                "done": True,
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped,
                "rejected": rejected,
                "session": {"version": 2},
            },
        }
        return client

    def test_create_session_sends_contract_v2(self):
        """create_session 应发送 contractVersion=2 和必需参数。"""
        client = self._make_v2_client()
        rows = [{"date": "2025-06-01", "shiftType": "Day"}]
        sync_via_api("fuel", rows, {"日期": "date"}, client)

        client.create_session.assert_called_once()
        call_args = client.create_session.call_args
        assert call_args[0][0] == "fuel_consumption"  # table
        assert call_args[1]["expected_batches"] == 1
        assert call_args[1]["total_rows"] == 1
        assert call_args[1]["conflict_policy"] == "SKIP"

    def test_send_batch_sends_contract_v2(self):
        """send_batch 应发送 contractVersion=2、expectedVersion 和 rowOffset。"""
        client = self._make_v2_client()
        rows = [{"date": "2025-06-01", "shiftType": "Day"}]
        sync_via_api("fuel", rows, {"日期": "date"}, client)

        client.send_batch.assert_called_once()
        call_args = client.send_batch.call_args
        assert call_args[1]["expected_version"] == 1
        assert call_args[1]["row_offset"] == 0

    def test_confirm_batch_sends_contract_v2(self):
        """confirm_batch 应发送 contractVersion=2 和 expectedVersion。"""
        client = self._make_v2_client()
        rows = [{"date": "2025-06-01", "shiftType": "Day"}]
        sync_via_api("fuel", rows, {"日期": "date"}, client)

        client.confirm_batch.assert_called_once()
        call_args = client.confirm_batch.call_args
        assert call_args[0][0] == "fuel_consumption"  # table
        assert call_args[1]["expected_version"] == 1

    def test_windowed_confirm_loops_until_done(self):
        """窗口化 confirm 应循环调用直到 done=true。"""
        client = self._make_v2_client()
        # 第一次 confirm: done=false, 第二次: done=true
        client.confirm_batch.side_effect = [
            {"data": {"done": False, "inserted": 5, "updated": 0, "skipped": 0, "rejected": 0, "session": {"version": 2}}},
            {"data": {"done": True, "inserted": 5, "updated": 0, "skipped": 0, "rejected": 0, "session": {"version": 3}}},
        ]

        rows = [{"date": f"2025-06-{i:02d}"} for i in range(1, 11)]
        result = sync_via_api("fuel", rows, {"日期": "date"}, client)

        assert client.confirm_batch.call_count == 2
        assert result["success"] == 10  # 5 + 5
        assert result["skipped"] == 0

    def test_session_limit_returns_all_failed(self):
        """SessionLimitReachedError 应返回全部行失败。"""
        from func.sync.api_client import SessionLimitReachedError

        client = MagicMock()
        client.create_session.side_effect = SessionLimitReachedError(
            "Too many", status_code=409, error_code="IMPORT_SESSION_LIMIT_REACHED",
        )

        rows = [{"date": "2025-06-01"}, {"date": "2025-06-02"}]
        result = sync_via_api("fuel", rows, {"日期": "date"}, client)

        assert result["success"] == 0
        assert result["failed"] == 2

    def test_session_expired_returns_failed(self):
        """SessionExpiredError during confirm 应返回失败。"""
        from func.sync.api_client import SessionExpiredError

        client = MagicMock()
        client.create_session.return_value = {
            "data": {"session": {"id": "session-1", "version": 1}},
        }
        client.send_batch.return_value = {
            "data": {
                "receipt": {"result": {"success": 1, "skipped": 0, "failed": 0}},
                "session": {"version": 1},
            },
        }
        client.confirm_batch.side_effect = SessionExpiredError(
            "Expired", status_code=410, error_code="SESSION_EXPIRED",
        )

        rows = [{"date": "2025-06-01"}]
        result = sync_via_api("fuel", rows, {"日期": "date"}, client)

        assert result["failed"] == 1
        client.cancel_import.assert_called_once()


# ---------------------------------------------------------------------------
# sync (integration with mocks)
# ---------------------------------------------------------------------------


class TestSyncAPI:
    @patch("func.sync_to_minebase.MineBaseAPIClient")
    def test_sync_fuel_dry_run(self, mock_api_cls, tmp_path, sample_fuel_excel, sample_mapping):
        import shutil
        mapping_path, _ = sample_mapping
        shutil.copy(sample_fuel_excel, tmp_path / "Fuel.xlsx")

        results = sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True, mapping_file=mapping_path)
        assert "fuel" in results
        assert results["fuel"]["success"] == 0  # dry-run

    @patch("func.sync_to_minebase.MineBaseAPIClient")
    def test_sync_no_files(self, mock_api_cls, tmp_path, sample_mapping):
        mapping_path, _ = sample_mapping
        results = sync(tmp_path, mode="api", mapping_file=mapping_path)
        assert results == {}


class TestSyncDB:
    @patch("func.sync_to_minebase.MineBaseDBClient")
    def test_sync_fuel_dry_run(self, mock_db_cls, tmp_path, sample_fuel_excel, sample_mapping):
        import shutil
        mapping_path, _ = sample_mapping
        shutil.copy(sample_fuel_excel, tmp_path / "Fuel.xlsx")

        results = sync(tmp_path, mode="database", data_types=["fuel"], dry_run=True, mapping_file=mapping_path)
        assert "fuel" in results
        assert results["fuel"]["success"] == 0


# ---------------------------------------------------------------------------
# config_loader MineBase 配置
# ---------------------------------------------------------------------------


class TestMineBaseConfig:
    def test_get_minebase_config_defaults(self, monkeypatch):
        from func import config_loader
        monkeypatch.setattr(config_loader, "_CONFIG_FILE", pathlib.Path("/nonexistent"))
        monkeypatch.setattr(config_loader, "_USER_CONFIG_FILE", pathlib.Path("/nonexistent"))
        cfg = config_loader.get_minebase_config()
        assert cfg["mode"] == "api"
        assert cfg["api"]["url"] == "http://localhost:3000"
        assert cfg["database"]["port"] == 5432

    def test_get_minebase_mode_default(self, monkeypatch):
        from func import config_loader
        monkeypatch.setattr(config_loader, "_CONFIG_FILE", pathlib.Path("/nonexistent"))
        monkeypatch.setattr(config_loader, "_USER_CONFIG_FILE", pathlib.Path("/nonexistent"))
        assert config_loader.get_minebase_mode() == "api"


# ---------------------------------------------------------------------------
# discover_files glob vs keyword priority
# ---------------------------------------------------------------------------


class TestDiscoverFilesGlobPriority:
    def test_discover_files_prefers_glob_over_keywords(self, tmp_path):
        """当 processed output 和 raw input 同时存在时，discover_files 选择 processed output。"""
        # 创建 processed output
        (tmp_path / "Fuel.xlsx").write_bytes(b"")
        # 创建 raw input（关键字匹配能命中）
        (tmp_path / "Fuel report Normount 6-2026.xlsx").write_bytes(b"")

        found = discover_files(tmp_path, keywords={"fuel": ["Fuel report"]})
        assert "fuel" in found
        assert found["fuel"][0].name == "Fuel.xlsx"

    def test_discover_files_glob_finds_processed_outputs(self, tmp_path):
        """验证 discover_files 通过 DATA_TYPE_REGISTRY pattern 找到所有 processed output。"""
        (tmp_path / "Fuel.xlsx").write_bytes(b"")
        (tmp_path / "电力消耗统计.xlsx").write_bytes(b"")
        (tmp_path / "合并产量.xlsx").write_bytes(b"")
        (tmp_path / "202506_工作效率表.xlsx").write_bytes(b"")

        found = discover_files(tmp_path, year=2025, month=6, keywords={})
        assert found["fuel"][0].name == "Fuel.xlsx"
        assert found["electrical"][0].name == "电力消耗统计.xlsx"
        assert found["production"][0].name == "合并产量.xlsx"
        assert found["operation"][0].name == "合并产量.xlsx"
        assert found["work_efficiency"][0].name == "202506_工作效率表.xlsx"

    def test_discover_files_keyword_fallback_when_no_glob_match(self, tmp_path):
        """当不存在 processed output 时，关键字匹配作为回退。"""
        # 只有 raw input，没有 Fuel.xlsx
        (tmp_path / "Fuel report Normount 6-2026.xlsx").write_bytes(b"")

        found = discover_files(tmp_path, keywords={"fuel": ["Fuel report"]})
        assert "fuel" in found
        assert found["fuel"][0].name == "Fuel report Normount 6-2026.xlsx"

    def test_discover_files_mixed_directory_priority(self, tmp_path):
        """全类型混合目录：所有 data type 都应优先解析到 processed output。"""
        # processed outputs
        (tmp_path / "Fuel.xlsx").write_bytes(b"")
        (tmp_path / "电力消耗统计.xlsx").write_bytes(b"")
        (tmp_path / "合并产量.xlsx").write_bytes(b"")
        (tmp_path / "202506_工作效率表.xlsx").write_bytes(b"")

        # raw inputs (关键字匹配可能命中)
        (tmp_path / "Fuel report Normount 6-2026.xlsx").write_bytes(b"")
        (tmp_path / "Цахилгааны зарцуулалт.xlsx").write_bytes(b"")
        (tmp_path / "夜班日报_2025-06.xlsx").write_bytes(b"")

        keywords = {
            "fuel": ["Fuel report"],
            "electrical": ["Цахилгааны"],
            "production": ["夜班日报"],
        }
        found = discover_files(tmp_path, year=2025, month=6, keywords=keywords)

        assert found["fuel"][0].name == "Fuel.xlsx"
        assert found["electrical"][0].name == "电力消耗统计.xlsx"
        assert found["production"][0].name == "合并产量.xlsx"
        assert found["operation"][0].name == "合并产量.xlsx"
        assert "202506_工作效率表" in found["work_efficiency"][0].name


# ---------------------------------------------------------------------------
# work_efficiency mapping: date + shiftType
# ---------------------------------------------------------------------------


class TestWorkEfficiencyMapping:
    def test_work_efficiency_mapping_includes_date_and_shift(self):
        """默认列映射中 work_efficiency 应包含 日期->date 和 班次->shiftType。"""
        mapping = load_column_mapping()
        we_mapping = mapping.get("work_efficiency") or mapping.get("worktime", {})
        assert we_mapping, "work_efficiency mapping not found"
        assert we_mapping["日期"] == "date"
        assert we_mapping["班次"] == "shiftType"

    def test_read_and_map_work_efficiency_with_date_shift(self, tmp_path):
        """read_and_map_excel 正确映射 work_efficiency 的 日期、班次、设备名称、应运行分钟。"""
        mapping = load_column_mapping()
        we_mapping = mapping.get("work_efficiency") or mapping.get("worktime", {})

        df = pd.DataFrame({
            "日期": ["2025-06-01", "2025-06-02"],
            "班次": ["Day", "Night"],
            "设备名称": ["CAT785D-01", "NTE240-02"],
            "应运行分钟": [720, 600],
        })
        path = tmp_path / "test_work_efficiency.xlsx"
        df.to_excel(path, index=False)

        rows = read_and_map_excel(path, None, we_mapping)
        assert len(rows) == 2
        assert rows[0]["date"] == "2025-06-01"
        assert rows[0]["shiftType"] == "Day"
        assert rows[0]["equipmentName"] == "CAT785D-01"
        assert rows[0]["plannedMinutes"] == 720
        assert rows[1]["date"] == "2025-06-02"
        assert rows[1]["shiftType"] == "Night"


# ---------------------------------------------------------------------------
# _df_to_mapped_rows
# ---------------------------------------------------------------------------


class TestDfToMappedRows:
    def test_basic_mapping(self):
        """df with 日期/班次/设备名称 -> mapped to date/shiftType/equipmentName."""
        df = pd.DataFrame({
            "日期": ["2025-06-01", "2025-06-02"],
            "班次": ["Day", "Night"],
            "设备名称": ["CAT785D-01", "NTE240-02"],
        })
        mapping = {"日期": "date", "班次": "shiftType", "设备名称": "equipmentName"}
        rows = _df_to_mapped_rows(df, mapping)
        assert len(rows) == 2
        assert rows[0] == {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01"}
        assert rows[1] == {"date": "2025-06-02", "shiftType": "Night", "equipmentName": "NTE240-02"}

    def test_nan_handling(self):
        """NaN values skipped in output."""
        df = pd.DataFrame({
            "日期": ["2025-06-01"],
            "班次": ["Day"],
            "设备名称": ["CAT785D-01"],
            "油品消耗": [float("nan")],
        })
        mapping = {"日期": "date", "班次": "shiftType", "设备名称": "equipmentName", "油品消耗": "consumption"}
        rows = _df_to_mapped_rows(df, mapping)
        assert len(rows) == 1
        assert "consumption" not in rows[0]
        assert rows[0]["date"] == "2025-06-01"

    def test_date_conversion(self):
        """pd.Timestamp -> YYYY-MM-DD string."""
        df = pd.DataFrame({
            "日期": [pd.Timestamp("2025-06-15"), pd.Timestamp("2025-07-01")],
            "班次": ["Day", "Night"],
        })
        mapping = {"日期": "date", "班次": "shiftType"}
        rows = _df_to_mapped_rows(df, mapping)
        assert rows[0]["date"] == "2025-06-15"
        assert rows[1]["date"] == "2025-07-01"

    def test_empty_dataframe(self):
        """Empty df returns empty list."""
        df = pd.DataFrame(columns=["日期", "班次", "设备名称"])
        mapping = {"日期": "date", "班次": "shiftType", "设备名称": "equipmentName"}
        rows = _df_to_mapped_rows(df, mapping)
        assert rows == []


# ---------------------------------------------------------------------------
# _process_fuel_file
# ---------------------------------------------------------------------------


class TestProcessFuelFile:
    @patch("func.sync_to_minebase._df_to_mapped_rows")
    @patch("func.excel_fuel.process_diesel_data")
    def test_process_fuel_returns_mapped_rows(self, mock_diesel, mock_df_map):
        """patch process_diesel_data to return {'油耗信息': df}, verify _process_fuel_file returns correctly mapped rows."""
        from func.sync_to_minebase import _process_fuel_file

        test_df = pd.DataFrame({"col": [1]})
        mock_diesel.return_value = {"油耗信息": test_df}
        expected_rows = [
            {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "fuelName": "0#柴油", "consumption": 150.5},
        ]
        mock_df_map.return_value = expected_rows

        result = _process_fuel_file(Path("/fake/Fuel.xlsx"), year=2025)
        assert result == expected_rows
        mock_diesel.assert_called_once()
        mock_df_map.assert_called_once()

    @patch("func.excel_fuel.process_diesel_data", side_effect=RuntimeError("boom"))
    def test_process_fuel_handles_error(self, mock_diesel):
        """patch to raise exception, verify returns empty list."""
        from func.sync_to_minebase import _process_fuel_file

        result = _process_fuel_file(Path("/fake/Fuel.xlsx"), year=2025)
        assert result == []


# ---------------------------------------------------------------------------
# _process_electrical_file
# ---------------------------------------------------------------------------


class TestProcessElectricalFile:
    @patch("func.sync_to_minebase._df_to_mapped_rows")
    @patch("func.excel_electrical.parse_excel_data")
    def test_process_electrical_returns_mapped_rows(self, mock_parse, mock_df_map):
        """patch parse_excel_data to return {'电力消耗': df}, verify mapping."""
        from func.sync_to_minebase import _process_electrical_file

        test_df = pd.DataFrame({"col": [1]})
        mock_parse.return_value = {"电力消耗": test_df}
        expected_rows = [
            {"date": "2025-06-01", "shiftType": "Night", "equipmentName": "EX-001", "consumption": 500.0},
        ]
        mock_df_map.return_value = expected_rows

        result = _process_electrical_file(Path("/fake/电力消耗统计.xlsx"), year=2025)
        assert result == expected_rows
        mock_parse.assert_called_once()
        mock_df_map.assert_called_once()

    @patch("func.excel_electrical.parse_excel_data", return_value={})
    def test_process_electrical_handles_empty(self, mock_parse):
        """patch to return empty dict, verify returns empty list."""
        from func.sync_to_minebase import _process_electrical_file

        result = _process_electrical_file(Path("/fake/电力消耗统计.xlsx"), year=2025)
        assert result == []


# ---------------------------------------------------------------------------
# _process_production_file
# ---------------------------------------------------------------------------


class TestProcessProductionFile:
    @patch("func.sync_to_minebase._df_to_mapped_rows")
    @patch("func.excel_production_enhanced.MiningDataProcessor")
    def test_process_production_returns_both_types(self, mock_processor_cls, mock_df_map):
        """patch MiningDataProcessor.process_file, verify production and operation rows."""
        from func.sync_to_minebase import _process_production_file

        running_df = pd.DataFrame({"col": [1]})
        production_df = pd.DataFrame({"col": [2]})
        mock_processor_cls.return_value.process_single_file.return_value = (running_df, production_df)

        prod_rows = [{"date": "2025-06-01", "truckName": "CAT785D-01"}]
        ops_rows = [{"date": "2025-06-01", "equipmentName": "CAT785D-01"}]
        mock_df_map.side_effect = [prod_rows, ops_rows]

        result = _process_production_file(Path("/fake/合并产量.xlsx"))
        assert result == {"production": prod_rows, "operation": ops_rows}
        mock_processor_cls.return_value.process_single_file.assert_called_once()

    @patch("func.sync_to_minebase._df_to_mapped_rows")
    @patch("func.excel_production_enhanced.MiningDataProcessor")
    def test_process_production_handles_error(self, mock_processor_cls, mock_df_map):
        """patch to raise, verify returns empty dicts."""
        from func.sync_to_minebase import _process_production_file

        mock_processor_cls.return_value.process_single_file.side_effect = RuntimeError("boom")

        result = _process_production_file(Path("/fake/合并产量.xlsx"))
        assert result == {"production": [], "operation": []}


# ---------------------------------------------------------------------------
# TestSyncWithProcessors
# ---------------------------------------------------------------------------


class TestSyncWithProcessors:
    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    def test_sync_fuel_via_processor(self, mock_discover, mock_fuel_proc, mock_sync_api, tmp_path):
        """full integration - patch processor AND api client, verify sync() sends fuel data."""
        mock_discover.return_value = {"fuel": [tmp_path / "Fuel.xlsx"]}
        mock_fuel_proc.return_value = [
            {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "fuelName": "0#柴油", "consumption": 150.5},
        ]
        mock_sync_api.return_value = {"success": 1, "skipped": 0, "failed": 0, "warnings": []}

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            results = sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True)

        assert "fuel" in results
        assert results["fuel"]["success"] == 1
        mock_fuel_proc.assert_called_once()

    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_electrical_file")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    def test_sync_processor_failure_continues(self, mock_discover, mock_fuel_proc, mock_elec_proc, mock_sync_api, tmp_path):
        """one type fails, others still sync."""
        mock_discover.return_value = {
            "fuel": [tmp_path / "Fuel.xlsx"],
            "electrical": [tmp_path / "电力消耗统计.xlsx"],
        }
        mock_fuel_proc.side_effect = RuntimeError("boom")
        mock_elec_proc.return_value = [
            {"date": "2025-06-01", "shiftType": "Night", "equipmentName": "EX-001", "consumption": 500.0},
        ]
        mock_sync_api.return_value = {"success": 1, "skipped": 0, "failed": 0, "warnings": []}

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            results = sync(tmp_path, mode="api", data_types=["fuel", "electrical"], dry_run=True)

        # fuel fails and gets empty result, electrical still syncs
        assert results["fuel"] == {"success": 0, "skipped": 0, "failed": 0, "warnings": []}
        assert results["electrical"]["success"] == 1


# ---------------------------------------------------------------------------
# TestResolveFksForDb — data_type routing
# ---------------------------------------------------------------------------


class TestResolveFksForDb:
    """_resolve_fks_for_db must route production rows through the
    truck/excavator/material branch, not the generic equipment branch."""

    def test_production_row_uses_truck_fk_branch(self):
        """production rows with truckName/excavatorName should resolve via
        the production_record branch (truckId, excavatorId), NOT be rejected
        with '缺少设备名称'."""
        mock_db = MagicMock()
        mock_db.resolve_equipment_id.return_value = "equip-uuid-1"
        mock_db.resolve_material_type_id.return_value = "mat-uuid-1"

        row = {
            "date": "2025-06-15",
            "shiftType": "夜班",
            "truckName": "CAT785D-01",
            "excavatorName": "EX-001",
            "materialTypeName": "铜矿",
            "tripCount": 10,
            "production": 350.0,
        }

        # sync() passes "production" (the registry key), not "production_record" (the table name)
        result = _resolve_fks_for_db("production", row, mock_db)

        # Should NOT be None — production rows must be accepted
        assert result is not None
        assert result["truckId"] == "equip-uuid-1"
        assert result["excavatorId"] == "equip-uuid-1"
        assert result["materialTypeId"] == "mat-uuid-1"
        assert result["truckName"] == "CAT785D-01"

    def test_operation_row_uses_equipment_fk_branch(self):
        """operation rows with equipmentName should resolve via the generic
        equipment branch (equipmentId)."""
        mock_db = MagicMock()
        mock_db.resolve_equipment_id.return_value = "equip-uuid-2"

        row = {
            "date": "2025-06-15",
            "shiftType": "夜班",
            "equipmentName": "CAT785D-01",
            "tripCount": 10,
        }

        result = _resolve_fks_for_db("operation", row, mock_db)

        assert result is not None
        assert result["equipmentId"] == "equip-uuid-2"

    def test_production_row_missing_truck_returns_none(self):
        """production row without truckName should be rejected."""
        mock_db = MagicMock()
        row = {
            "date": "2025-06-15",
            "shiftType": "夜班",
            "excavatorName": "EX-001",
        }

        result = _resolve_fks_for_db("production", row, mock_db)
        assert result is None

    def test_production_missing_truck_warning_value_shows_empty_placeholder(self):
        """production row missing truckName should emit warning with '（空）' as value."""
        mock_db = MagicMock()
        warnings = []
        row = {
            "_row_num": 5,
            "date": "2025-06-15",
            "shiftType": "夜班",
            "excavatorName": "EX-001",
        }

        _resolve_fks_for_db("production", row, mock_db, warnings=warnings)

        assert len(warnings) == 1
        assert warnings[0]["field"] == "truckName"
        assert warnings[0]["value"] == "（空）"
        assert warnings[0]["row"] == 5

    def test_production_missing_excavator_warning_value_shows_empty_placeholder(self):
        """production row missing excavatorName should emit warning with '（空）'."""
        mock_db = MagicMock()
        mock_db.resolve_equipment_id.return_value = "truck-uuid"
        warnings = []
        row = {
            "_row_num": 8,
            "date": "2025-06-15",
            "shiftType": "夜班",
            "truckName": "CAT785D-01",
        }

        _resolve_fks_for_db("production", row, mock_db, warnings=warnings)

        assert len(warnings) == 1
        assert warnings[0]["field"] == "excavatorName"
        assert warnings[0]["value"] == "（空）"

    def test_generic_missing_equipment_warning_value_shows_empty_placeholder(self):
        """generic row missing equipmentName should emit warning with '（空）'."""
        mock_db = MagicMock()
        warnings = []
        row = {
            "_row_num": 12,
            "date": "2025-06-15",
            "shiftType": "Day",
        }

        _resolve_fks_for_db("fuel", row, mock_db, warnings=warnings)

        assert len(warnings) == 1
        assert warnings[0]["field"] == "equipmentName"
        assert warnings[0]["value"] == "（空）"
        assert warnings[0]["row"] == 12


class TestLedgerToggleSplit:
    """sync() 应支持 use_equipment_ledger 和 use_oil_ledger 独立控制。"""

    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    @patch("func.config_loader.load_oil_ledger_cache")
    @patch("func.config_loader.load_equipment_ledger_cache")
    def test_use_equipment_ledger_true_loads_only_equipment(
        self, mock_eq_cache, mock_oil_cache, mock_discover, mock_fuel, mock_api, tmp_path,
    ):
        """use_equipment_ledger=True 应只加载设备台账，油品台账为 None。"""
        mock_discover.return_value = {"fuel": [tmp_path / "Fuel.xlsx"]}
        mock_fuel.return_value = [{"date": "2025-06-01", "equipmentName": "CAT785D-01"}]
        mock_api.return_value = {"success": 1, "skipped": 0, "failed": 0, "warnings": []}
        mock_eq_cache.return_value = [{"标准名称": "CAT785D", "别名": ["卡特785"]}]
        mock_oil_cache.return_value = None

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True,
                 use_equipment_ledger=True, use_oil_ledger=False)

        mock_eq_cache.assert_called_once()
        mock_oil_cache.assert_not_called()

    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    @patch("func.config_loader.load_oil_ledger_cache")
    @patch("func.config_loader.load_equipment_ledger_cache")
    def test_use_oil_ledger_true_loads_only_oil(
        self, mock_eq_cache, mock_oil_cache, mock_discover, mock_fuel, mock_api, tmp_path,
    ):
        """use_oil_ledger=True 应只加载油品台账，设备台账为 None。"""
        mock_discover.return_value = {"fuel": [tmp_path / "Fuel.xlsx"]}
        mock_fuel.return_value = [{"date": "2025-06-01", "fuelName": "0#柴油"}]
        mock_api.return_value = {"success": 1, "skipped": 0, "failed": 0, "warnings": []}
        mock_eq_cache.return_value = None
        mock_oil_cache.return_value = [{"标准名称": "0号柴油", "编码": "OIL-001"}]

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True,
                 use_equipment_ledger=False, use_oil_ledger=True)

        mock_eq_cache.assert_not_called()
        mock_oil_cache.assert_called_once()

    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    @patch("func.config_loader.load_oil_ledger_cache")
    @patch("func.config_loader.load_equipment_ledger_cache")
    def test_use_ledger_true_backward_compat(
        self, mock_eq_cache, mock_oil_cache, mock_discover, mock_fuel, mock_api, tmp_path,
    ):
        """use_ledger=True 应同时启用设备和油品台账（向后兼容）。"""
        mock_discover.return_value = {"fuel": [tmp_path / "Fuel.xlsx"]}
        mock_fuel.return_value = [{"date": "2025-06-01"}]
        mock_api.return_value = {"success": 1, "skipped": 0, "failed": 0, "warnings": []}
        mock_eq_cache.return_value = [{"标准名称": "CAT785D"}]
        mock_oil_cache.return_value = [{"标准名称": "0号柴油"}]

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True, use_ledger=True)

        mock_eq_cache.assert_called_once()
        mock_oil_cache.assert_called_once()

    def test_default_oil_ledger_on_equipment_ledger_off(self):
        """默认值: use_equipment_ledger=False, use_oil_ledger=True。"""
        import inspect
        sig = inspect.signature(sync)
        assert sig.parameters["use_equipment_ledger"].default is False
        assert sig.parameters["use_oil_ledger"].default is True


# ---------------------------------------------------------------------------
# sync_via_api 值解析与空值兜底
# ---------------------------------------------------------------------------


class TestSyncApiValueFallback:
    """sync_via_api 应正确解析 API 返回中的 value 字段，空值显示为 '（空）'。"""

    def _make_api_client(self, warnings=None, errors=None):
        """创建 mock API 客户端，模拟 contractVersion 2 的 send_batch/confirm 响应。"""
        client = MagicMock()
        client.create_session.return_value = {
            "data": {"session": {"id": "session-1", "version": 1}},
        }
        batch_result = {"success": 1, "skipped": 0, "failed": 0}
        if warnings:
            batch_result["warnings"] = warnings
        if errors:
            batch_result["errors"] = errors
        client.send_batch.return_value = {
            "data": {
                "receipt": {"result": batch_result},
                "session": {"version": 1},
            },
        }
        client.confirm_batch.return_value = {
            "data": {
                "done": True,
                "inserted": 1,
                "updated": 0,
                "skipped": 0,
                "rejected": 0,
                "session": {"version": 2},
            },
        }
        return client

    def test_api_warning_value_from_standard_key(self):
        """API 返回的 warning.value 有值时正常传递。"""
        client = self._make_api_client(warnings=[
            {"row": 3, "field": "equipmentName", "value": "旧设备", "message": "未匹配"},
        ])
        result = sync_via_api("fuel", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "旧设备"

    def test_api_warning_value_from_rawvalue_fallback(self):
        """API 返回 warning 时，若 value 为空但有 rawValue，应取 rawValue。"""
        client = self._make_api_client(warnings=[
            {"row": 3, "field": "fuelName", "value": "", "rawValue": "柴油-10", "message": "未匹配"},
        ])
        result = sync_via_api("fuel", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "柴油-10"

    def test_api_warning_value_from_raw_value_fallback(self):
        """API 返回 warning 时，若 value 为空但有 raw_value，应取 raw_value。"""
        client = self._make_api_client(warnings=[
            {"row": 3, "field": "fuelName", "value": "", "raw_value": "柴油-10", "message": "未匹配"},
        ])
        result = sync_via_api("fuel", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "柴油-10"

    def test_api_warning_value_all_empty_shows_placeholder(self):
        """API 返回 warning 的 value 与 rawValue 均为空时，显示 '（空）'。"""
        client = self._make_api_client(warnings=[
            {"row": 3, "field": "equipmentName", "value": "", "rawValue": "", "message": "缺少设备"},
        ])
        result = sync_via_api("fuel", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "（空）"

    def test_api_error_value_none_shows_placeholder(self):
        """API 返回 error 的 value 为 None 时，显示 '（空）'。"""
        client = self._make_api_client(errors=[
            {"row": 1, "field": "truckName", "value": None, "message": "缺少矿卡"},
        ])
        result = sync_via_api("production", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "（空）"

    def test_api_error_value_from_originalvalue_fallback(self):
        """API 返回 error 时，若 value 为空但有 originalValue，应取 originalValue。"""
        client = self._make_api_client(errors=[
            {"row": 2, "field": "consumption", "value": "", "originalValue": "99999", "message": "超出范围"},
        ])
        result = sync_via_api("electrical", [{"date": "2025-01-01"}], {"日期": "date"}, client)
        assert len(result["warnings"]) == 1
        assert result["warnings"][0]["value"] == "99999"


# ---------------------------------------------------------------------------
# export_dry_run_to_excel
# ---------------------------------------------------------------------------


class TestExportDryRunToExcel:
    """试运行预览导出功能测试。"""

    def _make_column_mapping(self):
        """构建模拟的列映射配置。"""
        return {
            "fuel": {
                "日期": "date",
                "班次": "shiftType",
                "设备名称": "equipmentName",
                "油品种类": "fuelName",
                "油品消耗": "consumption",
            },
            "electrical": {
                "日期": "date",
                "设备名称": "equipmentName",
                "电力消耗": "consumption",
            },
        }

    def test_export_creates_multi_sheet_excel(self, tmp_path):
        """多个数据类型应各自生成独立 sheet。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {
            "fuel": [
                {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "consumption": 150.5},
                {"date": "2025-06-02", "shiftType": "Night", "equipmentName": "CAT785D-02", "consumption": 200.0},
            ],
            "electrical": [
                {"date": "2025-06-01", "equipmentName": "EX-001", "consumption": 500.0},
            ],
        }
        out_path = export_dry_run_to_excel(
            dry_run_rows, column_mapping=self._make_column_mapping(), input_dir=tmp_path,
        )
        assert Path(out_path).exists()

        # 读取 sheet 名称
        xls = pd.ExcelFile(out_path)
        assert "汇总" in xls.sheet_names
        assert "油耗数据" in xls.sheet_names
        assert "电耗数据" in xls.sheet_names

    def test_export_summary_sheet_counts(self, tmp_path):
        """汇总 sheet 应包含各数据类型的记录数。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {
            "fuel": [{"date": "2025-06-01"}, {"date": "2025-06-02"}],
            "production": [{"date": "2025-06-01"}],
        }
        out_path = export_dry_run_to_excel(dry_run_rows, input_dir=tmp_path)

        df = pd.read_excel(out_path, sheet_name="汇总")
        fuel_row = df[df["数据类型"] == "油耗数据"].iloc[0]
        assert fuel_row["记录数"] == 2
        prod_row = df[df["数据类型"] == "生产数据"].iloc[0]
        assert prod_row["记录数"] == 1

    def test_export_removes_internal_keys(self, tmp_path):
        """导出时应剥离 _row_num 等内部元数据列。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {
            "fuel": [
                {"date": "2025-06-01", "equipmentName": "CAT785D", "_row_num": 2},
            ],
        }
        out_path = export_dry_run_to_excel(dry_run_rows, input_dir=tmp_path)

        # 用 openpyxl 读取，第二行是数据库字段名，第三行起是数据
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb["油耗数据"]
        # 第一行：源列名，第二行：数据库字段名
        row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        assert "_row_num" not in row1
        assert "_row_num" not in row2
        wb.close()

    def test_export_dual_headers_with_mapping(self, tmp_path):
        """有列映射时应生成双重表头：第一行源列名，第二行字段名。"""
        from func.sync.export import export_dry_run_to_excel
        from openpyxl import load_workbook

        dry_run_rows = {
            "fuel": [
                {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "consumption": 150.5},
            ],
        }
        out_path = export_dry_run_to_excel(
            dry_run_rows, column_mapping=self._make_column_mapping(), input_dir=tmp_path,
        )

        wb = load_workbook(out_path)
        ws = wb["油耗数据"]
        row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        row2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        wb.close()

        # 第一行是源列名（中文）
        assert "日期" in row1
        assert "班次" in row1
        assert "设备名称" in row1
        assert "油品消耗" in row1
        # 第二行是数据库字段名（英文）
        assert "date" in row2
        assert "shiftType" in row2
        assert "equipmentName" in row2
        assert "consumption" in row2

    def test_export_no_dual_headers_without_mapping(self, tmp_path):
        """无列映射时不应插入第二行表头（只有一行字段名）。"""
        from func.sync.export import export_dry_run_to_excel
        from openpyxl import load_workbook

        dry_run_rows = {
            "fuel": [
                {"date": "2025-06-01", "consumption": 150.5},
            ],
        }
        out_path = export_dry_run_to_excel(dry_run_rows, input_dir=tmp_path)

        wb = load_workbook(out_path)
        ws = wb["油耗数据"]
        assert ws.max_row == 2  # 1 header + 1 data row（无第二行表头）
        row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert row1 == ["date", "consumption"]
        wb.close()

    def test_export_with_warnings(self, tmp_path):
        """有警告时应额外生成异常行 sheet。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {"fuel": [{"date": "2025-06-01"}]}
        warnings = [
            ("fuel", {"row": 2, "field": "equipmentName", "value": "未知设备", "message": "未在台账中找到"}),
        ]
        out_path = export_dry_run_to_excel(dry_run_rows, warnings=warnings, input_dir=tmp_path)

        xls = pd.ExcelFile(out_path)
        assert "异常行" in xls.sheet_names
        df = pd.read_excel(out_path, sheet_name="异常行")
        assert len(df) == 1
        assert df.iloc[0]["字段"] == "equipmentName"

    def test_export_no_warnings_no_sheet(self, tmp_path):
        """无警告时不应生成异常行 sheet。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {"fuel": [{"date": "2025-06-01"}]}
        out_path = export_dry_run_to_excel(dry_run_rows, input_dir=tmp_path)

        xls = pd.ExcelFile(out_path)
        assert "异常行" not in xls.sheet_names

    def test_export_empty_rows(self, tmp_path):
        """空数据类型应生成空 sheet。"""
        from func.sync.export import export_dry_run_to_excel

        dry_run_rows = {"fuel": []}
        out_path = export_dry_run_to_excel(dry_run_rows, input_dir=tmp_path)
        assert Path(out_path).exists()

        df = pd.read_excel(out_path, sheet_name="油耗数据")
        assert len(df) == 0


# ---------------------------------------------------------------------------
# sync_engines dry_run returns rows
# ---------------------------------------------------------------------------


class TestDryRunEnginesReturnRows:
    """sync_via_api / sync_via_db 在 dry_run 模式应返回 dry_run_rows。"""

    def test_sync_via_api_dry_run_returns_rows(self):
        """API 模式 dry_run 应返回 dry_run_rows 且不调用 API。"""
        rows = [
            {"date": "2025-06-01", "equipmentName": "CAT785D-01"},
            {"date": "2025-06-02", "equipmentName": "CAT785D-02"},
        ]
        client = MagicMock()
        result = sync_via_api("fuel", rows, {}, client, dry_run=True)

        assert result["success"] == 0
        assert "dry_run_rows" in result
        assert len(result["dry_run_rows"]) == 2
        client.login.assert_not_called()
        client.create_session.assert_not_called()

    def test_sync_via_api_normal_no_dry_run_rows(self):
        """正常模式不应返回 dry_run_rows。"""
        client = MagicMock()
        client.create_session.return_value = {
            "data": {"session": {"id": "session-1", "version": 1}},
        }
        client.send_batch.return_value = {
            "data": {
                "receipt": {"result": {"success": 1, "skipped": 0, "failed": 0}},
                "session": {"version": 1},
            },
        }
        client.confirm_batch.return_value = {
            "data": {
                "done": True,
                "inserted": 1,
                "updated": 0,
                "skipped": 0,
                "rejected": 0,
                "session": {"version": 2},
            },
        }

        result = sync_via_api("fuel", [{"date": "2025-06-01"}], {"日期": "date"}, client, dry_run=False)
        assert "dry_run_rows" not in result


# ---------------------------------------------------------------------------
# core.sync dry_run integration
# ---------------------------------------------------------------------------


class TestDryRunIntegration:
    """sync() dry_run 模式应生成预览 Excel。"""

    @patch("func.sync_to_minebase.sync_via_api")
    @patch("func.sync_to_minebase._process_fuel_file")
    @patch("func.sync_to_minebase.discover_files")
    def test_dry_run_generates_excel_file(self, mock_discover, mock_fuel_proc, mock_sync_api, tmp_path, sample_mapping):
        """dry_run 应生成预览 Excel 并将路径写入 results['_dry_run_file']。"""
        mapping_path, _ = sample_mapping
        mock_discover.return_value = {"fuel": [tmp_path / "Fuel.xlsx"]}
        mock_fuel_proc.return_value = [
            {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "fuelName": "0#柴油", "consumption": 150.5},
            {"date": "2025-06-02", "shiftType": "Night", "equipmentName": "CAT785D-02", "fuelName": "0#柴油", "consumption": 200.0},
        ]
        # sync_via_api 在 dry_run 模式下会被实际调用（未 mock 穿透），
        # 但因为我们用了 @patch，返回值需要包含 dry_run_rows
        mock_sync_api.return_value = {
            "success": 0, "skipped": 0, "failed": 0, "warnings": [],
            "dry_run_rows": [
                {"date": "2025-06-01", "shiftType": "Day", "equipmentName": "CAT785D-01", "fuelName": "0#柴油", "consumption": 150.5},
                {"date": "2025-06-02", "shiftType": "Night", "equipmentName": "CAT785D-02", "fuelName": "0#柴油", "consumption": 200.0},
            ],
        }

        with patch("func.sync_to_minebase.MineBaseAPIClient") as mock_api_cls, \
             patch("func.sync_to_minebase.get_minebase_api_config", return_value={"url": "http://test", "username": "u", "password": "p"}):
            mock_api_cls.return_value = MagicMock()
            results = sync(tmp_path, mode="api", data_types=["fuel"], dry_run=True, mapping_file=mapping_path)

        assert "_dry_run_file" in results
        assert Path(results["_dry_run_file"]).exists()
        # 验证结果中不再包含 dry_run_rows
        assert "dry_run_rows" not in results.get("fuel", {})
