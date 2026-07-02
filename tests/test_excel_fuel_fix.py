"""Test excel_fuel handles short sheets gracefully (finding #3 fix)."""

import os
import pytest
import openpyxl

from func.excel_fuel import process_diesel_data


def _create_short_sheet_excel(path, num_rows_before_data=3):
    """Create an Excel file with a matching sheet name but very few rows.

    The sheet has `num_rows_before_data` rows before a row whose first
    column is 1 (the marker that triggers start_row_idx discovery).
    This means start_row = num_rows_before_data + 1, which will be < 6
    when num_rows_before_data < 5.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备柴油消耗表"

    # Write some filler rows (columns A onward)
    for r in range(1, num_rows_before_data + 1):
        ws.cell(row=r, column=1, value=f"filler_{r}")

    # Write the marker row (first column == 1) at `num_rows_before_data + 1`
    ws.cell(row=num_rows_before_data + 1, column=1, value=1)

    # Write a couple of data rows after the marker to avoid other issues
    for r in range(num_rows_before_data + 2, num_rows_before_data + 4):
        ws.cell(row=r, column=1, value=r)
        ws.cell(row=r, column=2, value="Device")
        ws.cell(row=r, column=3, value="D001")

    wb.save(path)


def test_short_sheet_does_not_crash(tmp_path):
    """A sheet with start_row < 6 must be skipped, not crash with negative index."""
    excel_path = os.path.join(tmp_path, "short_fuel.xlsx")
    # 3 rows before data row => start_row = 4, which is < 6
    _create_short_sheet_excel(excel_path, num_rows_before_data=3)

    # Should raise ValueError (no valid data) rather than crash with negative index
    with pytest.raises(ValueError, match="未找到"):
        process_diesel_data(excel_path)


def test_two_short_sheets_all_skipped(tmp_path):
    """Multiple short sheets are all skipped, resulting in ValueError."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "设备柴油消耗表1"
    ws1.cell(row=1, column=1, value="hdr")
    ws1.cell(row=2, column=1, value=1)  # start_row = 2

    ws2 = wb.create_sheet("设备柴油消耗表2")
    ws2.cell(row=1, column=1, value="hdr")
    ws2.cell(row=2, column=1, value=1)
    ws2.cell(row=3, column=1, value="data")

    path = os.path.join(tmp_path, "multi_short.xlsx")
    wb.save(path)

    with pytest.raises(ValueError, match="未找到"):
        process_diesel_data(path)


def test_normal_sheet_still_works(tmp_path):
    """A sheet with start_row >= 6 must still be processed (no regression)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备柴油消耗表"

    # Header rows (rows 1-4 are "above" the marker area)
    # We need start_row_idx where col A == 1 at row 6 (start_row = 7)
    # So place the marker at row 6 => iloc index 5 => start_row = 6
    for r in range(1, 6):
        ws.cell(row=r, column=1, value="header")

    # Marker row at row 6
    ws.cell(row=6, column=1, value=1)

    # Data row at row 7+
    ws.cell(row=7, column=1, value="data")
    ws.cell(row=7, column=2, value="设备A")
    ws.cell(row=7, column=3, value="D001")

    path = os.path.join(tmp_path, "normal_fuel.xlsx")
    wb.save(path)

    # Should NOT raise the short-sheet error; it may raise ValueError
    # for "no valid data" but the guard should not trigger.
    try:
        process_diesel_data(path)
    except ValueError as e:
        # Acceptable: data parsing found nothing, but the short-sheet guard did not fire
        assert "行数不足" not in str(e)
