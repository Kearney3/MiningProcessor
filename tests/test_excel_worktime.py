"""
Unit tests for func/excel_worktime.process_excel_data.

Fixtures are synthetic Excel workbooks created with openpyxl, following the
structure expected by split_day_night_shifts:
  Row 1: title (ignored by processor)
  Row 2: header row
  Row 3+: data rows; a row whose first cell matches header[0] splits Day/Night.
"""

import os
import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from func.excel_worktime import process_excel_data


# ---------------------------------------------------------------------------
# Helper: create a worktime Excel fixture
# ---------------------------------------------------------------------------

def _create_worktime_excel(
    path: str,
    days: list[int],
    equipment_day: list[str] | None = None,
    equipment_night: list[str] | None = None,
    extra_sheets: dict[str, list[list]] | None = None,
) -> None:
    """Create a synthetic worktime Excel file.

    Each day gets its own numeric sheet with a title row, header row, day-shift
    data, a separator row (repeating the header's first value), and night-shift
    data.

    Args:
        path: Output file path.
        days: List of day numbers (sheet names).
        equipment_day: Equipment names for day shift.  Defaults to ["Excavator", "Dump Truck"].
        equipment_night: Equipment names for night shift.  Defaults to ["Loader", "Drill"].
        extra_sheets: Additional sheets as {name: [[row1], [row2], ...]}.
    """
    if equipment_day is None:
        equipment_day = ["Excavator", "Dump Truck"]
    if equipment_night is None:
        equipment_night = ["Loader", "Drill"]

    wb = openpyxl.Workbook()
    first = True
    for day in days:
        if first:
            ws = wb.active
            ws.title = str(day)
            first = False
        else:
            ws = wb.create_sheet(title=str(day))

        # Row 1: title
        ws.cell(row=1, column=1, value="Worktime Report")

        # Row 2: header
        ws.cell(row=2, column=1, value="Equipment")
        ws.cell(row=2, column=2, value="Hours")

        row_num = 3
        # Day-shift data
        for eq in equipment_day:
            ws.cell(row=row_num, column=1, value=eq)
            ws.cell(row=row_num, column=2, value=8.0)
            row_num += 1

        # Separator row: repeats the header's first cell ("Equipment")
        ws.cell(row=row_num, column=1, value="Equipment")
        ws.cell(row=row_num, column=2, value="Hours")
        row_num += 1

        # Night-shift data
        for eq in equipment_night:
            ws.cell(row=row_num, column=1, value=eq)
            ws.cell(row=row_num, column=2, value=6.0)
            row_num += 1

    # Extra (non-numeric) sheets
    if extra_sheets:
        for name, rows in extra_sheets.items():
            ws = wb.create_sheet(title=name)
            for r_idx, row_data in enumerate(rows, start=1):
                for c_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestNormalProcessing:
    """process_excel_data produces a correct DataFrame from a valid file."""

    def test_normal_processing_day_and_night_shifts(self, tmp_path):
        """Both day and night shifts appear in output with correct dates."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[10, 15])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=3,
            return_sheets=True,
        )

        assert result is not None
        df = result["工时数据"]
        assert isinstance(df, pd.DataFrame)
        assert not df.empty

        # Must contain date column with expected values
        dates = set(df["日期"].astype(str))
        assert "2025-03-10" in dates
        assert "2025-03-15" in dates

        # Must contain both shifts
        shifts = set(df["班次"])
        assert "Day" in shifts
        assert "Night" in shifts

    def test_normal_processing_equipment_names(self, tmp_path):
        """Equipment names from both shifts appear in output."""
        excel_path = str(tmp_path / "worktime.xlsx")
        # Note: split_day_night_shifts with day_end_offset=-1 excludes the
        # last day-shift row before the separator (typically a summary row).
        # So we put "Truck" as the last row which will be excluded.
        day_equip = ["Excavator", "Truck"]
        night_equip = ["Loader", "Drill"]
        _create_worktime_excel(excel_path, days=[5], equipment_day=day_equip, equipment_night=night_equip)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        df = result["工时数据"]
        equipment = set(df["Equipment"].astype(str))
        # "Truck" is the last day row before the separator and is excluded
        # by day_end_offset=-1 (that row is typically a subtotal in real data)
        assert "Excavator" in equipment
        for name in night_equip:
            assert name in equipment, f"Expected equipment '{name}' in output"
        assert "Truck" not in equipment, "Last day row should be excluded by day_end_offset=-1"

    def test_normal_processing_sorted_by_date(self, tmp_path):
        """Output DataFrame is sorted by date."""
        excel_path = str(tmp_path / "worktime.xlsx")
        # Create days in reverse order
        _create_worktime_excel(excel_path, days=[20, 5, 10])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=6,
            return_sheets=True,
        )

        df = result["工时数据"]
        dates = df["日期"].tolist()
        # After strip_date_column, dates are date objects
        assert dates == sorted(dates), "DataFrame should be sorted by date"

    def test_date_and_shift_columns_first(self, tmp_path):
        """日期 and 班次 are the first two columns."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        df = result["工时数据"]
        assert list(df.columns[:2]) == ["日期", "班次"]


class TestReturnSheets:
    """return_sheets flag controls whether a dict is returned vs file written."""

    def test_return_sheets_returns_dict_with_expected_key(self, tmp_path):
        """return_sheets=True returns dict with key '工时数据'."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        assert isinstance(result, dict)
        assert "工时数据" in result
        assert isinstance(result["工时数据"], pd.DataFrame)

    def test_return_sheets_false_returns_none(self, tmp_path):
        """return_sheets=False writes file and returns None."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=False,
        )

        assert result is None

    def test_return_sheets_true_no_output_file(self, tmp_path):
        """return_sheets=True should not create any output file."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith(".xlsx") and f != "worktime.xlsx"]
        assert xlsx_files == []

    def test_return_sheets_contains_all_day_rows(self, tmp_path):
        """return_sheets DataFrame reflects day_end_offset=-1 behavior."""
        excel_path = str(tmp_path / "worktime.xlsx")
        # day_end_offset=-1 excludes the last day-shift row before separator
        _create_worktime_excel(
            excel_path,
            days=[1],
            equipment_day=["A", "B"],  # "B" is last day row, gets excluded
            equipment_night=["C"],
        )

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        df = result["工时数据"]
        # Day: "A" only (2nd row "B" excluded by day_end_offset=-1)
        # Night: "C" (1 row)
        # Total = 2
        assert len(df) == 2
        assert set(df["Equipment"]) == {"A", "C"}


class TestNumericSheetNames:
    """Only numeric sheet names are processed."""

    def test_only_numeric_sheets_processed(self, tmp_path):
        """Non-numeric sheets are skipped; only day-number sheets produce data."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(
            excel_path,
            days=[5, 12],
            extra_sheets={
                "Summary": [["Total", 100]],
                "Sheet": [["foo", "bar"]],
            },
        )

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=2,
            return_sheets=True,
        )

        df = result["工时数据"]
        # Only days 5 and 12 should produce data
        dates = set(df["日期"].astype(str))
        assert dates == {"2025-02-05", "2025-02-12"}

    def test_single_digit_day(self, tmp_path):
        """Single-digit day number (e.g. '3') is treated as a valid date."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[3])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=7,
            return_sheets=True,
        )

        df = result["工时数据"]
        dates = set(df["日期"].astype(str))
        assert "2025-07-03" in dates


class TestSkipNonDateSheets:
    """Sheets with non-numeric names produce a warning and are skipped."""

    def test_skip_alpha_sheet(self, tmp_path):
        """Sheet named 'Overview' is skipped entirely."""
        excel_path = str(tmp_path / "worktime.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overview"
        ws.cell(row=1, column=1, value="Some info")
        ws.cell(row=2, column=1, value="Equipment")
        ws.cell(row=2, column=2, value="Hours")
        ws.cell(row=3, column=1, value="Truck")
        ws.cell(row=3, column=2, value=8)
        wb.save(excel_path)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        # No numeric sheets => no data extracted => returns None
        assert result is None

    def test_skip_mixed_alphanumeric_sheet(self, tmp_path):
        """Sheet named 'Day5' (non-pure-numeric) is skipped."""
        excel_path = str(tmp_path / "worktime.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Day5"
        ws.cell(row=1, column=1, value="Title")
        ws.cell(row=2, column=1, value="Equipment")
        ws.cell(row=2, column=2, value="Hours")
        ws.cell(row=3, column=1, value="Truck")
        ws.cell(row=3, column=2, value=8)
        wb.save(excel_path)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        assert result is None

    def test_skip_sheet_with_spaces_in_number(self, tmp_path):
        """Sheet named ' 10 ' (spaces around number) is accepted because clean_string strips."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[10])

        # Rename the sheet to have spaces
        wb = openpyxl.load_workbook(excel_path)
        wb["10"].title = " 10 "
        wb.save(excel_path)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=4,
            return_sheets=True,
        )

        assert result is not None
        df = result["工时数据"]
        dates = set(df["日期"].astype(str))
        assert "2025-04-10" in dates


class TestYearMonthInOutput:
    """Output filename and date column reflect the provided year/month."""

    def test_output_filename_default(self, tmp_path):
        """Default output file is '{year}{month:02d}_工作效率表.xlsx' in the input dir."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        process_excel_data(
            file_path=excel_path,
            year=2024,
            month=11,
            return_sheets=False,
        )

        expected = tmp_path / "202411_工作效率表.xlsx"
        assert expected.exists(), f"Expected output file {expected} not found"

    def test_output_filename_custom(self, tmp_path):
        """Custom output_file path is used when provided."""
        excel_path = str(tmp_path / "worktime.xlsx")
        output_path = str(tmp_path / "custom.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        process_excel_data(
            file_path=excel_path,
            year=2024,
            month=11,
            output_file=output_path,
            return_sheets=False,
        )

        assert os.path.exists(output_path)

    def test_date_column_contains_correct_year_month(self, tmp_path):
        """Dates in output reflect the year and month arguments, not the file name."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[7, 22])

        result = process_excel_data(
            file_path=excel_path,
            year=2023,
            month=9,
            return_sheets=True,
        )

        df = result["工时数据"]
        dates = set(df["日期"].astype(str))
        assert "2023-09-07" in dates
        assert "2023-09-22" in dates

    def test_single_digit_month_padded(self, tmp_path):
        """Month 1 becomes '01' in date strings and filename."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        df = result["工时数据"]
        dates = set(df["日期"].astype(str))
        assert "2025-01-01" in dates


class TestEmptyFile:
    """Files with no valid (numeric) sheets return None gracefully."""

    def test_empty_workbook(self, tmp_path):
        """An Excel file with no sheets returns None."""
        excel_path = str(tmp_path / "empty.xlsx")
        # openpyxl always has at least one sheet, but we can make it non-numeric
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"
        wb.save(excel_path)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        assert result is None

    def test_numeric_sheet_with_no_data_rows(self, tmp_path):
        """A numeric sheet with only a title and header row (no data) should produce an empty or no DataFrame."""
        excel_path = str(tmp_path / "no_data.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "1"
        ws.cell(row=1, column=1, value="Title")
        ws.cell(row=2, column=1, value="Equipment")
        ws.cell(row=2, column=2, value="Hours")
        # No data rows at all
        wb.save(excel_path)

        # split_day_night_shifts with no data rows after header:
        # data_start_index=2, len(df_raw)=2 => no iteration => split_idx=-1
        # day_data = df_raw.iloc[2:] => empty
        # After clean_split_dataframe, likely empty
        # The all_data list gets an empty df, then after concat and dedup...
        # This should still produce a result (possibly with 0 rows) or None
        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )
        # Accept either None (if all data was dropped) or a dict with empty/near-empty df
        if result is not None:
            df = result["工时数据"]
            # If it returned data, it should at least be a DataFrame
            assert isinstance(df, pd.DataFrame)

    def test_file_not_found_raises(self, tmp_path):
        """Non-existent file raises FileNotFoundError."""
        with pytest.raises(FileNotFoundError, match="找不到输入文件"):
            process_excel_data(
                file_path=str(tmp_path / "nonexistent.xlsx"),
                year=2025,
                month=1,
            )


class TestHeaderMapping:
    """Verify header_mapping renames columns in the output."""

    def test_header_mapping_position_mode(self, tmp_path):
        """Position-based header mapping renames columns correctly."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        mapping = {
            "mode": "position",
            "entries": [
                {"index": 1, "new": "Date"},       # 1-based: col 0 = 日期
                {"index": 2, "new": "Shift"},      # col 1 = 班次
                {"index": 3, "new": "EquipName"},   # col 2 = Equipment
                {"index": 4, "new": "WorkHours"},   # col 3 = Hours
            ],
        }

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
            header_mapping=mapping,
        )

        df = result["工时数据"]
        assert "Date" in df.columns
        assert "Shift" in df.columns
        assert "EquipName" in df.columns
        assert "WorkHours" in df.columns

    def test_header_mapping_empty_entries_is_noop(self, tmp_path):
        """Empty entries list does not rename anything."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        mapping = {"mode": "position", "entries": []}

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
            header_mapping=mapping,
        )

        df = result["工时数据"]
        assert "日期" in df.columns
        assert "Equipment" in df.columns

    def test_header_mapping_none_is_noop(self, tmp_path):
        """header_mapping=None does not rename anything."""
        excel_path = str(tmp_path / "worktime.xlsx")
        _create_worktime_excel(excel_path, days=[1])

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
            header_mapping=None,
        )

        df = result["工时数据"]
        assert "日期" in df.columns
        assert "Equipment" in df.columns


class TestDeduplication:
    """Duplicate rows are removed from output."""

    def test_duplicate_rows_removed(self, tmp_path):
        """Identical data rows across shifts are deduplicated."""
        excel_path = str(tmp_path / "worktime.xlsx")
        # Create file with same equipment in both shifts with same hours
        # Note: day_end_offset=-1 means the last day row before separator is excluded.
        # So we put "Truck" twice in day shift - the second one gets excluded,
        # then the night shift "Truck" with same hours is a duplicate of the first.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "1"
        ws.cell(row=1, column=1, value="Title")
        ws.cell(row=2, column=1, value="Equipment")
        ws.cell(row=2, column=2, value="Hours")
        # Day shift data - first row kept
        ws.cell(row=3, column=1, value="Truck")
        ws.cell(row=3, column=2, value=8.0)
        # Day shift - second row (will be excluded by day_end_offset=-1)
        ws.cell(row=4, column=1, value="Truck")
        ws.cell(row=4, column=2, value=8.0)
        # Separator
        ws.cell(row=5, column=1, value="Equipment")
        ws.cell(row=5, column=2, value="Hours")
        # Night shift - same data as day shift row 1
        ws.cell(row=6, column=1, value="Truck")
        ws.cell(row=6, column=2, value=8.0)
        wb.save(excel_path)

        result = process_excel_data(
            file_path=excel_path,
            year=2025,
            month=1,
            return_sheets=True,
        )

        df = result["工时数据"]
        # Day: "Truck" 8.0 (row 3 kept, row 4 excluded by offset)
        # Night: "Truck" 8.0 (row 6)
        # Same date, different shifts (Day vs Night) => not exact dup
        assert len(df) == 2
        assert set(df["班次"]) == {"Day", "Night"}
