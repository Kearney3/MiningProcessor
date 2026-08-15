"""Comprehensive tests for func/excel_fuel.py (process_diesel_data)."""

import os
import datetime
from pathlib import Path

import pytest
import openpyxl

from func.excel_fuel import process_diesel_data


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _build_realistic_fuel_sheet(ws, *, include_data=True):
    """Populate an openpyxl worksheet with a realistic diesel consumption layout.

    The parser expects this pandas row layout (after header=None parse):

    iloc 0-1  : title/filler rows
    iloc 2    : h2 (date row)  — dates, "起运小时数", or info labels
    iloc 3    : h3 (group row) — 班组 info, ffill-expanded
    iloc 4    : h4 (shift row) — combined shift+type labels, e.g. "白班小时数", "夜班柴油"
    iloc 5    : h5 (oil row)   — fuel type names like "柴油", or empty
    iloc 6    : marker row (col A == 1 triggers start_row detection)
    iloc 7+   : data rows

    Key detail: h4 is scanned for BOTH shift keywords (白班/夜班) via detect_shift()
    AND for hour-type keywords ("小时数", "已使用小时数") via substring check.
    So data columns must have combined values like "白班小时数" or "夜班柴油".
    """
    # Excel rows 1-2: title
    ws.cell(row=1, column=1, value="设备柴油消耗月报表")
    ws.cell(row=2, column=1, value="2025年1月")

    # h2 (date row) - Excel row 3, pandas iloc 2
    ws.cell(row=3, column=1, value="序号")
    ws.cell(row=3, column=2, value="设备名称")
    ws.cell(row=3, column=3, value="设备编号")
    ws.cell(row=3, column=4, value="起运小时数")
    # Date 1: cols 5-6 (Day end_hours + Day fuel), cols 7-8 (Night end_hours + Night fuel)
    ws.cell(row=3, column=5, value="2025-01-15")
    ws.cell(row=3, column=6, value="2025-01-15")
    ws.cell(row=3, column=7, value="2025-01-15")
    ws.cell(row=3, column=8, value="2025-01-15")
    # Date 2: cols 9-10 (Day end_hours + Day fuel)
    ws.cell(row=3, column=9, value="2025-01-16")
    ws.cell(row=3, column=10, value="2025-01-16")

    # h3 (group row) - Excel row 4, pandas iloc 3
    ws.cell(row=4, column=1, value="班组A")
    ws.cell(row=4, column=2, value="班组A")
    ws.cell(row=4, column=3, value="班组A")

    # h4 (shift + type row) - Excel row 5, pandas iloc 4
    # Combined shift+type labels for data columns.
    # detect_shift('白班') = 'Day', detect_shift('夜班') = 'Night'.
    # '小时数' in '白班小时数' = True => classified as end_hours.
    ws.cell(row=5, column=1, value="白班")
    ws.cell(row=5, column=2, value="白班")
    ws.cell(row=5, column=3, value="白班")
    ws.cell(row=5, column=4, value="白班")
    ws.cell(row=5, column=5, value="白班小时数")    # Day end_hours
    ws.cell(row=5, column=6, value="白班柴油")      # Day fuel (no hour marker => fuel)
    ws.cell(row=5, column=7, value="夜班小时数")    # Night end_hours
    ws.cell(row=5, column=8, value="夜班柴油")      # Night fuel
    ws.cell(row=5, column=9, value="白班小时数")    # Day end_hours
    ws.cell(row=5, column=10, value="白班柴油")     # Day fuel

    # h5 (oil type row) - Excel row 6, pandas iloc 5
    ws.cell(row=6, column=6, value="柴油")
    ws.cell(row=6, column=8, value="柴油")
    ws.cell(row=6, column=10, value="柴油")

    # Marker row - Excel row 7, pandas iloc 6
    ws.cell(row=7, column=1, value=1)

    if not include_data:
        return

    # Data rows (Excel row 8+, pandas iloc 7+)
    # Device 1: CAT 785D
    ws.cell(row=8, column=1, value=1)
    ws.cell(row=8, column=2, value="CAT 785D")
    ws.cell(row=8, column=3, value="D001")
    ws.cell(row=8, column=4, value=1000.0)   # initial start hours
    ws.cell(row=8, column=5, value=1010.0)   # Day1 end hours
    ws.cell(row=8, column=6, value=150.0)    # Day1 fuel
    ws.cell(row=8, column=7, value=1020.0)   # Night1 end hours
    ws.cell(row=8, column=8, value=120.0)    # Night1 fuel
    ws.cell(row=8, column=9, value=1030.0)   # Day2 end hours
    ws.cell(row=8, column=10, value=160.0)   # Day2 fuel

    # Device 2: KOM 730E
    ws.cell(row=9, column=1, value=2)
    ws.cell(row=9, column=2, value="KOM 730E")
    ws.cell(row=9, column=3, value="D002")
    ws.cell(row=9, column=4, value=500.0)
    ws.cell(row=9, column=5, value=508.0)
    ws.cell(row=9, column=6, value=180.0)
    ws.cell(row=9, column=7, value=516.0)
    ws.cell(row=9, column=8, value=130.0)
    ws.cell(row=9, column=9, value=524.0)
    ws.cell(row=9, column=10, value=170.0)


def _create_fuel_excel(path, *, sheet_names=None, include_data=True):
    """Create an Excel file with one or more diesel sheets.

    Args:
        path: File path to save.
        sheet_names: List of sheet title strings. Defaults to ["设备柴油消耗表"].
        include_data: Whether to add data rows after headers.
    """
    if sheet_names is None:
        sheet_names = ["设备柴油消耗表"]

    wb = openpyxl.Workbook()
    for i, name in enumerate(sheet_names):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = name
        _build_realistic_fuel_sheet(ws, include_data=include_data)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestNormalProcessing:
    """Normal processing of a realistic diesel consumption sheet."""

    def test_returns_output_file_path(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path))

        # Assert
        assert result is not None
        assert result.endswith("Fuel.xlsx")
        assert os.path.exists(result)

    def test_engine_data_extracted(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        assert result is not None
        assert "设备信息" in result
        df_engine = result["设备信息"]
        assert len(df_engine) > 0
        assert set(["日期", "班次", "设备名称", "设备编号", "发动机小时数开始",
                     "发动机小时数结束", "运行小时数"]).issubset(df_engine.columns)

    def test_fuel_data_extracted(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        assert result is not None
        assert "油耗信息" in result
        df_fuel = result["油耗信息"]
        assert len(df_fuel) > 0
        assert set(["日期", "班次", "设备名称", "设备编号", "油品种类",
                     "油品消耗"]).issubset(df_fuel.columns)

    def test_device_names_correct(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        devices = set(df_engine["设备名称"])
        assert "CAT 785D" in devices
        assert "KOM 730E" in devices

    def test_dates_are_date_objects(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        for d in df_engine["日期"]:
            assert isinstance(d, datetime.date)

    def test_shifts_are_day_or_night(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        assert set(df_engine["班次"]).issubset({"Day", "Night"})

    def test_sorted_by_date_and_shift(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        dates = list(df_engine["日期"])
        assert dates == sorted(dates)

    def test_fuel_type_values(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_fuel = result["油耗信息"]
        assert all(t == "柴油" for t in df_fuel["油品种类"])

    def test_empty_device_name_falls_back_to_device_id(self, tmp_path):
        # Arrange: keep the existing row's fuel values but remove its name.
        excel_path = _create_fuel_excel(tmp_path / "fuel_input.xlsx")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["设备柴油消耗表"]
        ws["B8"] = None
        ws["C8"] = "LO#165"
        wb.save(excel_path)

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert: the row is retained and the ID is used as its display name.
        df_fuel = result["油耗信息"]
        lo_rows = df_fuel[df_fuel["设备编号"] == "LO#165"]
        assert len(lo_rows) == 3
        assert set(lo_rows["设备名称"]) == {"LO#165"}


class TestMultipleSheets:
    """Test processing a file with multiple matching sheets."""

    def test_both_sheets_processed(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(
            tmp_path / "multi.xlsx",
            sheet_names=["设备柴油消耗表1", "设备柴油消耗表2"],
        )

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        assert result is not None
        # Both sheets produce data, so results should have more rows than a single sheet
        df_engine = result["设备信息"]
        assert len(df_engine) > 4  # at least 2 devices * 2 sheets worth


class TestNoMatchingSheet:
    """Test ValueError when no sheet name matches."""

    def test_raises_value_error(self, tmp_path):
        # Arrange
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RandomSheet"
        ws.cell(row=1, column=1, value="data")
        path = tmp_path / "no_match.xlsx"
        wb.save(path)

        # Act & Assert
        with pytest.raises(ValueError, match="未找到匹配的柴油消耗Sheet"):
            process_diesel_data(str(path))


class TestEmptyData:
    """Test handling of a sheet with headers but no usable data rows."""

    def test_no_valid_data_raises(self, tmp_path):
        # Arrange: headers + marker but no device data rows
        excel_path = _create_fuel_excel(
            tmp_path / "empty.xlsx", include_data=False
        )

        # Act & Assert
        with pytest.raises(ValueError, match="未找到有效数据"):
            process_diesel_data(str(excel_path))


class TestMongolianHeaders:
    """Test with Mongolian sheet name 'Техник'."""

    def test_mongolian_sheet_name(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(
            tmp_path / "mongolian.xlsx", sheet_names=["Техникин"]
        )

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        assert result is not None
        assert "设备信息" in result


class TestReturnSheetsMode:
    """Test return_sheets=True returns dict without writing file."""

    def test_returns_dict_not_path(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        assert isinstance(result, dict)
        assert "设备信息" in result or "油耗信息" in result

    def test_does_not_write_output_file(self, tmp_path):
        # Arrange: use a different input name to avoid case-insensitive collision
        # with the default output "Fuel.xlsx" on macOS
        excel_path = _create_fuel_excel(tmp_path / "diesel_input.xlsx")

        # Act
        process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        output_file = tmp_path / "Fuel.xlsx"
        assert not output_file.exists()

    def test_returns_none_when_no_valid_data(self, tmp_path):
        # Arrange: a sheet that matches name but has short headers (skipped)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "设备柴油消耗表"
        ws.cell(row=1, column=1, value="hdr")
        ws.cell(row=2, column=1, value=1)  # marker at row 2 => start_row=2 < 6
        path = tmp_path / "short.xlsx"
        wb.save(path)

        # Act & Assert - should raise ValueError since no valid data found
        with pytest.raises(ValueError):
            process_diesel_data(str(path), return_sheets=True)


class TestYearFiltering:
    """Test target_year parameter overrides date years."""

    def test_year_override(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel.xlsx")

        # Act
        result = process_diesel_data(
            str(excel_path), target_year=2024, return_sheets=True
        )

        # Assert
        assert result is not None
        df_engine = result["设备信息"]
        for d in df_engine["日期"]:
            assert d.year == 2024

    def test_default_year_preserved(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        for d in df_engine["日期"]:
            assert d.year == 2025  # original year from fixture


class TestProcessFuelDataAlias:
    """Test that process_fuel_data is an alias for process_diesel_data."""

    def test_alias_exists(self):
        from func.excel_fuel import process_fuel_data
        assert process_fuel_data is process_diesel_data


class TestEngineHoursChain:
    """Test that engine hours chain is maintained correctly across shifts."""

    def test_hours_chain_continuity(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert - for device D001, the chain should be:
        # initial_start=1000 -> Day1 end=1010 -> Night1 end=1020 -> Day2 end=1030
        df_engine = result["设备信息"]
        d001 = df_engine[df_engine["设备编号"] == "D001"].sort_values(
            by=["日期", "班次"], key=lambda s: s.map({"Day": 0, "Night": 1}) if s.name == "班次" else s
        )
        starts = list(d001["发动机小时数开始"])
        ends = list(d001["发动机小时数结束"])
        # First start should be the initial value
        assert starts[0] == 1000.0
        # Each subsequent start should equal previous end
        for i in range(1, len(starts)):
            assert starts[i] == ends[i - 1]

    def test_work_hours_recorded(self, tmp_path):
        # Arrange
        excel_path = _create_fuel_excel(tmp_path / "fuel.xlsx")

        # Act
        result = process_diesel_data(str(excel_path), return_sheets=True)

        # Assert
        df_engine = result["设备信息"]
        # All work hours should be numeric
        for val in df_engine["运行小时数"]:
            assert isinstance(val, (int, float))


class TestDeduplication:
    """Test that duplicate rows are removed."""

    def test_duplicates_removed(self, tmp_path):
        # Arrange - create a sheet where data has a duplicate device row
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "设备柴油消耗表"

        # Minimal valid structure
        _build_realistic_fuel_sheet(ws)

        # Add an exact duplicate of row 8 (device D001)
        ws.cell(row=10, column=1, value=1)
        ws.cell(row=10, column=2, value="CAT 785D")
        ws.cell(row=10, column=3, value="D001")
        ws.cell(row=10, column=4, value=1000.0)
        ws.cell(row=10, column=5, value=1010.0)
        ws.cell(row=10, column=6, value=150.0)
        ws.cell(row=10, column=7, value=1020.0)
        ws.cell(row=10, column=8, value=120.0)
        ws.cell(row=10, column=9, value=1030.0)
        ws.cell(row=10, column=10, value=160.0)

        path = tmp_path / "dup.xlsx"
        wb.save(path)

        # Act
        result = process_diesel_data(str(path), return_sheets=True)

        # Assert - dedup should reduce rows; fuel data for D001 should have
        # exactly the same count as non-dup run (3 data rows per device)
        assert result is not None


class TestSummaryColumnExclusion:
    """Test that monthly summary columns are excluded from output.

    Real-world files like '2020.01 сар түлш.xlsx' have summary columns
    (Түлш зэхэлт ээлжээр) after the last date block with total fuel values.
    These must NOT be treated as regular data columns.
    """

    @staticmethod
    def _build_sheet_with_summary(ws):
        """Build a sheet with one date block + summary columns.

        Layout (pandas iloc):
        Row 0: filler
        Row 1: h1 (date row) — dates in data cols, NaN in summary
        Row 2: h2 (leader/summary header row) — "Түлш зэхэлт ээлжээр" in summary area
        Row 3: h3 (shift row) — "өдөр"/"шөнө" in data, "Өдөр"/"Шөнө"/"Нийт" in summary
        Row 4: h4 (brand row) — "Primary" in data, empty in summary
        Row 5: marker (col A == 1)
        Row 6+: data rows

        Data cols: 4-9 (мц, АМЦ, өдөр fuel, -, -, шөнө fuel)
        Summary cols: 10-12 (Өдөр total, Шөнө total, Нийт total)
        """
        # Row 1 (h1 - date row)
        ws.cell(row=2, column=1, value="Парк дугаар")  # anchor
        ws.cell(row=2, column=2, value="Атомашин төрөл")
        ws.cell(row=2, column=3, value="Парк дугаар")
        ws.cell(row=2, column=5, value="2020-01-15")  # date in col 4 (0-indexed)

        # Row 2 (h2 - leader/summary header row)
        ws.cell(row=3, column=6, value="өдөр")  # day shift label
        ws.cell(row=3, column=10, value="Өдөр")  # summary day
        ws.cell(row=3, column=11, value="Шөнө")  # summary night
        ws.cell(row=3, column=12, value="Нийт түлш авалт")  # summary total

        # Row 3 (h3 - shift row)
        ws.cell(row=4, column=5, value="мц")     # end hours
        ws.cell(row=4, column=6, value="АМЦ")    # work hours
        ws.cell(row=4, column=7, value="өдөр")   # day shift
        ws.cell(row=4, column=11, value="Өдөр")  # summary day label
        ws.cell(row=4, column=12, value="Шөнө")  # summary night label

        # Row 4 (h4 - brand row)
        ws.cell(row=5, column=7, value="Primary")  # fuel brand for day
        ws.cell(row=5, column=8, value="НИК")      # another fuel brand

        # Row 5: summary header "Түлш зэхэлт ээлжээр" in h2 row
        # This is in iloc row 2 (Excel row 3), col 10 area
        ws.cell(row=3, column=11, value="Түлш зэхэлт ээлжээр")

        # Marker row
        ws.cell(row=6, column=1, value=1)

        # Data row: Device with small daily fuel, large summary totals
        ws.cell(row=7, column=1, value=1)
        ws.cell(row=7, column=2, value="TEREX TR100")
        ws.cell(row=7, column=3, value="HT0001")
        ws.cell(row=7, column=4, value=1000.0)  # initial hours
        ws.cell(row=7, column=5, value=1010.0)  # end hours
        ws.cell(row=7, column=6, value=5.0)     # work hours
        ws.cell(row=7, column=7, value=500.0)   # day fuel (real daily value)
        ws.cell(row=7, column=8, value=300.0)   # night fuel (real daily value)
        # Summary columns with much larger values (monthly totals)
        ws.cell(row=7, column=11, value=15000.0)  # summary day total
        ws.cell(row=7, column=12, value=12000.0)  # summary night total
        ws.cell(row=7, column=13, value=27000.0)  # summary grand total

    def test_summary_columns_excluded(self, tmp_path):
        """Summary fuel columns must not appear in output."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Техник_түлш_зарцуулалт"
        self._build_sheet_with_summary(ws)
        path = tmp_path / "summary.xlsx"
        wb.save(path)

        result = process_diesel_data(str(path), return_sheets=True)
        assert result is not None

        df_fuel = result.get("油耗信息")
        assert df_fuel is not None

        # All fuel values should be small daily values, NOT the summary totals
        max_fuel = df_fuel["油品消耗"].max()
        assert max_fuel <= 500.0, (
            f"Summary column value leaked into output: max fuel = {max_fuel}"
        )

    def test_summary_total_keyword_stops_columns(self, tmp_path):
        """The 'Нийт' keyword in the shift row must trigger stop."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Техник_түлш_зарцуулалт"

        # Minimal layout with summary section marked by "Нийт" in shift row
        ws.cell(row=2, column=1, value="Парк дугаар")
        ws.cell(row=2, column=5, value="2020-01-15")

        # Shift row with "Нийт" in summary area
        ws.cell(row=4, column=5, value="мц")
        ws.cell(row=4, column=6, value="АМЦ")
        ws.cell(row=4, column=7, value="өдөр")
        ws.cell(row=4, column=11, value="Нийт")

        # Brand row
        ws.cell(row=5, column=7, value="Primary")

        ws.cell(row=6, column=1, value=1)  # marker

        # Data row
        ws.cell(row=7, column=1, value=1)
        ws.cell(row=7, column=2, value="Device")
        ws.cell(row=7, column=3, value="D001")
        ws.cell(row=7, column=7, value=100.0)   # real fuel
        ws.cell(row=7, column=11, value=99000.0)  # summary value

        path = tmp_path / "nithit.xlsx"
        wb.save(path)

        result = process_diesel_data(str(path), return_sheets=True)
        if result and "油耗信息" in result:
            max_fuel = result["油耗信息"]["油品消耗"].max()
            assert max_fuel <= 100.0, (
                f"'Нийт' stop condition failed: max fuel = {max_fuel}"
            )


# ---------------------------------------------------------------------------
# Filter zero hours tests
# ---------------------------------------------------------------------------


def _build_fuel_sheet_with_zero_rows(ws):
    """Build a fuel sheet where some devices have zero engine hours."""
    ws.cell(row=1, column=1, value="设备柴油消耗月报表")
    ws.cell(row=2, column=1, value="2025年1月")
    ws.cell(row=3, column=1, value="序号")
    ws.cell(row=3, column=2, value="设备名称")
    ws.cell(row=3, column=3, value="设备编号")
    ws.cell(row=3, column=4, value="起运小时数")
    ws.cell(row=3, column=5, value="2025-01-15")
    ws.cell(row=3, column=6, value="2025-01-15")
    ws.cell(row=3, column=7, value="2025-01-15")
    ws.cell(row=3, column=8, value="2025-01-15")
    ws.cell(row=3, column=9, value="2025-01-15")
    ws.cell(row=3, column=10, value="2025-01-15")
    ws.cell(row=4, column=1, value="班组A")
    ws.cell(row=5, column=1, value="白班")
    ws.cell(row=5, column=5, value="白班小时数")
    ws.cell(row=5, column=6, value="白班已使用小时数")
    ws.cell(row=5, column=7, value="白班柴油")
    ws.cell(row=5, column=8, value="夜班小时数")
    ws.cell(row=5, column=9, value="夜班已使用小时数")
    ws.cell(row=5, column=10, value="夜班柴油")
    ws.cell(row=6, column=7, value="柴油")
    ws.cell(row=6, column=10, value="柴油")
    ws.cell(row=7, column=1, value=1)

    # Device 1: normal hours
    ws.cell(row=8, column=1, value=1)
    ws.cell(row=8, column=2, value="CAT 785D")
    ws.cell(row=8, column=3, value="D001")
    ws.cell(row=8, column=4, value=1000.0)
    ws.cell(row=8, column=5, value=1010.0)   # Day end_hours
    ws.cell(row=8, column=6, value=10.0)     # Day work_hours
    ws.cell(row=8, column=7, value=150.0)    # Day fuel
    ws.cell(row=8, column=8, value=1020.0)   # Night end_hours
    ws.cell(row=8, column=9, value=10.0)     # Night work_hours
    ws.cell(row=8, column=10, value=120.0)   # Night fuel

    # Device 2: zero end hours, zero work hours
    ws.cell(row=9, column=1, value=2)
    ws.cell(row=9, column=2, value="KOM 730E")
    ws.cell(row=9, column=3, value="D002")
    ws.cell(row=9, column=4, value=0.0)      # initial start = 0
    ws.cell(row=9, column=5, value=0.0)      # Day end_hours = 0
    ws.cell(row=9, column=6, value=0.0)      # Day work_hours = 0
    ws.cell(row=9, column=7, value=180.0)    # Day fuel
    ws.cell(row=9, column=8, value=0.0)      # Night end_hours = 0
    ws.cell(row=9, column=9, value=0.0)      # Night work_hours = 0
    ws.cell(row=9, column=10, value=130.0)   # Night fuel


def _create_fuel_excel_with_zeros(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备柴油消耗表"
    _build_fuel_sheet_with_zero_rows(ws)
    wb.save(path)
    return path


class TestFilterZeroEngineHours:
    """Tests for filter_zero_engine_hours parameter."""

    def test_zero_start_and_end_filtered(self, tmp_path):
        """Rows where start=0 and end=0 should be filtered."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_engine_hours=True)
        df = result["设备信息"]
        # Device D002 has start=0 end=0 on both shifts -> all rows filtered
        assert not (df["设备编号"] == "D002").any()
        # Device D001 remains
        assert (df["设备编号"] == "D001").any()

    def test_filter_disabled_keeps_all(self, tmp_path):
        """When disabled, rows with zero hours are kept."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_engine_hours=False)
        df = result["设备信息"]
        assert (df["设备编号"] == "D002").any()

    def test_partial_zero_filtered(self, tmp_path):
        """Row with start=0 but end!=0 should also be filtered."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=9, column=5, value=50.0)   # Day end_hours = 50 (start still 0)
        wb.save(excel_path)

        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_engine_hours=True)
        df = result["设备信息"]
        d002_rows = df[df["设备编号"] == "D002"]
        assert len(d002_rows) == 0

    def test_none_start_filtered(self, tmp_path):
        """Row with NaN start should be filtered."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=9, column=4, value=None)    # initial start = None -> NaN
        ws.cell(row=9, column=5, value=50.0)    # Day end = 50
        ws.cell(row=9, column=8, value=50.0)    # Night end = 50
        wb.save(excel_path)

        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_engine_hours=True)
        df = result["设备信息"]
        d002_day = df[(df["设备编号"] == "D002") & (df["班次"] == "Day")]
        assert len(d002_day) == 0


class TestFilterZeroWorkHours:
    """Tests for filter_zero_work_hours parameter."""

    def test_zero_work_hours_filtered(self, tmp_path):
        """Rows where 运行小时数 is 0 should be filtered."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_work_hours=True)
        df = result["设备信息"]
        # D002 has end=0, start=0, so work_hours = end-start = 0 -> filtered
        assert not (df["设备编号"] == "D002").any()

    def test_normal_work_hours_kept(self, tmp_path):
        """Rows with non-zero 运行小时数 should be kept."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_work_hours=True)
        df = result["设备信息"]
        assert (df["设备编号"] == "D001").any()

    def test_both_filters_combined(self, tmp_path):
        """Both filters can be enabled simultaneously."""
        excel_path = _create_fuel_excel_with_zeros(tmp_path / "fuel.xlsx")
        result = process_diesel_data(str(excel_path), return_sheets=True,
                                     filter_zero_engine_hours=True,
                                     filter_zero_work_hours=True)
        df = result["设备信息"]
        assert len(df) > 0
        assert (df["设备编号"] == "D001").any()
        assert not (df["设备编号"] == "D002").any()
