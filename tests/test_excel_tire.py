import pandas as pd
from unittest.mock import patch
from openpyxl import Workbook, load_workbook

from func.anomaly import detect_and_filter
from func.anomaly.rules import AnomalyConfig
from func.equipment_ledger import EquipmentLedger
from func.excel_tire import (
    TIRE_OUTPUT_SHEET,
    parse_tire_workbook,
    process_tire_data,
)
from func.ledger_postprocess import match_sheets


def _make_tire_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "轮胎数据"
    headers = [
        "胎号",
        "品牌",
        "型号",
        "当前状态",
        "所属车辆1",
        "安装位置1",
        "开始日期1",
        "开始时间1",
        "开始公里数1",
        "结束日期1",
        "结束时间1",
        "结束公里数1",
        "所属车辆2",
        "安装位置2",
        "开始日期2",
        "开始时间2",
        "开始公里数2",
        "结束日期2",
        "结束时间2",
        "结束公里数2",
        "寿命合计",
        "标准寿命",
    ]
    ws.append(headers)
    row = [
        "T-001",
        "Brand",
        "Model",
        "在用",
        "Truck-01",
        1,
        "2025.020.1",
        100,
        "8,000",
        "2025-02-05",
        150,
        8500,
        "Truck-02",
        2,
        "2025.03",
        200,
        9000,
        None,
        999,
        9999,
        50,
        100,
    ]
    ws.append(row)
    ws.append(row)  # duplicate installation periods
    ws.append(
        [
            "T-002",
            "Brand",
            "Model",
            "在用",
            "Truck-03",
            1,
            "2024-01-01",
            10,
            1000,
            "2024-01-10",
            20,
            1200,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            0,
            100,
        ]
    )
    ws.row_dimensions[4].hidden = True
    ws.column_dimensions["C"].hidden = True

    ignored = wb.create_sheet("ignored")
    ignored["A1"] = "not a tire table"
    wb.save(path)


def test_parse_all_sheets_deduplicates_and_recalculates(tmp_path):
    source = tmp_path / "tire.xlsx"
    _make_tire_workbook(source)

    rows, metadata = parse_tire_workbook(
        source,
        skip_hidden_rows=True,
        skip_hidden_cols=True,
    )

    assert len(rows) == 2
    assert metadata["deduplication"]["removed"] == 2
    assert metadata["skipped_sheets"][0]["sheet"] == "ignored"
    assert {row["源表"] for row in rows} == {"轮胎数据"}
    assert rows[0]["型号"] is None  # hidden column is not read

    ended = next(row for row in rows if row["周期状态"] == "已结束")
    running = next(row for row in rows if row["周期状态"] == "运行中")
    assert ended["安装日期"] == "2025-02-01"
    assert "2025.020.1→2025-02-01" in ended["异常原因"]
    assert ended["寿命（时间）"] == 50
    assert ended["寿命（里程）"] == 500
    assert ended["磨损程度"] == 0.5
    assert running["拆卸时使用时间"] is None
    assert running["拆卸时公里数"] is None
    assert running["寿命（时间）"] is None
    assert [row["安装次数"] for row in rows] == [1, 2]


def test_tire_anomaly_rules_cover_time_and_mileage():
    df = pd.DataFrame(
        {
            "寿命（时间）": [10, 20, 1000],
            "寿命（里程）": [100, 200, 10000],
        }
    )
    config = AnomalyConfig(
        enabled=True,
        flag_anomalies=True,
        use_sigma=False,
        use_percentile=False,
        thresholds={
            "tire": {
                "寿命（时间）": {"min": 0, "max": 100},
                "寿命（里程）": {"min": 0, "max": 1000},
            }
        },
    )
    config._anomaly_records = []

    result, anomalies = detect_and_filter(df, "tire", config=config)

    assert result["异常值"].tolist() == [False, False, True]
    assert anomalies is not None
    assert {row["异常列"] for row in config._anomaly_records} == {
        "寿命（时间）",
        "寿命（里程）",
    }
    assert config._anomaly_records[0]["数据类型"] == "轮胎寿命"


def test_equipment_ledger_matches_installation_vehicle():
    ledger = EquipmentLedger()
    ledger._df = pd.DataFrame(
        [
            {
                "设备名称": "Truck-01",
                "设备编号": "001",
                "公司": "矿山",
                "标准设备名称": "标准卡车",
                "标准设备编号": "STD-001",
                "标准公司名称": "标准公司",
            }
        ]
    )
    ledger._build_search_cache()

    matched = match_sheets(
        {"轮胎": pd.DataFrame({"安装车辆": ["Truck-01"]})},
        equipment_ledger=ledger,
    )

    result = matched["轮胎"]
    assert result.loc[0, "标准设备名称"] == "标准卡车"
    assert result.loc[0, "标准设备编号"] == "STD-001"


def test_process_tire_data_uses_shared_formatted_output(tmp_path):
    source = tmp_path / "tire.xlsx"
    output = tmp_path / "轮胎寿命统计.xlsx"
    _make_tire_workbook(source)

    process_tire_data(source, output_file=output, skip_hidden_rows=True)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [TIRE_OUTPUT_SHEET]
    sheet = workbook[TIRE_OUTPUT_SHEET]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:V3"
    life_col = next(cell.column for cell in sheet[1] if cell.value == "寿命（时间）")
    assert sheet.cell(2, life_col).number_format == "General"
    date_col = next(cell.column for cell in sheet[1] if cell.value == "安装日期")
    assert sheet.cell(2, date_col).number_format == "yyyy-mm-dd"


def test_tire_is_available_through_orchestration_and_tauri_rpc(tmp_path):
    source = tmp_path / "tire.xlsx"
    _make_tire_workbook(source)

    with patch("func.excel_tire.process_tire_data") as process_mock:
        from func.orchestration import process_single

        result = process_single("tire", str(source))

    process_mock.assert_called_once()
    assert result["output_file"] == str(tmp_path / "轮胎寿命统计.xlsx")

    import tauri_bridge

    with patch("func.orchestration.process_single", return_value={"output_file": "out.xlsx"}) as rpc_mock:
        rpc_result = tauri_bridge._process_tire({"path": str(source)})

    assert rpc_result == {"output_file": "out.xlsx"}
    rpc_mock.assert_called_once()
    assert rpc_mock.call_args.args[0] == "tire"
