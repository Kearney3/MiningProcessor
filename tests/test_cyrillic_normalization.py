"""Cyrillic homoglyph normalization integration tests.

Verifies that equipment names with Cyrillic look-alike characters are
normalized to Latin equivalents throughout the processing pipeline:

- MiningDataProcessor._normalize() and safe_str()
- process_diesel_data (excel_fuel)
- parse_excel_data (excel_electrical)
- EquipmentLedger matching
- enrich_dataframe_device (ledger_enrichment)
"""

import openpyxl
import pandas as pd

from func.equipment_ledger import EquipmentLedger
from func.excel_production_enhanced import MiningDataProcessor
from func.ledger_enrichment import enrich_dataframe_device
from func.string_utils import clean_equipment_name, normalize_cyrillic_homoglyphs

# ===========================================================================
# Shared fixtures
# ===========================================================================

# Real-world Cyrillic/Latin mixed equipment names from Mongolian mine Excel
CYRILLIC_C_777 = "СAT 777 #18KH"   # С is Cyrillic U+0421
LATIN_C_777 = "CAT 777 #18KH"
CYRILLIC_KAT = "КАТ777"             # All Cyrillic: К, А, Т
LATIN_KAT = "KAT777"
CYRILLIC_NTE = "NТЕ240"             # Т is Cyrillic U+0422
LATIN_NTE = "NTE240"


# ===========================================================================
# 1. MiningDataProcessor (excel_production_enhanced)
# ===========================================================================


class TestProductionProcessorNormalize:
    """Test _normalize() and safe_str() in MiningDataProcessor."""

    @staticmethod
    def _proc():
        return MiningDataProcessor(device_load_map={}, raw_start=-1)

    # --- _normalize: uppercase + strip spaces + Cyrillic -> Latin ---

    def test_normalize_cyrillic_c(self):
        proc = self._proc()
        assert proc._normalize("СAT 777") == "CAT777"

    def test_normalize_all_cyrillic(self):
        proc = self._proc()
        assert proc._normalize("КАТ777") == "KAT777"

    def test_normalize_mixed_script_with_hash(self):
        proc = self._proc()
        assert proc._normalize("СAT 777 #18KH") == "CAT777#18KH"

    def test_normalize_already_latin_unchanged(self):
        proc = self._proc()
        assert proc._normalize("CAT 777") == "CAT777"

    # --- safe_str (clean_equipment_name wrapper) ---

    def test_safe_str_cyrillic_c(self):
        proc = self._proc()
        assert proc.safe_str("СAT 777 #18KH") == "CAT 777 #18KH"

    def test_safe_str_all_cyrillic(self):
        proc = self._proc()
        assert proc.safe_str("КАТ777") == "KAT777"

    def test_safe_str_none_returns_empty(self):
        proc = self._proc()
        assert proc.safe_str(None) == ""

    # --- get_load_capacity with Cyrillic truck names ---

    def test_get_load_capacity_cyrillic_model(self):
        """Cyrillic model name in truck name should still match the load map."""
        proc = MiningDataProcessor(
            device_load_map={"CAT777": 100}, raw_start=-1
        )
        assert proc.get_load_capacity("СAT 777 Unit01") == 100

    def test_get_load_capacity_all_cyrillic_model(self):
        proc = MiningDataProcessor(
            device_load_map={"KAT777": 100}, raw_start=-1
        )
        assert proc.get_load_capacity("КАТ777") == 100

    def test_get_load_capacity_cyrillic_in_load_map_key(self):
        """Load map key itself has Cyrillic - should match Latin truck name."""
        proc = MiningDataProcessor(
            device_load_map={"СAT777": 100}, raw_start=-1
        )
        assert proc.get_load_capacity("CAT 777") == 100


# ===========================================================================
# 2. excel_fuel: process_diesel_data
# ===========================================================================


class TestFuelCyrillicNormalization:
    """Verify that fuel data processing normalizes Cyrillic in device names."""

    @staticmethod
    def _build_fuel_xlsx(path, device_name, device_id):
        """Build a minimal fuel Excel with sheet name matching the parser filter."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "设备柴油消耗月报表"

        # Title rows
        ws.cell(row=1, column=1, value="设备柴油消耗月报表")
        ws.cell(row=2, column=1, value="2025年1月")

        # h2: date row
        ws.cell(row=3, column=1, value="序号")
        ws.cell(row=3, column=2, value="设备名称")
        ws.cell(row=3, column=3, value="设备编号")
        ws.cell(row=3, column=4, value="起运小时数")
        ws.cell(row=3, column=5, value="2025-01-15")
        ws.cell(row=3, column=6, value="2025-01-15")

        # h3: group row
        ws.cell(row=4, column=2, value="班组A")

        # h4: shift+type row
        ws.cell(row=5, column=5, value="白班小时数")
        ws.cell(row=5, column=6, value="白班柴油")

        # h5: oil type row
        ws.cell(row=6, column=6, value="柴油")

        # Marker row
        ws.cell(row=7, column=1, value=1)

        # Data row
        ws.cell(row=8, column=1, value=1)
        ws.cell(row=8, column=2, value=device_name)
        ws.cell(row=8, column=3, value=device_id)
        ws.cell(row=8, column=4, value=1000.0)
        ws.cell(row=8, column=5, value=1010.0)
        ws.cell(row=8, column=6, value=150.0)

        wb.save(path)
        wb.close()

    def test_fuel_normalizes_cyrillic_c_device_name(self, tmp_path):
        from func.excel_fuel import process_diesel_data

        xlsx = tmp_path / "fuel.xlsx"
        self._build_fuel_xlsx(str(xlsx), "СAT 785D", "D001")
        result = process_diesel_data(str(xlsx), return_sheets=True)

        assert isinstance(result, dict)
        df_engine = result["设备信息"]
        names = df_engine["设备名称"].tolist()
        assert all("С" not in n for n in names), (
            f"Cyrillic С found in output names: {names}"
        )
        assert any("CAT" in n for n in names), f"Expected Latin CAT in {names}"

    def test_fuel_normalizes_cyrillic_device_id(self, tmp_path):
        from func.excel_fuel import process_diesel_data

        xlsx = tmp_path / "fuel.xlsx"
        self._build_fuel_xlsx(str(xlsx), "CAT 785D", "СAT-001")
        result = process_diesel_data(str(xlsx), return_sheets=True)

        assert isinstance(result, dict)
        df_engine = result["设备信息"]
        ids = df_engine["设备编号"].tolist()
        assert all("С" not in str(i) for i in ids), (
            f"Cyrillic С found in output IDs: {ids}"
        )


# ===========================================================================
# 3. excel_electrical: parse_excel_data
# ===========================================================================


class TestElectricalCyrillicNormalization:
    """Verify that electrical data parsing normalizes Cyrillic in device names."""

    @staticmethod
    def _build_electrical_xlsx(path, device_label):
        """Build a minimal electrical Excel with a Cyrillic device label."""
        rows = [
            ["日期", "", "", "", "2025-01-15", "2025-01-16"],
            [device_label, None, None, None, 100.0, 200.0],
        ]
        df = pd.DataFrame(rows)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Electrical", index=False, header=False)

    def test_electrical_normalizes_cyrillic_c(self, tmp_path):
        from func.excel_electrical import parse_excel_data

        xlsx = tmp_path / "elec.xlsx"
        self._build_electrical_xlsx(str(xlsx), "EX-СAT01 电力总消耗")
        # Use return_sheets=True to get result without writing to disk
        result = parse_excel_data(str(xlsx), return_sheets=True)

        assert result is not None
        df = result["电力消耗"]
        names = df["设备名称"].tolist()
        assert all("С" not in n for n in names), (
            f"Cyrillic С found in output names: {names}"
        )
        assert any("CAT" in n for n in names), f"Expected Latin CAT in {names}"

    def test_electrical_preserves_latin(self, tmp_path):
        from func.excel_electrical import parse_excel_data

        xlsx = tmp_path / "elec.xlsx"
        self._build_electrical_xlsx(str(xlsx), "EX-CAT01 电力总消耗")
        result = parse_excel_data(str(xlsx), return_sheets=True)

        assert result is not None
        df = result["电力消耗"]
        assert len(df) >= 2  # Two date columns -> two records


# ===========================================================================
# 4. EquipmentLedger: matching with Cyrillic names
# ===========================================================================


class TestEquipmentLedgerCyrillic:
    """Verify that EquipmentLedger matching handles Cyrillic names."""

    @staticmethod
    def _make_ledger(tmp_path, records):
        """Create a ledger Excel with the given records and load it."""
        df = pd.DataFrame(records, columns=[
            "设备名称", "设备编号", "公司",
            "标准设备名称", "标准设备编号", "标准公司名称",
        ])
        path = str(tmp_path / "ledger.xlsx")
        df.to_excel(path, index=False)
        ledger = EquipmentLedger()
        ledger.load(path)
        return ledger

    def test_match_by_name_with_cyrillic_c(self, tmp_path):
        """Query with Cyrillic С should match Latin entry in ledger."""
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "#18KH", "A公司",
            "CAT 777 #18KH", "HT#18KH", "标准公司",
        ]])
        # match() returns "标准名称" key (not "标准设备名称")
        result = ledger.match(CYRILLIC_C_777)
        assert result is not None, "Cyrillic С name should match Latin CAT 777"
        assert result["标准名称"] == "CAT 777 #18KH"

    def test_match_by_name_with_all_cyrillic(self, tmp_path):
        """All-Cyrillic КАТ777 should match Latin KAT777 in ledger."""
        ledger = self._make_ledger(tmp_path, [[
            "KAT777", "K001", "B公司",
            "KAT777 Standard", "SK001", "标准B",
        ]])
        result = ledger.match(CYRILLIC_KAT)
        assert result is not None, "All-Cyrillic КАТ777 should match Latin KAT777"

    def test_match_device_with_cyrillic_name_and_id(self, tmp_path):
        """match_device with Cyrillic name + id should normalize both."""
        ledger = self._make_ledger(tmp_path, [[
            "NTE240", "#1101", "C公司",
            "NTE240 HT#1101", "HT#1101", "标准C",
        ]])
        # Use just the base name "NТЕ240" (Cyrillic Т) to match raw name "NTE240"
        result = ledger.match_device(name=CYRILLIC_NTE, device_id="#1101")
        assert result is not None, "Cyrillic Т in NTE240 should match Latin NTE240"
        assert result["标准设备名称"] == "NTE240 HT#1101"

    def test_match_by_id_with_cyrillic(self, tmp_path):
        """match_by_id with Cyrillic chars should normalize before lookup."""
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "СAT-001", "D公司",
            "CAT 777 Standard", "CAT-001", "标准D",
        ]])
        result = ledger.match_by_id("CAT-001")
        assert result is not None, "Latin CAT-001 should match normalized Cyrillic entry"

    def test_ledger_entry_with_cyrillic_name_gets_normalized(self, tmp_path):
        """Ledger entry itself has Cyrillic - should be normalized on load."""
        ledger = self._make_ledger(tmp_path, [[
            "СAT 777", "#18KH", "A公司",
            "CAT 777 #18KH", "HT#18KH", "标准公司",
        ]])
        result = ledger.match("CAT 777")
        assert result is not None, "Ledger with Cyrillic entry should match Latin query"

    def test_ledger_entry_with_cyrillic_id_gets_normalized(self, tmp_path):
        """Ledger entry with Cyrillic in device ID should normalize."""
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "СAT-001", "A公司",
            "CAT 777 Standard", "CAT-001", "标准A",
        ]])
        result = ledger.match_by_id("CAT-001")
        assert result is not None, "Latin CAT-001 should match Cyrillic СAT-001 entry"

    def test_bidirectional_cyrillic_match(self, tmp_path):
        """Both entry and query have Cyrillic - should still match."""
        ledger = self._make_ledger(tmp_path, [[
            "СAT 777", "СAT-001", "A公司",
            "CAT 777 Standard", "CAT-001", "标准A",
        ]])
        result = ledger.match("СAT 777")
        assert result is not None, "Cyrillic entry + Cyrillic query should match"

    def test_match_device_returns_standard_name(self, tmp_path):
        """match_device result should contain normalized standard info."""
        ledger = self._make_ledger(tmp_path, [[
            "СAT 777", "СAT-001", "A公司",
            "CAT 777 Standard", "CAT-001", "标准A",
        ]])
        result = ledger.match_device(name="СAT 777", device_id="СAT-001")
        assert result is not None
        assert result["标准设备名称"] == "CAT 777 Standard"
        assert result["标准设备编号"] == "CAT-001"


# ===========================================================================
# 5. ledger_enrichment: enrich_dataframe_device with Cyrillic names
# ===========================================================================


class TestLedgerEnrichmentCyrillic:
    """Verify that enrich_dataframe_device normalizes Cyrillic names."""

    @staticmethod
    def _make_ledger(tmp_path, records):
        df = pd.DataFrame(records, columns=[
            "设备名称", "设备编号", "公司",
            "标准设备名称", "标准设备编号", "标准公司名称",
        ])
        path = str(tmp_path / "ledger.xlsx")
        df.to_excel(path, index=False)
        ledger = EquipmentLedger()
        ledger.load(path)
        return ledger

    def test_enrich_normalizes_cyrillic_name_column(self, tmp_path):
        """DataFrame with Cyrillic device names should match ledger."""
        # Raw name "CAT 777" and raw ID "#18KH" match the pair cache keys
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "#18KH", "A公司",
            "CAT 777 Standard", "#18KH", "标准公司",
        ]])

        # Use name that matches ledger raw name after normalization
        df = pd.DataFrame({
            "设备名称": ["СAT 777", "CAT 785D"],
            "设备编号": ["#18KH", "#002"],
        })

        result_df = enrich_dataframe_device(
            df, name_col="设备名称", id_col="设备编号",
            equipment_ledger=ledger,
        )

        assert result_df.loc[0, "标准设备名称"] == "CAT 777 Standard", (
            f"Cyrillic name not matched, got: {result_df.loc[0, '标准设备名称']}"
        )

    def test_enrich_normalizes_cyrillic_in_id_column(self, tmp_path):
        """DataFrame with Cyrillic in device ID should match ledger."""
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "CAT-001", "A公司",
            "CAT 777 Standard", "CAT-001", "标准公司",
        ]])

        df = pd.DataFrame({
            "设备名称": ["CAT 777"],
            "设备编号": ["СAT-001"],
        })

        result_df = enrich_dataframe_device(
            df, name_col="设备名称", id_col="设备编号",
            equipment_ledger=ledger,
        )

        assert result_df.loc[0, "标准设备名称"] == "CAT 777 Standard"

    def test_enrich_without_id_col_normalizes_cyrillic(self, tmp_path):
        """Name-only enrichment should normalize Cyrillic."""
        ledger = self._make_ledger(tmp_path, [[
            "KAT777", "K001", "B公司",
            "KAT777 Standard", "SK001", "标准B",
        ]])

        df = pd.DataFrame({"设备名称": [CYRILLIC_KAT]})

        result_df = enrich_dataframe_device(
            df, name_col="设备名称",
            equipment_ledger=ledger,
        )

        assert result_df.loc[0, "标准设备名称"] == "KAT777 Standard"

    def test_enrich_unmatched_returns_empty(self, tmp_path):
        """Unmatched Cyrillic name should get empty standard fields."""
        ledger = self._make_ledger(tmp_path, [[
            "CAT 777", "#18KH", "A公司",
            "CAT 777 Standard", "HT#18KH", "标准公司",
        ]])

        df = pd.DataFrame({"设备名称": ["XYZ999"]})

        result_df = enrich_dataframe_device(
            df, name_col="设备名称",
            equipment_ledger=ledger,
        )

        assert result_df.loc[0, "标准设备名称"] == ""


# ===========================================================================
# 6. Filter keyword normalization (excel_production_enhanced)
# ===========================================================================


class TestFilterKeywordNormalization:
    """Verify that Cyrillic filter keywords match safe_str()-normalized values.

    After safe_str() was changed to use clean_equipment_name() (which normalizes
    Cyrilillic homoglyphs), all hardcoded Cyrillic filter keywords must also be
    normalized to stay in the same comparison space.
    """

    _N = staticmethod(normalize_cyrillic_homoglyphs)

    # --- Summary row filter (line 336) ---

    def test_summary_row_filter_cyrillic_n(self):
        """'Нийт' (Cyrillic Н) must be filtered from safe_str() output."""
        truck_name = clean_equipment_name("Нийт")
        assert self._N("Нийт") in truck_name

    def test_summary_row_filter_latin_h(self):
        """'Hийт' (Latin H, already in Excel) must also be filtered."""
        truck_name = clean_equipment_name("Hийт")
        assert self._N("Нийт") in truck_name

    def test_summary_row_filter_does_not_match_real_truck(self):
        """'Нийт' filter must NOT match a real truck name like 'CAT 777'."""
        truck_name = clean_equipment_name("СAT 777 #18KH")
        assert self._N("Нийт") not in truck_name

    # --- Total trips column detection (line 250) ---

    def test_total_trips_column_cyrillic(self):
        """'Нийт рейс' (Cyrillic Н) must match safe_str()-normalized header."""
        col_val = clean_equipment_name("Нийт рейс")
        keywords = ["总趟次", self._N("Нийт рейс")]
        assert any(k in col_val for k in keywords)

    def test_total_trips_column_latin_h(self):
        """'Hийт рейс' (Latin H) must also match."""
        col_val = clean_equipment_name("Hийт рейс")
        keywords = ["总趟次", self._N("Нийт рейс")]
        assert any(k in col_val for k in keywords)

    # --- Column matching keywords (lines 286-291) ---

    def test_moto_column_match(self):
        """'Мото' (Cyrillic М) must match safe_str()-normalized header."""
        col_val = clean_equipment_name("Мото Эхэлсэн")
        assert self._N("Мото") in col_val

    def test_end_column_match(self):
        """'Дууссан' has no homoglyphs, must match directly."""
        col_val = clean_equipment_name("Мото Дууссан")
        assert self._N("Мото") in col_val
        assert self._N("Дууссан") in col_val

    def test_km_column_match(self):
        """'км-ын' has no homoglyphs (lowercase к), must match directly."""
        col_val = clean_equipment_name("км-ын Дууссан")
        assert self._N("км-ын") in col_val

    def test_company_column_match(self):
        """'Компани' (Cyrillic К) must match safe_str()-normalized header."""
        col_val = clean_equipment_name("Компани")
        assert self._N("Компани") in col_val

    # --- Excavator exclude list (line 297) ---

    def test_excavator_exclude_moto(self):
        """'Мото' in exclude list must match safe_str()-normalized excavator name."""
        name = clean_equipment_name("Мото")
        exclude = [self._N("Мото"), self._N("Эхэлсэн"), self._N("Компани"),
                   self._N("км"), self._N("Дууссан")]
        assert any(k in name for k in exclude)

    def test_excavator_exclude_kompani(self):
        """'Компани' in exclude list must match safe_str()-normalized name."""
        name = clean_equipment_name("Компани")
        exclude = [self._N("Мото"), self._N("Эхэлсэн"), self._N("Компани"),
                   self._N("км"), self._N("Дууссан")]
        assert any(k in name for k in exclude)

    def test_excavator_exclude_does_not_match_real_excavator(self):
        """Exclude list must NOT match a real excavator name."""
        name = clean_equipment_name("СAT 777 #18KH")
        exclude = [self._N("Мото"), self._N("Эхэлсэн"), self._N("Компани"),
                   self._N("км"), self._N("Дууссан")]
        assert not any(k in name for k in exclude)
