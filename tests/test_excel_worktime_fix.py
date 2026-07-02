"""Test return_sheets flag is checked BEFORE write_formatted_excel (finding #1 fix)."""

import os
import openpyxl
import pytest

from func.excel_worktime import process_excel_data


def _create_minimal_worktime_excel(path):
    """Create a minimal Excel file with a numeric sheet name (simulating a day)."""
    wb = openpyxl.Workbook()
    # Default sheet "Sheet" -> rename to "15" (day 15)
    ws = wb.active
    ws.title = "15"
    # Write minimal data that split_day_night_shifts can process: at least
    # one row with "早班" or "白班" and one row with "夜班" plus some data.
    ws.cell(row=1, column=1, value="早班")
    ws.cell(row=2, column=1, value="设备A")
    ws.cell(row=2, column=2, value="工时")
    ws.cell(row=3, column=1, value="夜班")
    ws.cell(row=4, column=1, value="设备B")
    ws.cell(row=4, column=2, value="工时")

    wb.save(path)


def test_return_sheets_true_no_file_written(tmp_path):
    """return_sheets=True must NOT create an output file on disk."""
    excel_path = os.path.join(tmp_path, "worktime.xlsx")
    _create_minimal_worktime_excel(excel_path)

    result = process_excel_data(
        file_path=excel_path,
        year=2025,
        month=6,
        return_sheets=True,
    )

    # Must return the sheets dict
    assert result is not None
    assert "工时数据" in result

    # No unexpected file should exist in tmp_path beyond our input
    written_files = [f for f in os.listdir(tmp_path) if f.endswith(".xlsx") and f != "worktime.xlsx"]
    assert len(written_files) == 0, (
        f"return_sheets=True should not write a file, but found: {written_files}"
    )


def test_return_sheets_false_writes_file(tmp_path):
    """return_sheets=False must create the output file on disk."""
    excel_path = os.path.join(tmp_path, "worktime.xlsx")
    _create_minimal_worktime_excel(excel_path)

    result = process_excel_data(
        file_path=excel_path,
        year=2025,
        month=6,
        return_sheets=False,
    )

    # Must return None (no sheets dict when not requested)
    assert result is None

    # An output file should have been written
    written_files = [f for f in os.listdir(tmp_path) if f.endswith(".xlsx") and f != "worktime.xlsx"]
    assert len(written_files) == 1, (
        f"return_sheets=False should write exactly one file, but found: {written_files}"
    )


def test_return_sheets_true_with_output_file_arg(tmp_path):
    """return_sheets=True must NOT write even when output_file is explicitly provided."""
    excel_path = os.path.join(tmp_path, "worktime.xlsx")
    output_path = os.path.join(tmp_path, "custom_output.xlsx")
    _create_minimal_worktime_excel(excel_path)

    result = process_excel_data(
        file_path=excel_path,
        year=2025,
        month=6,
        output_file=output_path,
        return_sheets=True,
    )

    assert result is not None
    assert "工时数据" in result
    # The explicitly-provided output_file must NOT exist
    assert not os.path.exists(output_path), (
        "return_sheets=True must not write even when output_file is provided"
    )
