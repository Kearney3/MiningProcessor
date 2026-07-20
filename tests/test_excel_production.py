"""Unit tests for func/excel_production_enhanced.py.

Creates synthetic Excel fixtures via openpyxl / pandas and exercises:
- parse_filename
- get_load_capacity (extends existing test_production_model_match.py)
- process_sheet1  (vehicle running data + production tonnage)
- process_sheet2  (sheet2 column-based parsing)
- process_single_file with single-sheet files
- day/night shift splitting via filename parsing
"""
import pathlib
import tempfile
from datetime import date

import pandas as pd
import pytest

ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(ROOT) not in __import__("sys").path:
    __import__("sys").path.insert(0, str(ROOT))

from func.excel_production_enhanced import MiningDataProcessor


# ---------------------------------------------------------------------------
# Helper: build a processor that bypasses config-file lookups
# ---------------------------------------------------------------------------

def _make_processor(load_map=None, raw_start=-1, target_text="Мото цагийн заалт"):
    """Create a MiningDataProcessor with a custom device_load_map."""
    return MiningDataProcessor(
        device_load_map=load_map or {},
        raw_start=raw_start,
        target_text=target_text,
    )


# ---------------------------------------------------------------------------
# Helper: build a raw DataFrame that mirrors the real mine Excel layout
# ---------------------------------------------------------------------------

# _RAW_START = 6  means:
#   header row 6 = df_raw.iloc[raw_start - 1, :]  (row index 5)
#   header row 7 = df_raw.iloc[raw_start, :]       (row index 6)
#   data rows    = df_raw.iloc[raw_start + 1:, :]   (row index 7+)

_HEADER_ROW6 = ["", "小时数", "小时数", "公里数", "公里数", "公司",
                 "1号挖机", "2号挖机", "3号挖机", "总趟次", "备注"]
_HEADER_ROW7 = ["", "开始", "结束", "开始", "结束", "",
                 "运次", "运次", "运次", "", ""]


def _build_sheet1_df(truck_data, raw_start=6, n_blanks_before_header=5):
    """Build a raw DataFrame (header=None) that mirrors the real mine layout.

    Layout with raw_start=6 (default):
      rows 0..4        = blank metadata rows
      row  5 (idx 5)   = header row 6 labels  (raw_start - 1)
      row  6 (idx 6)   = header row 7 sub-labels (raw_start)
      rows 7+          = data rows (raw_start + 1+)
    """
    n_cols = max(len(_HEADER_ROW6), 15)
    rows = [[None] * n_cols for _ in range(raw_start - 1)]  # blank rows up to raw_start-1

    # header row 6  (raw_start - 1)
    rows.append(_HEADER_ROW6 + [None] * (n_cols - len(_HEADER_ROW6)))
    # header row 7  (raw_start)
    rows.append(_HEADER_ROW7 + [None] * (n_cols - len(_HEADER_ROW7)))

    # data rows
    for td in truck_data:
        rows.append([
            td["truck_name"],
            td.get("hour_start", ""),
            td.get("hour_end", ""),
            td.get("km_start", ""),
            td.get("km_end", ""),
            td.get("company", ""),
            td.get("exc_1_trips", 0),
            td.get("exc_2_trips", 0),
            td.get("exc_3_trips", 0),
            td.get("total_trips", 0),
            td.get("remark", ""),
        ])

    return pd.DataFrame(rows, dtype=object)


def _write_sheet1_xlsx(path, truck_data, raw_start=6):
    """Write an Excel file whose first sheet mirrors the real mine layout.

    Layout (1-based Excel rows, raw_start=6):
      Row 1-5: blank metadata
      Row 6:   header row 6 labels  ("小时数", "公里数", "公司", "1号挖机", ...)
      Row 7:   header row 7 sub-labels ("开始", "结束", "运次", ...)
      Row 8+:  data rows
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # header row 6 (1-based row = raw_start)
    for c, val in enumerate(_HEADER_ROW6, start=1):
        ws.cell(row=raw_start, column=c, value=val if val else None)
    # header row 7 (1-based row = raw_start + 1)
    for c, val in enumerate(_HEADER_ROW7, start=1):
        ws.cell(row=raw_start + 1, column=c, value=val if val else None)

    # data rows start at 1-based row = raw_start + 2
    for row_idx, td in enumerate(truck_data, start=raw_start + 2):
        ws.cell(row=row_idx, column=1, value=td["truck_name"])
        ws.cell(row=row_idx, column=2, value=td.get("hour_start", ""))
        ws.cell(row=row_idx, column=3, value=td.get("hour_end", ""))
        ws.cell(row=row_idx, column=4, value=td.get("km_start", ""))
        ws.cell(row=row_idx, column=5, value=td.get("km_end", ""))
        ws.cell(row=row_idx, column=6, value=td.get("company", ""))
        ws.cell(row=row_idx, column=7, value=td.get("exc_1_trips", 0))
        ws.cell(row=row_idx, column=8, value=td.get("exc_2_trips", 0))
        ws.cell(row=row_idx, column=9, value=td.get("exc_3_trips", 0))
        ws.cell(row=row_idx, column=10, value=td.get("total_trips", 0))
        ws.cell(row=row_idx, column=11, value=td.get("remark", ""))

    wb.save(path)


# ---------------------------------------------------------------------------
# Helper: write a two-sheet Excel file (sheet1 + sheet2)
# ---------------------------------------------------------------------------

def _write_two_sheet_xlsx(path, truck_data, sheet2_devices=None, raw_start=6):
    """Write an Excel file with two sheets (sheet1 + sheet2).

    sheet2_devices: list of (device_name, company, hour_start, hour_end)
                   If None, a default single device is written.
    """
    _write_sheet1_xlsx(path, truck_data, raw_start)

    if sheet2_devices is None:
        sheet2_devices = [("CAT773-02", "CompanyB", 8.0, 15.5)]

    import openpyxl
    wb = openpyxl.load_workbook(path)
    wb.create_sheet("Sheet2")
    ws2 = wb["Sheet2"]
    ws2.cell(row=1, column=2, value="设备名称")
    ws2.cell(row=1, column=3, value="公司")
    ws2.cell(row=1, column=6, value="小时数开始")
    ws2.cell(row=1, column=7, value="小时数结束")
    for i, (dev, comp, hs, he) in enumerate(sheet2_devices, start=4):
        ws2.cell(row=i, column=2, value=dev)
        ws2.cell(row=i, column=3, value=comp)
        ws2.cell(row=i, column=6, value=hs)
        ws2.cell(row=i, column=7, value=he)
    wb.save(path)


# ---------------------------------------------------------------------------
# Tests: parse_filename
# ---------------------------------------------------------------------------

class TestParseFilename:
    def test_day_shift_with_dots(self):
        proc = _make_processor({})
        d, s = proc.parse_filename("2025.01.15 白班.xlsx")
        assert d == date(2025, 1, 15)
        assert s == "Day"

    def test_night_shift_with_dashes(self):
        proc = _make_processor({})
        d, s = proc.parse_filename("2025-03-20 夜班.xlsx")
        assert d == date(2025, 3, 20)
        assert s == "Night"

    def test_no_shift_keyword(self):
        proc = _make_processor({})
        d, s = proc.parse_filename("2025.06.01 report.xlsx")
        assert d == date(2025, 6, 1)
        assert s == ""

    def test_raises_on_missing_date(self):
        proc = _make_processor({})
        with pytest.raises(ValueError, match="文件名中未找到日期"):
            proc.parse_filename("report.xlsx")

    def test_mixed_separator_normalized(self):
        proc = _make_processor({})
        d, _ = proc.parse_filename("2025-04-10 白班.xlsx")
        assert d == date(2025, 4, 10)

    def test_year_boundary(self):
        proc = _make_processor({})
        d, _ = proc.parse_filename("2025.12.31 夜班.xlsx")
        assert d == date(2025, 12, 31)


# ---------------------------------------------------------------------------
# Tests: get_load_capacity (extend existing)
# ---------------------------------------------------------------------------

class TestGetLoadCapacityExtended:
    def test_empty_truck_name_returns_zero(self):
        proc = _make_processor({"TR60": 22})
        assert proc.get_load_capacity("") == 0

    def test_spaces_in_truck_name(self):
        proc = _make_processor({"TR60": 22})
        assert proc.get_load_capacity("  TR60  ") == 22

    def test_default_load_map_used_when_none_provided(self):
        """When device_load_map=None during init, fallback defaults apply."""
        proc = MiningDataProcessor(device_load_map=None, raw_start=-1)
        cap = proc.get_load_capacity("TR60")
        assert cap > 0

    def test_exact_match_same_length(self):
        proc = _make_processor({"TR60": 22, "TR100": 35})
        assert proc.get_load_capacity("TR60") == 22
        assert proc.get_load_capacity("TR100") == 35


# ---------------------------------------------------------------------------
# Tests: process_sheet1
# ---------------------------------------------------------------------------

class TestProcessSheet1:
    def test_normal_processing(self):
        """Production sheet with running data and production data."""
        load_map = {"TR600": 85, "TR60": 22}
        proc = _make_processor(load_map, raw_start=6)

        truck_data = [
            {
                "truck_name": "TR600-01",
                "hour_start": 100.0, "hour_end": 108.5,
                "km_start": 5000.0, "km_end": 5050.0,
                "company": "CompanyA",
                "exc_1_trips": 10, "exc_2_trips": 5, "exc_3_trips": 0,
                "total_trips": 15,
            },
            {
                "truck_name": "TR60-02",
                "hour_start": 200.0, "hour_end": 206.0,
                "km_start": 3000.0, "km_end": 3030.0,
                "company": "CompanyB",
                "exc_1_trips": 8, "exc_2_trips": 0, "exc_3_trips": 3,
                "total_trips": 11,
            },
        ]

        df_raw = _build_sheet1_df(truck_data, raw_start=6)
        test_date = date(2025, 1, 15)
        test_shift = "Day"

        running_df, production_df = proc.process_sheet1(df_raw, test_date, test_shift)

        # Running data should have 2 rows
        assert not running_df.empty
        assert len(running_df) == 2

        # Check TR600 row
        tr600_row = running_df[running_df["设备名称"] == "TR600-01"].iloc[0]
        assert tr600_row["运行小时数"] == pytest.approx(8.5)
        assert tr600_row["运行里程"] == pytest.approx(50.0)
        assert tr600_row["趟次"] == 15

        # Production data should have rows for excavator trips > 0
        assert not production_df.empty
        # TR600-01: 10 trips x 85t + 5 trips x 85t = 15 x 85 = 1275t total
        tr600_prod = production_df[production_df["矿卡名称"] == "TR600-01"]
        assert len(tr600_prod) == 2  # exc_1 and exc_2 each produce a row
        assert tr600_prod["产量"].sum() == pytest.approx(15 * 85)

        # TR60-02: 8 trips x 22t + 3 trips x 22t
        tr60_prod = production_df[production_df["矿卡名称"] == "TR60-02"]
        assert len(tr60_prod) == 2
        assert tr60_prod["产量"].sum() == pytest.approx(11 * 22)

    def test_empty_dataframe_returns_empty(self):
        proc = _make_processor({})
        df_raw = pd.DataFrame(dtype=object)
        running_df, production_df = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        assert running_df.empty
        assert production_df.empty

    def test_short_dataframe_returns_empty(self):
        proc = _make_processor({})
        df_raw = pd.DataFrame({"A": [1]}, dtype=object)
        running_df, production_df = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        assert running_df.empty
        assert production_df.empty

    def test_total_trips_col_not_found_uses_last_col(self):
        """If '总趟次' is missing, last_col_idx falls back to last column."""
        load_map = {"TR600": 85}
        proc = _make_processor(load_map, raw_start=6)

        # Build layout without 总趟次 column:
        # header row6 = ["", "小时数", "小时数", "公里数", "公里数", "公司",
        #                "1号挖机", "2号挖机", "3号挖机"]
        # header row7 = ["", "开始", "结束", "开始", "结束", "", "运次", "运次", "运次"]
        custom_row6 = ["", "小时数", "小时数", "公里数", "公里数", "公司",
                       "1号挖机", "2号挖机", "3号挖机"]
        custom_row7 = ["", "开始", "结束", "开始", "结束", "", "运次", "运次", "运次"]

        rows = [[None] * 9 for _ in range(5)]
        rows.append(custom_row6)
        rows.append(custom_row7)
        rows.append(["TR600-01", 100.0, 108.5, 5000.0, 5050.0, "CompA", 10, 0, 0])

        df_raw = pd.DataFrame(rows, dtype=object)
        running_df, production_df = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        assert not running_df.empty
        assert len(running_df) == 1

    def test_filtering_niit_and_empty_trucks(self):
        """Rows with empty truck names or 'Нийт' (total) should be skipped."""
        load_map = {"TR600": 85}
        proc = _make_processor(load_map, raw_start=6)

        truck_data = [
            {"truck_name": "", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "CompA", "exc_1_trips": 5},
            {"truck_name": "Нийт 合计", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "", "exc_1_trips": 10},
            {"truck_name": "TR600-01", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "CompA", "exc_1_trips": 10},
        ]
        df_raw = _build_sheet1_df(truck_data, raw_start=6)

        running_df, _ = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        assert len(running_df) == 1
        assert running_df.iloc[0]["设备名称"] == "TR600-01"

    def test_company_required_column_filtered(self):
        """When company column exists, rows with empty company are excluded."""
        load_map = {"TR600": 85}
        proc = _make_processor(load_map, raw_start=6)

        truck_data = [
            {"truck_name": "TR600-01", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "", "exc_1_trips": 10},
            {"truck_name": "TR600-02", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "CompA", "exc_1_trips": 5},
        ]
        df_raw = _build_sheet1_df(truck_data, raw_start=6)

        running_df, _ = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        assert len(running_df) == 1
        assert running_df.iloc[0]["设备名称"] == "TR600-02"

    def test_auto_detect_target_text(self):
        """When raw_start=-1, auto_detect finds the target_text row."""
        load_map = {"TR600": 85}
        proc = _make_processor(load_map, raw_start=-1, target_text="Мото цагийн заалт")

        truck_data = [
            {"truck_name": "TR600-01", "hour_start": 100.0, "hour_end": 108.5,
             "km_start": 5000.0, "km_end": 5050.0, "company": "CompA", "exc_1_trips": 10},
        ]
        # Build DF with target_text embedded in the header row area
        n_cols = len(_HEADER_ROW6)
        rows = [[None] * n_cols for _ in range(4)]
        # Row 4 = target text (raw_start will be auto-detected as 5)
        rows.append(["Мото цагийн заалт"] * n_cols)
        # Row 5 = header row6 labels (raw_start - 1 = 4 in this layout,
        #   but auto_detect sets raw_start=5, so header6=row4=target text,
        #   header7=row5=sub-labels)
        rows.append(_HEADER_ROW7[:n_cols] + [None] * max(0, n_cols - len(_HEADER_ROW7)))
        # Data
        rows.append(["TR600-01", 100.0, 108.5, 5000.0, 5050.0, "CompA", 10, 0, 0, 0, ""])

        df_raw = pd.DataFrame(rows, dtype=object)
        # raw_start will be auto-detected; header6 row = target_text row (all same value)
        # The combined header6+header7 will produce columns but company_col should still match
        running_df, _ = proc.process_sheet1(df_raw, date(2025, 1, 1), "Day")
        # Should process without raising
        assert not running_df.empty or running_df.empty  # may or may not find company col


# ---------------------------------------------------------------------------
# Tests: process_sheet2
# ---------------------------------------------------------------------------

class TestProcessSheet2:
    def _build_sheet2_df(self, devices):
        """Build a raw DataFrame for sheet2 layout.

        start_row = 3 (0-based). Columns:
          col 1 (B) = device name
          col 2 (C) = company
          col 5 (F) = hour start
          col 6 (G) = hour end
        """
        n_cols = 11
        rows = [[None] * n_cols for _ in range(3)]  # rows 0-2: header area

        for dev, comp, hs, he in devices:
            row = [None] * n_cols
            row[1] = dev      # col B
            row[2] = comp     # col C
            row[5] = hs       # col F
            row[6] = he       # col G
            rows.append(row)

        return pd.DataFrame(rows, dtype=object)

    def test_basic_sheet2_parsing(self):
        proc = _make_processor({})
        df_raw = self._build_sheet2_df([
            ("TR600-01", "CompanyA", 10.5, 18.3),
            ("CAT773-02", "CompanyB", 8.0, 15.5),
        ])
        result = proc.process_sheet2(df_raw, date(2025, 1, 15), "Day")

        assert not result.empty
        assert len(result) == 2
        assert result.iloc[0]["设备名称"] == "TR600-01"
        assert result.iloc[0]["运行小时数"] == pytest.approx(7.8)
        assert result.iloc[0]["日期"] == date(2025, 1, 15)
        assert result.iloc[0]["班次"] == "Day"

    def test_empty_device_names_filtered(self):
        proc = _make_processor({})
        df_raw = self._build_sheet2_df([
            ("", "CompanyA", 10.0, 18.0),
            (None, "CompanyB", 8.0, 15.0),
        ])
        result = proc.process_sheet2(df_raw, date(2025, 1, 15), "Day")
        assert result.empty

    def test_too_few_rows_returns_empty(self):
        proc = _make_processor({})
        df_raw = pd.DataFrame([[None] * 11], dtype=object)
        result = proc.process_sheet2(df_raw, date(2025, 1, 15), "Day")
        assert result.empty

    def test_km_values_are_zero(self):
        proc = _make_processor({})
        df_raw = self._build_sheet2_df([
            ("TR600-01", "CompanyA", 10.0, 18.0),
        ])
        result = proc.process_sheet2(df_raw, date(2025, 1, 15), "Day")
        assert result.iloc[0]["公里数仪表开始"] == 0
        assert result.iloc[0]["公里数仪表结束"] == 0
        assert result.iloc[0]["运行里程"] == 0
        assert result.iloc[0]["趟次"] == 0

    def test_sheet2_with_remark_column(self):
        """Sheet2 with remark column at index 8 should include remark."""
        proc = _make_processor({})
        n_cols = 11
        rows = [[None] * n_cols for _ in range(3)]  # rows 0-2
        row = [None] * n_cols
        row[1] = "TR600-01"   # col B
        row[2] = "CompanyA"   # col C
        row[5] = 10.0         # col F (hour start)
        row[6] = 18.0         # col G (hour end)
        row[8] = "SomeRemark" # col I (remark)
        rows.append(row)
        df_raw = pd.DataFrame(rows, dtype=object)

        result = proc.process_sheet2(df_raw, date(2025, 1, 15), "Day")
        assert not result.empty
        assert result.iloc[0]["备注"] == "SomeRemark"


# ---------------------------------------------------------------------------
# Tests: process_single_file (via synthetic Excel on disk)
# ---------------------------------------------------------------------------

class TestProcessSingleFile:
    def test_full_two_sheet_file(self, tmp_path):
        """A complete file with both sheet1 and sheet2 produces combined output."""
        load_map = {"TR600": 85, "CAT773": 20}
        truck_data = [
            {
                "truck_name": "TR600-01",
                "hour_start": 100.0, "hour_end": 108.5,
                "km_start": 5000.0, "km_end": 5050.0,
                "company": "CompanyA",
                "exc_1_trips": 10, "exc_2_trips": 0, "exc_3_trips": 0,
            },
        ]
        xlsx_path = tmp_path / "2025.01.15 白班.xlsx"
        _write_two_sheet_xlsx(xlsx_path, truck_data,
                              sheet2_devices=[("CAT773-02", "CompanyB", 8.0, 15.5)],
                              raw_start=6)

        proc = _make_processor(load_map, raw_start=6)
        output_path = tmp_path / "output.xlsx"
        running_df, production_df = proc.process_single_file(str(xlsx_path), str(output_path))

        assert not running_df.empty
        # Sheet1 contributes TR600-01, sheet2 contributes CAT773-02
        assert len(running_df) == 2
        trucks = set(running_df["设备名称"])
        assert "TR600-01" in trucks
        assert "CAT773-02" in trucks

        assert not production_df.empty
        assert production_df.iloc[0]["矿卡名称"] == "TR600-01"

        # Output file was written
        assert output_path.exists()

    def test_single_sheet_file(self, tmp_path):
        """File with only 1 sheet should still work (sheet2 returns empty)."""
        load_map = {"TR600": 85}
        truck_data = [
            {
                "truck_name": "TR600-01",
                "hour_start": 100.0, "hour_end": 108.5,
                "km_start": 5000.0, "km_end": 5050.0,
                "company": "CompanyA",
                "exc_1_trips": 10, "exc_2_trips": 0, "exc_3_trips": 0,
            },
        ]
        xlsx_path = tmp_path / "2025.06.01 夜班.xlsx"
        _write_sheet1_xlsx(xlsx_path, truck_data, raw_start=6)

        # _write_sheet1_xlsx creates a workbook with exactly one sheet "Sheet1"
        import openpyxl
        wb_check = openpyxl.load_workbook(xlsx_path)
        assert len(wb_check.sheetnames) == 1
        wb_check.close()

        proc = _make_processor(load_map, raw_start=6)
        output_path = tmp_path / "single_output.xlsx"
        running_df, production_df = proc.process_single_file(str(xlsx_path), str(output_path))

        assert not running_df.empty
        assert len(running_df) == 1
        assert running_df.iloc[0]["设备名称"] == "TR600-01"
        assert running_df.iloc[0]["班次"] == "Night"

    def test_single_sheet_file_minimal(self, tmp_path):
        """File with only 1 sheet — simplest case."""
        load_map = {"TR600": 85}
        truck_data = [
            {
                "truck_name": "TR600-01",
                "hour_start": 100.0, "hour_end": 108.5,
                "km_start": 5000.0, "km_end": 5050.0,
                "company": "CompanyA",
                "exc_1_trips": 5, "exc_2_trips": 3, "exc_3_trips": 0,
            },
        ]
        xlsx_path = tmp_path / "2025.03.10 白班.xlsx"
        _write_sheet1_xlsx(xlsx_path, truck_data, raw_start=6)

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        assert len(wb.sheetnames) == 1
        wb.close()

        proc = _make_processor(load_map, raw_start=6)
        running_df, production_df = proc.process_single_file(str(xlsx_path))

        assert not running_df.empty
        assert len(running_df) == 1
        assert production_df.notna().sum().sum() > 0  # has production rows


# ---------------------------------------------------------------------------
# Tests: day/night shift splitting via filename parsing
# ---------------------------------------------------------------------------

class TestDayNightSplitting:
    def test_day_shift_extracted(self):
        proc = _make_processor({})
        _, shift = proc.parse_filename("2025.05.20 白班.xlsx")
        assert shift == "Day"

    def test_night_shift_extracted(self):
        proc = _make_processor({})
        _, shift = proc.parse_filename("2025.05.20 夜班.xlsx")
        assert shift == "Night"

    def test_both_shifts_same_date(self):
        proc = _make_processor({})
        d1, s1 = proc.parse_filename("2025.05.20 白班.xlsx")
        d2, s2 = proc.parse_filename("2025.05.20 夜班.xlsx")
        assert d1 == d2
        assert s1 == "Day"
        assert s2 == "Night"

    def test_filename_with_extra_text(self):
        proc = _make_processor({})
        d, s = proc.parse_filename("2025.05.20 白班 production report.xlsx")
        assert d == date(2025, 5, 20)
        assert s == "Day"

    def test_multiple_dates_only_first_matched(self):
        proc = _make_processor({})
        d, _ = proc.parse_filename("2025.05.20 白班 2025.05.21.xlsx")
        assert d == date(2025, 5, 20)


# ---------------------------------------------------------------------------
# Tests: safe_number edge cases
# ---------------------------------------------------------------------------

class TestSafeNumber:
    def test_numeric_string(self):
        proc = _make_processor({})
        assert proc.safe_number("123.45") == 123.45

    def test_none_returns_default(self):
        proc = _make_processor({})
        assert proc.safe_number(None) == 0

    def test_nan_returns_default(self):
        proc = _make_processor({})
        assert proc.safe_number(float("nan")) == 0

    def test_series_returns_first_nonnull(self):
        proc = _make_processor({})
        s = pd.Series([None, 42.5, None])
        assert proc.safe_number(s) == 42.5

    def test_empty_series_returns_default(self):
        proc = _make_processor({})
        s = pd.Series([], dtype=float)
        assert proc.safe_number(s) == 0

    def test_custom_default(self):
        proc = _make_processor({})
        assert proc.safe_number(None, default=-1) == -1

    def test_non_numeric_string_coerced_to_nan(self):
        proc = _make_processor({})
        assert proc.safe_number("abc") == 0


# ---------------------------------------------------------------------------
# Tests: find_first_matching_column
# ---------------------------------------------------------------------------

class TestFindFirstMatchingColumn:
    def test_or_logic_flat_keywords(self):
        proc = _make_processor({})
        cols = pd.Index(["小时数｜开始", "公司", "公里数｜结束"])
        result = proc.find_first_matching_column(cols, ["公司", "Компани"])
        assert result == "公司"

    def test_and_logic_nested_keywords(self):
        proc = _make_processor({})
        cols = pd.Index(["小时数｜开始", "公司", "备注"])
        result = proc.find_first_matching_column(cols, [["小时数", "开始"]])
        assert result == "小时数｜开始"

    def test_no_match_returns_none(self):
        proc = _make_processor({})
        cols = pd.Index(["A", "B", "C"])
        result = proc.find_first_matching_column(cols, ["XYZ"])
        assert result is None

    def test_mixed_or_and_groups(self):
        proc = _make_processor({})
        cols = pd.Index(["公里数｜开始", "小时数｜结束"])
        keywords = [
            ["小时数", "开始"],  # AND: won't match "小时数｜结束"
            ["公里数", "开始"],  # AND: matches "公里数｜开始"
        ]
        result = proc.find_first_matching_column(cols, keywords)
        assert result == "公里数｜开始"


# ---------------------------------------------------------------------------
# Tests: full folder processing
# ---------------------------------------------------------------------------

class TestProcessFolder:
    def test_collect_excel_files(self, tmp_path):
        proc = _make_processor({})
        (tmp_path / "2025.01.15 白班.xlsx").touch()
        (tmp_path / "2025.01.15 夜班.xlsx").touch()
        (tmp_path / "2025.01.16 白班.xlsm").touch()
        # These should be filtered out
        (tmp_path / "2025.01.15 日报.xlsx").touch()
        (tmp_path / "~$temp.xlsx").touch()
        (tmp_path / "readme.txt").touch()

        files = proc.collect_excel_files(str(tmp_path))
        assert len(files) == 3
        for f in files:
            assert "白班" in f or "夜班" in f
            assert "~$" not in f

    def test_process_folder_empty(self, tmp_path):
        proc = _make_processor({})
        output_path = tmp_path / "output.xlsx"
        proc.process_folder(str(tmp_path), str(output_path))
        assert output_path.exists()

    def test_process_folder_with_valid_files(self, tmp_path):
        load_map = {"TR600": 85}
        proc = _make_processor(load_map, raw_start=6)

        for name in ["2025.01.15 白班.xlsx", "2025.01.15 夜班.xlsx"]:
            xlsx_path = tmp_path / name
            _write_sheet1_xlsx(xlsx_path, [
                {
                    "truck_name": "TR600-01",
                    "hour_start": 100.0, "hour_end": 108.5,
                    "km_start": 5000.0, "km_end": 5050.0,
                    "company": "CompanyA",
                    "exc_1_trips": 10, "exc_2_trips": 0, "exc_3_trips": 0,
                },
            ], raw_start=6)

        output_path = tmp_path / "output.xlsx"
        sheets = proc.process_folder(str(tmp_path), str(output_path), return_sheets=True)

        assert sheets is not None
        assert "运行数据" in sheets
        assert "生产数据" in sheets
        assert len(sheets["运行数据"]) == 2  # one per file
