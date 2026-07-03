"""Tests for hidden row/column detection and filtering utilities."""

import pathlib
import sys

import pandas as pd
import pytest
from openpyxl import Workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from func.excel_utils import (
    filter_hidden_from_df,
    get_column_letter,
    get_hidden_indices,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_test_workbook(
    path: pathlib.Path,
    hidden_rows: list[int] | None = None,
    hidden_cols: list[str] | None = None,
) -> str:
    """Create a simple 5×4 Excel file and optionally hide rows/columns."""
    wb = Workbook()
    ws = wb.active
    for r in range(1, 6):
        for c in range(1, 5):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")
    for r in (hidden_rows or []):
        ws.row_dimensions[r].hidden = True
    for col_letter in (hidden_cols or []):
        ws.column_dimensions[col_letter].hidden = True
    fpath = str(path / "test.xlsx")
    wb.save(fpath)
    return fpath


# ---------------------------------------------------------------------------
# get_column_letter
# ---------------------------------------------------------------------------

class TestGetColumnLetter:
    def test_single_letter(self):
        assert get_column_letter(1) == "A"
        assert get_column_letter(26) == "Z"

    def test_double_letter(self):
        assert get_column_letter(27) == "AA"


# ---------------------------------------------------------------------------
# get_hidden_indices
# ---------------------------------------------------------------------------

class TestGetHiddenIndices:
    def test_no_hidden(self, tmp_path):
        fpath = _create_test_workbook(tmp_path)
        rows, cols = get_hidden_indices(fpath, 0)
        assert rows == set()
        assert cols == set()

    def test_hidden_rows(self, tmp_path):
        fpath = _create_test_workbook(tmp_path, hidden_rows=[2, 4])
        rows, cols = get_hidden_indices(fpath, 0)
        assert rows == {2, 4}
        assert cols == set()

    def test_hidden_cols(self, tmp_path):
        fpath = _create_test_workbook(tmp_path, hidden_cols=["B", "D"])
        rows, cols = get_hidden_indices(fpath, 0)
        assert rows == set()
        assert cols == {"B", "D"}

    def test_hidden_rows_and_cols(self, tmp_path):
        fpath = _create_test_workbook(tmp_path, hidden_rows=[1], hidden_cols=["C"])
        rows, cols = get_hidden_indices(fpath, 0)
        assert rows == {1}
        assert cols == {"C"}

    def test_by_sheet_name(self, tmp_path):
        fpath = _create_test_workbook(tmp_path, hidden_rows=[3])
        rows, cols = get_hidden_indices(fpath, "Sheet")
        assert rows == {3}


# ---------------------------------------------------------------------------
# filter_hidden_from_df
# ---------------------------------------------------------------------------

class TestFilterHiddenFromDf:
    def _make_df(self):
        """5 rows × 4 columns, header=None."""
        return pd.DataFrame(
            [
                ["A1", "B1", "C1", "D1"],
                ["A2", "B2", "C2", "D2"],
                ["A3", "B3", "C3", "D3"],
                ["A4", "B4", "C4", "D4"],
                ["A5", "B5", "C5", "D5"],
            ]
        )

    def test_no_filter(self):
        df = self._make_df()
        result = filter_hidden_from_df(df, set(), set())
        assert len(result) == 5
        assert result.shape[1] == 4

    def test_filter_hidden_rows(self):
        df = self._make_df()
        # Excel row 2 → df index 1
        result = filter_hidden_from_df(df, {2}, set())
        assert len(result) == 4
        # Index is NOT reset — original indices preserved (0, 2, 3, 4)
        assert list(result.index) == [0, 2, 3, 4]
        # .iloc[] is positional and works correctly
        assert list(result.iloc[:, 0]) == ["A1", "A3", "A4", "A5"]

    def test_iloc_works_with_noncontiguous_index(self):
        """After filtering, .iloc[] still accesses by position, not index label."""
        df = self._make_df()
        result = filter_hidden_from_df(df, {2, 4}, set())
        assert len(result) == 3
        # Index: 0, 2, 4 (rows 1, 3, 5 — original positions preserved)
        assert list(result.index) == [0, 2, 4]
        # .iloc[1] should be the SECOND remaining row (A3), not index 1
        assert list(result.iloc[1]) == ["A3", "B3", "C3", "D3"]
        assert list(result.iloc[2]) == ["A5", "B5", "C5", "D5"]

    def test_index_preserves_original_row_numbers(self):
        """Content-based index lookups (like fuel processor) still work."""
        df = self._make_df()
        result = filter_hidden_from_df(df, {1}, set())
        # Row 1 was hidden → remaining: rows 2,3,4,5 (indices 1,2,3,4)
        # Search for 'A3' in column 0 → should find at original index 2
        matches = result[result.iloc[:, 0] == "A3"]
        assert matches.index[0] == 2

    def test_filter_hidden_cols(self):
        df = self._make_df()
        # Column B → second column
        result = filter_hidden_from_df(df, set(), {"B"})
        assert result.shape[1] == 3
        assert list(result.columns) == [0, 1, 2]
        # First row should be A1, C1, D1
        assert list(result.iloc[0]) == ["A1", "C1", "D1"]

    def test_filter_both(self):
        df = self._make_df()
        result = filter_hidden_from_df(df, {1, 5}, {"A", "D"})
        assert len(result) == 3
        assert result.shape[1] == 2
        # Index preserved: rows 2,3,4 (indices 1,2,3)
        assert list(result.index) == [1, 2, 3]
        # First remaining row (row 2) with cols B,C
        assert list(result.iloc[0]) == ["B2", "C2"]

    def test_filter_with_header(self):
        """When has_header=True, Excel row 2 maps to df index 0."""
        df = pd.DataFrame(
            [["B2", "C2", "D2"], ["B3", "C3", "D3"], ["B4", "C4", "D4"]],
            columns=["col_b", "col_c", "col_d"],
        )
        # Excel row 2 is the first data row (header was row 1, consumed by pandas)
        result = filter_hidden_from_df(df, {2}, set(), has_header=True)
        assert len(result) == 2

    def test_empty_filter_sets(self):
        df = self._make_df()
        result = filter_hidden_from_df(df, set(), set())
        assert result.equals(df)

    def test_out_of_range_rows_ignored(self):
        df = self._make_df()
        # Row 100 doesn't exist
        result = filter_hidden_from_df(df, {100}, set())
        assert len(result) == 5
