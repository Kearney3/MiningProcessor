"""台账匹配工具区域组件"""
import datetime
import asyncio
import logging
import threading
from pathlib import Path

import pandas as pd
import flet as ft



from .common import _log_message, _last_directory as _import_dir, _update_last_directory, _cell_text

try:
    from . import theme
except ImportError:
    import gui.theme as theme


def _strip_date_only_times(df: pd.DataFrame) -> pd.DataFrame:
    """对 datetime 列检测：若所有非空值的时间部分均为 00:00:00，
    则转换为 date 对象，避免 Excel 导出时出现多余的 ' 00:00:00'。"""
    for col in df.columns:
        if not pd.api.types.is_datetime64_any_dtype(df[col]):
            continue
        times = df[col].dropna().dt.time
        if times.empty:
            continue
        import datetime as _dt
        midnight = _dt.time(0, 0, 0)
        if (times == midnight).all():
            df[col] = df[col].apply(
                lambda v: v.date() if pd.notna(v) else v
            )
    return df


def create_ledger_match_section(
    page: ft.Page, log, ledger_refs: dict, oil_ledger_refs: dict
) -> tuple[ft.Container, dict]:
    """创建台账匹配工具区域，返回 (container, refs)"""

    PAGE_SIZE = 20
    _all_sheets: dict[str, pd.DataFrame] = {}  # sheet_name -> DataFrame (原始数据)
    _matched_all_sheets: dict[str, pd.DataFrame] = {}  # sheet_name -> 匹配后的完整数据
    _filtered_df: list[pd.DataFrame] = [None]  # 当前 sheet 经过排序/筛选后的视图
    _current_sheet = [""]
    _page = [0]
    _columns: list[str] = []
    _sort_column: list[str] = [None]  # 当前排序列
    _sort_ascending: list[bool] = [True]  # 排序方向
    _matched_sheets: dict[str, pd.DataFrame] = {}   # sheet_name -> 匹配成功的行
    _unmatched_sheets: dict[str, pd.DataFrame] = {}  # sheet_name -> 匹配失败的行
    _view_mode: list[str] = ["all"]  # "all" | "matched" | "unmatched"

    # --- 控件 ---
    file_label = ft.Text("未导入文件", size=12, color=ft.Colors.GREY)

    sheet_dropdown = ft.Dropdown(
        label="Sheet",
        width=200,
        dense=True,
        options=[],
    )

    name_dropdown = ft.Dropdown(
        label="设备名称列",
        hint_text="（可选）",
        width=180,
        dense=True,
        options=[],
        disabled=True,
    )
    id_dropdown = ft.Dropdown(
        label="设备编号列",
        hint_text="（可选）",
        width=180,
        dense=True,
        options=[],
        disabled=True,
    )
    oil_dropdown = ft.Dropdown(
        label="油品列",
        hint_text="（可选）",
        width=180,
        dense=True,
        options=[],
        disabled=True,
    )

    name_match_switch = ft.Switch(
        label="名称匹配", value=False,
    )
    id_match_switch = ft.Switch(
        label="编号匹配", value=False,
    )
    oil_match_switch = ft.Switch(
        label="油品匹配", value=False,
    )

    match_btn = theme.primary_btn("执行匹配", icon=ft.Icons.SEARCH, disabled=True)
    export_btn = theme.secondary_btn("导出结果", icon=ft.Icons.DOWNLOAD, disabled=True)

    _VIEW_LABELS = ["全部", "已匹配", "未匹配"]
    _VIEW_MODES = ["all", "matched", "unmatched"]

    def _on_view_segment_change(e):
        sel = e.control.selected  # set of selected values
        if not sel:
            return
        val = next(iter(sel))
        idx = _VIEW_MODES.index(val) if val in _VIEW_MODES else 0
        _on_view_change(idx)

    from flet.controls.material.segmented_button import Segment

    view_segment = ft.SegmentedButton(
        selected=["all"],
        allow_empty_selection=False,
        segments=[
            Segment(label=ft.Text("全部"), value="all"),
            Segment(label=ft.Text("已匹配"), value="matched"),
            Segment(label=ft.Text("未匹配"), value="unmatched"),
        ],
        on_change=_on_view_segment_change,
        disabled=True,  # 初始禁用，导入后启用
    )

    status_label = ft.Text("", size=12, color=theme.TEXT_SECONDARY)
    match_count_label = ft.Text("", size=12, color=theme.TEXT_SECONDARY)

    _import_progress_bar = ft.ProgressBar(
        value=0, height=6, visible=False, expand=True,
    )
    _import_progress_text = ft.Text("", size=12, color=theme.TEXT_SECONDARY, visible=False)
    _cancel_btn = ft.Button(
        "取消导入",
        icon=ft.Icons.CANCEL,
        visible=False,
        style=ft.ButtonStyle(bgcolor=theme.ERROR, color="#FFFFFF"),
        height=36,
    )
    _import_cancelled = threading.Event()

    # --- 表格 ---
    data_table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("等待导入数据..."))],
        rows=[],
        expand=True,
        sort_column_index=None,
        sort_ascending=True,
    )

    page_label = ft.Text("0 / 0", size=12, color=theme.TEXT_SECONDARY)
    prev_btn = ft.IconButton(
        icon=ft.Icons.CHEVRON_LEFT, tooltip="上一页", icon_size=18, disabled=True,
    )
    next_btn = ft.IconButton(
        icon=ft.Icons.CHEVRON_RIGHT, tooltip="下一页", icon_size=18, disabled=True,
    )

    # ========================================================================
    # 内部工具函数
    # ========================================================================
    def _get_current_df() -> pd.DataFrame | None:
        name = _current_sheet[0]
        if not name:
            return None
        # 优先返回匹配后的数据，如果没有则返回原始数据
        if name in _matched_all_sheets:
            return _matched_all_sheets[name]
        if name in _all_sheets:
            return _all_sheets[name]
        return None

    def _apply_filter_and_sort():
        """对当前 sheet 的 DataFrame 应用排序，更新 _filtered_df"""
        df = _get_current_df()
        if df is None:
            _filtered_df[0] = None
            return

        result = df.copy()

        # 排序
        col = _sort_column[0]
        if col and col in result.columns:
            ascending = _sort_ascending[0]
            try:
                result = result.sort_values(by=col, ascending=ascending, kind="stable")
            except Exception:
                logging.getLogger(__name__).debug("排序失败: col=%s", col)

        _filtered_df[0] = result

    def _total_pages():
        df = _get_view_df()
        if df is None or df.empty:
            return 1
        return max(1, (len(df) + PAGE_SIZE - 1) // PAGE_SIZE)

    def _update_page_controls():
        total = _total_pages()
        cur = _page[0]
        page_label.value = f"{cur + 1} / {total}"
        prev_btn.disabled = cur <= 0
        next_btn.disabled = cur >= total - 1

    def _show_import_progress(total: int):
        """显示导入进度 UI"""
        _import_cancelled.clear()
        _import_progress_bar.value = 0
        _import_progress_bar.visible = True
        _import_progress_text.value = f"正在解析 0/{total} 个 sheet..."
        _import_progress_text.visible = True
        _cancel_btn.visible = True
        match_btn.disabled = True

    def _update_import_progress(current: int, total: int, sheet_name: str):
        """更新导入进度"""
        _import_progress_bar.value = current / total if total > 0 else 0
        _import_progress_text.value = f"正在解析 {current}/{total}: {sheet_name}"

    def _hide_import_progress():
        """隐藏导入进度 UI"""
        _import_progress_bar.visible = False
        _import_progress_text.visible = False
        _cancel_btn.visible = False
        _import_progress_bar.value = 0

    def _on_cancel_import(e):
        _import_cancelled.set()
        _log_message(log, "正在取消导入...")

    _cancel_btn.on_click = _on_cancel_import

    def _on_name_toggle(e):
        name_dropdown.disabled = not name_match_switch.value
        name_dropdown.update()

    def _on_id_toggle(e):
        id_dropdown.disabled = not id_match_switch.value
        id_dropdown.update()

    def _on_oil_toggle(e):
        enabled = oil_match_switch.value
        oil_dropdown.disabled = not enabled
        oil_dropdown.update()

    def _on_view_change(tab_index: int):
        modes = ["all", "matched", "unmatched"]
        _view_mode[0] = modes[tab_index]
        _page[0] = 0
        build_table()

    name_match_switch.on_change = _on_name_toggle
    id_match_switch.on_change = _on_id_toggle
    oil_match_switch.on_change = _on_oil_toggle

    def _rebuild_columns(cols: list[str]):
        nonlocal _columns
        _columns = cols

        def on_sort_handler(col_idx):
            def handler(e):
                _sort_column[0] = cols[e.column_index]
                _sort_ascending[0] = e.ascending
                _apply_filter_and_sort()
                _page[0] = 0
                build_table()
            return handler

        if cols:
            data_table.columns = [
                ft.DataColumn(
                    ft.Text(c, size=13, no_wrap=True),
                    on_sort=on_sort_handler(c),
                )
                for c in cols
            ]
            if _sort_column[0] and _sort_column[0] in cols:
                data_table.sort_column_index = cols.index(_sort_column[0])
                data_table.sort_ascending = _sort_ascending[0]
            else:
                data_table.sort_column_index = None
        else:
            data_table.columns = [ft.DataColumn(ft.Text("等待导入数据..."))]

    def _get_view_df() -> pd.DataFrame | None:
        """根据当前视图模式返回对应的 DataFrame"""
        mode = _view_mode[0]
        sheet = _current_sheet[0]
        if mode == "matched":
            return _matched_sheets.get(sheet)
        elif mode == "unmatched":
            return _unmatched_sheets.get(sheet)
        return _filtered_df[0]

    _empty_state = ft.Column(
        [
            ft.Icon(ft.Icons.TABLE_CHART_OUTLINED, size=48, color=ft.Colors.GREY_300),
            ft.Text("暂无数据", size=14, color=theme.TEXT_SECONDARY, weight=ft.FontWeight.W_500),
            ft.Text("点击上方「导入文件」开始", size=12, color=ft.Colors.GREY_400),
        ],
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        alignment=ft.MainAxisAlignment.CENTER,
        spacing=8,
    )

    def build_table():
        _apply_filter_and_sort()
        df = _get_view_df()
        if df is None or df.empty:
            data_table.rows = []
            data_table.columns = [ft.DataColumn(ft.Text("等待导入数据..."))]
            _empty_state.visible = True
            _update_page_controls()
            page.update()
            return
        _empty_state.visible = False

        cols = list(df.columns)
        _rebuild_columns(cols)

        start = _page[0] * PAGE_SIZE
        end = start + PAGE_SIZE
        page_df = df.iloc[start:end]

        rows = []
        for row_idx, row in page_df.iterrows():
            cells = []
            for c in cols:
                cell_value = _cell_text(row[c])
                cells.append(ft.DataCell(ft.Text(cell_value, size=13, selectable=True)))
            rows.append(ft.DataRow(cells=cells))

        data_table.rows = rows
        _update_page_controls()
        page.update()

    def _prev(e):
        if _page[0] > 0:
            _page[0] -= 1
            build_table()

    def _next(e):
        if _page[0] < _total_pages() - 1:
            _page[0] += 1
            build_table()

    prev_btn.on_click = _prev
    next_btn.on_click = _next

    # ========================================================================
    # Sheet 切换 & 列名自动匹配
    # ========================================================================
    def _update_match_status(sheet_name: str = None):
        """更新匹配计数显示"""
        sheet = sheet_name or _current_sheet[0]
        matched = _matched_sheets.get(sheet)
        unmatched = _unmatched_sheets.get(sheet)
        if matched is not None and unmatched is not None:
            m = len(matched)
            u = len(unmatched)
            match_count_label.value = f"已匹配: {m}  |  未匹配: {u}"
        else:
            match_count_label.value = ""

    def _update_column_dropdowns(cols: list[str]):
        options = [ft.dropdown.Option(c) for c in cols]
        for dd in [name_dropdown, id_dropdown, oil_dropdown]:
            old_val = dd.value
            dd.options = options
            if old_val and old_val in cols:
                dd.value = old_val
            else:
                dd.value = None
        if not name_dropdown.value:
            for c in cols:
                if c.strip() in ("设备名称", "矿卡名称"):
                    name_dropdown.value = c
                    break
        if not id_dropdown.value:
            for c in cols:
                if c.strip() == "设备编号":
                    id_dropdown.value = c
                    break
        if not oil_dropdown.value:
            for c in cols:
                if c.strip() in ("油品种类", "油品名称"):
                    oil_dropdown.value = c
                    break

    def _on_sheet_change(sheet_name: str):
        logging.getLogger(__name__).debug(
            "_on_sheet_change called: sheet_name=%r, _all_sheets.keys()=%s",
            sheet_name, list(_all_sheets.keys())[:5],
        )
        if not sheet_name or sheet_name not in _all_sheets:
            return
        _current_sheet[0] = sheet_name
        _page[0] = 0
        _sort_column[0] = None
        _columns.clear()
        _sort_ascending[0] = True
        # 重置视图模式为"全部"，避免切换 sheet 后显示不一致
        _view_mode[0] = "all"
        view_segment.selected = ["all"]
        df = _all_sheets[sheet_name]
        _update_column_dropdowns(list(df.columns))
        _update_match_status(sheet_name)
        build_table()

    def _on_sheet_dropdown_change(e):
        logging.getLogger(__name__).debug(
            "sheet_dropdown.on_select fired: value=%r", e.control.value
        )
        _on_sheet_change(e.control.value)

    sheet_dropdown.on_select = _on_sheet_dropdown_change

    # ========================================================================
    # 文件导入
    # ========================================================================
    async def on_import(e):
        picker = ft.FilePicker()
        files = await picker.pick_files(
            dialog_title="导入 Excel 文件",
            allowed_extensions=["xlsx", "xls"],
            initial_directory=_import_dir[0] or None,
        )
        if not files:
            return
        path = files[0].path

        # 显示进度条
        _import_progress_bar.visible = True
        _import_progress_text.visible = True
        _cancel_btn.visible = True
        _import_cancelled.clear()
        match_btn.disabled = True
        page.update()

        parsed_sheets: dict[str, pd.DataFrame] = {}

        try:
            from openpyxl import load_workbook

            # 使用 read_only 模式打开，可以获取行数
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            total_sheets = len(sheet_names)

            if total_sheets == 0:
                _log_message(log, "文件中没有 sheet", level=logging.WARNING)
                _hide_import_progress()
                page.update()
                return

            for sheet_idx, sname in enumerate(sheet_names):
                if _import_cancelled.is_set():
                    break

                ws = wb[sname]
                # 获取行数（read_only 模式下 max_row 可用）
                total_rows = ws.max_row or 0
                total_cols = ws.max_column or 0

                if total_rows == 0 or total_cols == 0:
                    parsed_sheets[sname] = pd.DataFrame()
                    continue

                # 读取表头（第一行）
                headers = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    val = cell.value
                    if val is not None:
                        # strip 空白字符，避免列名匹配失败
                        headers.append(str(val).strip())
                    else:
                        headers.append(f"Col{cell.column - 1}")

                # 分批读取数据行
                rows_data = []
                batch_size = 500
                rows_read = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if _import_cancelled.is_set():
                        break

                    rows_data.append(list(row))
                    rows_read += 1

                    # 每 batch_size 行更新一次进度
                    if rows_read % batch_size == 0:
                        progress = rows_read / total_rows if total_rows > 0 else 0
                        _import_progress_bar.value = progress
                        _import_progress_text.value = f"正在导入 {sname}: {rows_read}/{total_rows} 行"
                        page.update()
                        await asyncio.sleep(0)

                # 构建 DataFrame
                if rows_data:
                    # 检查数据行长度与表头是否一致
                    header_len = len(headers)
                    row_lens = set(len(r) for r in rows_data)
                    if len(row_lens) > 1 or (row_lens and header_len not in row_lens):
                        logging.getLogger(__name__).debug(
                            "Sheet %r: header_len=%d, row_lengths=%s",
                            sname, header_len, row_lens,
                        )
                    df = pd.DataFrame(rows_data, columns=headers)
                    df = _strip_date_only_times(df)
                else:
                    df = pd.DataFrame(columns=headers)

                parsed_sheets[sname] = df
                _log_message(log, f"已导入 {sname}: {len(df)} 行, {len(df.columns)} 列")

            wb.close()

        except Exception as ex:
            file_label.value = f"读取失败: {ex}"
            file_label.color = theme.ERROR
            _log_message(log, f"读取文件失败: {ex}", level=logging.ERROR)
            _hide_import_progress()
            page.update()
            return

        _hide_import_progress()

        if _import_cancelled.is_set():
            _log_message(log, f"导入已取消（已解析 {len(parsed_sheets)}/{total_sheets} 个 sheet）")
            if not parsed_sheets:
                page.update()
                return

        # 使用已解析的数据
        _all_sheets.clear()
        _all_sheets.update(parsed_sheets)
        logging.getLogger(__name__).debug(
            "on_import: _all_sheets keys=%s", list(_all_sheets.keys())
        )

        _update_last_directory(path)
        file_label.value = Path(path).name
        file_label.color = ft.Colors.GREEN

        sheet_dropdown.options = [ft.dropdown.Option(s) for s in sheet_names]
        if sheet_names:
            first = sheet_names[0]
            sheet_dropdown.value = first
            # 直接调用 _on_sheet_change 确保初始化完整
            _on_sheet_change(first)
        else:
            _current_sheet[0] = ""

        match_btn.disabled = False
        view_segment.disabled = False
        build_table()
        loaded = len(parsed_sheets)
        _log_message(log, f"已导入: {path} ({loaded}/{total_sheets} 个 sheet)")
        page.update()
    # ========================================================================

    def _do_clear_impl():
        """清空的实际逻辑"""
        _hide_import_progress()
        _matched_sheets.clear()
        _unmatched_sheets.clear()
        _matched_all_sheets.clear()
        _view_mode[0] = "all"
        view_segment.selected = ["all"]
        view_segment.disabled = True
        _all_sheets.clear()
        _filtered_df[0] = None
        _current_sheet[0] = ""
        _page[0] = 0
        _sort_column[0] = None
        _columns.clear()
        _sort_ascending[0] = True
        file_label.value = "未导入文件"
        file_label.color = ft.Colors.GREY
        sheet_dropdown.options = []
        sheet_dropdown.value = None
        name_dropdown.options = []
        name_dropdown.value = None
        id_dropdown.options = []
        id_dropdown.value = None
        oil_dropdown.options = []
        oil_dropdown.value = None
        match_btn.disabled = True
        export_btn.disabled = True
        status_label.value = ""
        data_table.columns = [ft.DataColumn(ft.Text("等待导入数据..."))]
        data_table.rows = []
        _log_message(log, "已清空")
        page.update()

    def _do_clear_confirmed(e):
        page.pop_dialog()
        _do_clear_impl()

    _clear_confirm_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("确认清空"),
        content=ft.Text("确定要清空所有已导入数据和匹配结果吗？此操作不可撤销。"),
        actions=[
            ft.TextButton("取消", on_click=lambda e: page.pop_dialog()),
            ft.TextButton("确认清空", on_click=_do_clear_confirmed,
                          style=ft.ButtonStyle(color=theme.ERROR)),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    def on_clear(e):
        if not _all_sheets:
            return
        page.show_dialog(_clear_confirm_dialog)
    # 执行匹配
    # ========================================================================
    async def on_match(e):
        df = _get_current_df()
        if df is None or df.empty:
            _log_message(log, "没有数据可匹配", level=logging.WARNING)
            return

        eq_ledger = ledger_refs.get("get_ledger", lambda: None)()
        oil_ledger = oil_ledger_refs.get("get_oil", lambda: None)()

        if not eq_ledger and not oil_ledger:
            _log_message(log, "请先在设备台账或油品台账页导入台账", level=logging.WARNING)
            return

        name_col = name_dropdown.value if name_match_switch.value else None
        id_col = id_dropdown.value if id_match_switch.value else None
        oil_col = oil_dropdown.value if oil_match_switch.value else None

        if not name_col and not id_col and not oil_col:
            _log_message(log, "未启用任何匹配，跳过匹配")
            return

        # 初始化进度条
        _import_progress_bar.visible = True
        _import_progress_text.visible = True
        _cancel_btn.visible = True
        _import_cancelled.clear()
        
        # Loading 状态
        match_btn.disabled = True
        match_btn.text = "匹配中..."
        match_btn.icon = ft.Icons.HOURGLASS_TOP
        export_btn.disabled = True
        page.update()

        batch_size = 100
        total_rows = len(df)
        processed = 0

        try:
            result_df = df.copy()
            matched_count = 0

            # 设备匹配
            if eq_ledger and (name_col or id_col):
                # 检测是否同时存在矿卡名称和挖机名称（生产数据场景）
                has_truck_col = "矿卡名称" in result_df.columns
                has_excavator_col = "挖机名称" in result_df.columns
                
                if has_truck_col and has_excavator_col:
                    # 生产数据场景：分别匹配矿卡和挖机，添加后缀
                    # 计算总工作量（矿卡 + 挖机）
                    total_work = total_rows * 2
                    processed = 0
                    
                    # 匹配矿卡名称
                    std_names, std_ids, std_companies = [], [], []
                    for i in range(0, total_rows, batch_size):
                        if _import_cancelled.is_set():
                            _log_message(log, "匹配已取消")
                            break
                        
                        batch = df.iloc[i:i+batch_size]
                        for _, row in batch.iterrows():
                            n = str(row["矿卡名称"]) if "矿卡名称" in result_df.columns and not pd.isna(row.get("矿卡名称")) else None
                            i_val = str(row[id_col]) if id_col and id_col in result_df.columns and not pd.isna(row.get(id_col)) else None
                            r = eq_ledger.match_device(name=n, device_id=i_val)
                            if r:
                                std_names.append(r.get("标准设备名称", ""))
                                std_ids.append(r.get("标准设备编号", ""))
                                std_companies.append(r.get("标准公司名称", ""))
                                matched_count += 1
                            else:
                                std_names.append("")
                                std_ids.append("")
                                std_companies.append("")
                        
                        # 更新进度
                        processed += len(batch)
                        progress = processed / total_work
                        _import_progress_bar.value = progress
                        _import_progress_text.value = f"正在匹配矿卡: {processed}/{total_work}"
                        page.update()
                        
                        # 让出控制权，避免 UI 冻结
                        await asyncio.sleep(0)
                    
                    result_df["标准设备名称（矿卡）"] = std_names
                    result_df["标准设备编号（矿卡）"] = std_ids
                    result_df["标准公司名称（矿卡）"] = std_companies
                    
                    # 匹配挖机名称
                    std_names_ex, std_ids_ex, std_companies_ex = [], [], []
                    for i in range(0, total_rows, batch_size):
                        if _import_cancelled.is_set():
                            _log_message(log, "匹配已取消")
                            break
                        
                        batch = df.iloc[i:i+batch_size]
                        for _, row in batch.iterrows():
                            n = str(row["挖机名称"]) if "挖机名称" in result_df.columns and not pd.isna(row.get("挖机名称")) else None
                            r = eq_ledger.match_device(name=n, device_id=None)
                            if r:
                                std_names_ex.append(r.get("标准设备名称", ""))
                                std_ids_ex.append(r.get("标准设备编号", ""))
                                std_companies_ex.append(r.get("标准公司名称", ""))
                            else:
                                std_names_ex.append("")
                                std_ids_ex.append("")
                                std_companies_ex.append("")
                        
                        # 更新进度
                        processed += len(batch)
                        progress = processed / total_work
                        _import_progress_bar.value = progress
                        _import_progress_text.value = f"正在匹配挖机: {processed}/{total_work}"
                        page.update()
                        
                        # 让出控制权，避免 UI 冻结
                        await asyncio.sleep(0)
                    
                    result_df["标准设备名称（挖机）"] = std_names_ex
                    result_df["标准设备编号（挖机）"] = std_ids_ex
                    result_df["标准公司名称（挖机）"] = std_companies_ex
                else:
                    # 原有逻辑：单列匹配（非生产数据场景）
                    std_names, std_ids, std_companies = [], [], []
                    for i in range(0, total_rows, batch_size):
                        if _import_cancelled.is_set():
                            _log_message(log, "匹配已取消")
                            break
                        
                        batch = df.iloc[i:i+batch_size]
                        for _, row in batch.iterrows():
                            n = str(row[name_col]) if name_col and name_col in result_df.columns and not pd.isna(row.get(name_col)) else None
                            i_val = str(row[id_col]) if id_col and id_col in result_df.columns and not pd.isna(row.get(id_col)) else None
                            r = eq_ledger.match_device(name=n, device_id=i_val)
                            if r:
                                std_names.append(r.get("标准设备名称", ""))
                                std_ids.append(r.get("标准设备编号", ""))
                                std_companies.append(r.get("标准公司名称", ""))
                                matched_count += 1
                            else:
                                std_names.append("")
                                std_ids.append("")
                                std_companies.append("")
                        
                        # 更新进度
                        processed += len(batch)
                        progress = processed / total_rows
                        _import_progress_bar.value = progress
                        _import_progress_text.value = f"正在匹配第 {processed}/{total_rows} 行..."
                        page.update()
                        
                        # 让出控制权，避免 UI 冻结
                        await asyncio.sleep(0)
                    
                    result_df["标准设备名称"] = std_names
                    result_df["标准设备编号"] = std_ids
                    result_df["标准公司名称"] = std_companies

            # 油品匹配
            oil_matched = 0
            if oil_ledger and oil_col and oil_col in result_df.columns:
                std_oils = []
                for i in range(0, total_rows, batch_size):
                    if _import_cancelled.is_set():
                        _log_message(log, "匹配已取消")
                        break
                    
                    batch = df.iloc[i:i+batch_size]
                    for _, row in batch.iterrows():
                        v = row[oil_col]
                        r = oil_ledger.match(str(v)) if not pd.isna(v) else None
                        if r:
                            std_oils.append(r["标准名称"])
                            oil_matched += 1
                        else:
                            std_oils.append("")
                    
                    # 更新进度
                    processed += len(batch)
                    progress = processed / total_rows
                    _import_progress_bar.value = progress
                    _import_progress_text.value = f"正在匹配油品: {processed}/{total_rows}"
                    page.update()
                    
                    # 让出控制权，避免 UI 冻结
                    await asyncio.sleep(0)
                
                result_df["标准油品名称"] = std_oils
            # 保存匹配结果到单独的字典，不覆盖原始数据
            _matched_all_sheets[_current_sheet[0]] = result_df
            _page[0] = 0
            build_table()

            # 记录日志
            logging.getLogger(__name__).debug(
                "on_match: updated _matched_all_sheets[%r], columns=%s",
                _current_sheet[0], list(result_df.columns),
            )

            # 拆分匹配成功/失败的行
            sheet = _current_sheet[0]
            if eq_ledger and (name_col or id_col):
                if has_truck_col and has_excavator_col:
                    # 生产数据：合并两个匹配列的匹配状态
                    mask_truck = result_df["标准设备名称（矿卡）"].astype(str).str.len() > 0
                    mask_ex = result_df["标准设备名称（挖机）"].astype(str).str.len() > 0
                    mask = mask_truck | mask_ex
                else:
                    mask = result_df["标准设备名称"].astype(str).str.len() > 0
                _matched_sheets[sheet] = result_df[mask].copy()
                _unmatched_sheets[sheet] = result_df[~mask].copy()
            elif oil_ledger and oil_col:
                mask = result_df["标准油品名称"].astype(str).str.len() > 0
                _matched_sheets[sheet] = result_df[mask].copy()
                _unmatched_sheets[sheet] = result_df[~mask].copy()

            parts = []
            if eq_ledger and (name_col or id_col):
                total = len(result_df)
                parts.append(f"设备匹配: {matched_count}/{total}")
            if oil_ledger and oil_col:
                total = len(result_df)
                parts.append(f"油品匹配: {oil_matched}/{total}")
            status_label.value = "  |  ".join(parts)
            _update_match_status()
            _log_message(log, f"匹配完成: {status_label.value}")
        except Exception as ex:
            _log_message(log, f"匹配失败: {ex}", level=logging.ERROR)
        finally:
            # 恢复 UI
            _import_progress_bar.visible = False
            _import_progress_text.visible = False
            _cancel_btn.visible = False
            match_btn.disabled = False
            match_btn.text = "执行匹配"
            match_btn.icon = ft.Icons.SEARCH
            export_btn.disabled = not bool(_all_sheets)
            page.update()
    # ========================================================================
    # 导出
    # ========================================================================
    async def on_export(e):
        if not _all_sheets and not _matched_sheets and not _unmatched_sheets:
            _log_message(log, "没有数据可导出", level=logging.WARNING)
            return

        # 选择保存路径
        picker = ft.FilePicker()
        path = await picker.save_file(
            dialog_title="导出结果",
            file_name="匹配结果.xlsx",
            allowed_extensions=["xlsx"],
            initial_directory=_import_dir[0] or None,
        )
        if not path:
            return

        # 初始化进度条
        _import_progress_bar.visible = True
        _import_progress_text.visible = True
        _cancel_btn.visible = True
        _import_cancelled.clear()
        export_btn.disabled = True
        page.update()

        try:
            import xlsxwriter

            # 根据视图模式选择数据
            mode = _view_mode[0]
            if mode == "matched" and _matched_sheets:
                df_to_export = pd.concat(_matched_sheets.values(), ignore_index=True)
                sheet_name = "已匹配"
            elif mode == "unmatched" and _unmatched_sheets:
                df_to_export = pd.concat(_unmatched_sheets.values(), ignore_index=True)
                sheet_name = "未匹配"
            else:
                # 优先使用匹配后的完整数据，如果没有则使用原始数据
                export_data = _matched_all_sheets if _matched_all_sheets else _all_sheets
                df_to_export = pd.concat(export_data.values(), ignore_index=True)
                sheet_name = "全部"

            # 创建 xlsxwriter 对象
            workbook = xlsxwriter.Workbook(path)
            worksheet = workbook.add_worksheet(sheet_name)

            # 定义日期格式
            date_fmt = workbook.add_format({'num_format': 'yyyy-mm-dd'})
            datetime_fmt = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})

            # 写入表头
            headers = list(df_to_export.columns)
            for col, header in enumerate(headers):
                worksheet.write(0, col, header)

            # 识别日期列（包括 datetime64 和 date/datetime 对象）
            date_columns = set()
            for col in headers:
                if pd.api.types.is_datetime64_any_dtype(df_to_export[col]):
                    date_columns.add(col)
                else:
                    # 检查是否有 date/datetime 对象
                    sample = df_to_export[col].dropna().head(10)
                    if not sample.empty and sample.apply(
                        lambda v: isinstance(v, (datetime.date, datetime.datetime))
                    ).any():
                        date_columns.add(col)

            # 分批写入数据
            batch_size = 1000
            total_rows = len(df_to_export)

            for i in range(0, total_rows, batch_size):
                if _import_cancelled.is_set():
                    _log_message(log, "导出已取消")
                    break

                batch = df_to_export.iloc[i:i+batch_size]
                for row_idx, (_, row) in enumerate(batch.iterrows(), start=i+1):
                    for col_idx, col_name in enumerate(headers):
                        value = row[col_name]
                        # 处理 NaN 值
                        if pd.isna(value):
                            worksheet.write(row_idx, col_idx, "")
                        elif col_name in date_columns:
                            # 日期列使用日期格式
                            try:
                                if isinstance(value, pd.Timestamp):
                                    worksheet.write_datetime(row_idx, col_idx, value.to_pydatetime(), date_fmt)
                                elif isinstance(value, (datetime.date, datetime.datetime)):
                                    worksheet.write_datetime(row_idx, col_idx, value, date_fmt)
                                else:
                                    worksheet.write(row_idx, col_idx, value)
                            except Exception:
                                worksheet.write(row_idx, col_idx, str(value))
                        else:
                            worksheet.write(row_idx, col_idx, value)

                # 更新进度
                processed = min(i + batch_size, total_rows)
                progress = processed / total_rows
                _import_progress_bar.value = progress
                _import_progress_text.value = f"正在导出第 {processed}/{total_rows} 行..."
                page.update()

                await asyncio.sleep(0)
            # 完成
            if not _import_cancelled.is_set():
                workbook.close()
                _log_message(log, f"已导出: {path}")
                _update_last_directory(path)

        except Exception as ex:
            _log_message(log, f"导出失败: {ex}", level=logging.ERROR)
        finally:
            # 恢复 UI
            _import_progress_bar.visible = False
            _import_progress_text.visible = False
            _cancel_btn.visible = False
            export_btn.disabled = False
            page.update()
    match_btn.on_click = on_match
    export_btn.on_click = on_export

    # ========================================================================
    # 布局
    # ========================================================================
    table_wrapper = ft.Column(
        controls=[
            ft.Row(
                controls=[data_table],
                scroll=ft.ScrollMode.AUTO,
            ),
            _empty_state,
        ],
        scroll=ft.ScrollMode.AUTO,
        expand=True,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        alignment=ft.MainAxisAlignment.CENTER,
    )

    # ── 文件操作栏 ──
    file_row = ft.Row(
        [
            ft.Row(
                [
                    theme.secondary_btn("导入文件", icon=ft.Icons.UPLOAD, on_click=on_import),
                    theme.destructive_btn("清空", icon=ft.Icons.DELETE_SWEEP, on_click=on_clear),
                    file_label,
                ],
                spacing=8,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            sheet_dropdown,
        ],
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
    )

    # ── 进度条 ──
    progress_row = ft.Row(
        [_import_progress_bar, _import_progress_text, _cancel_btn],
        spacing=8,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
    )

    # ── 匹配配置（可折叠，2 列网格） ──
    match_config_grid = ft.ResponsiveRow(
        [
            ft.Container(
                ft.Row([name_match_switch, name_dropdown], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                col={"xs": 12, "md": 6},
            ),
            ft.Container(
                ft.Row([id_match_switch, id_dropdown], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                col={"xs": 12, "md": 6},
            ),
            ft.Container(
                ft.Row([oil_match_switch, oil_dropdown], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                col={"xs": 12, "md": 6},
            ),
        ],
        run_spacing=4,
        spacing=8,
        vertical_alignment=ft.CrossAxisAlignment.CENTER,
    )

    match_config_collapsible = theme.make_collapsible(
        title="匹配配置",
        subtitle="选择设备名称/编号/油品列进行匹配",
        icon=ft.Icons.TUNE,
        initially_expanded=True,
        content_controls=[match_config_grid],
    )

    # ── 操作栏（2 行） ──
    action_rows = ft.Column(
        [
            ft.Row([match_btn, export_btn], spacing=8),
            ft.Row([status_label, match_count_label], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ],
        spacing=6,
    )

    container = ft.Container(
        content=ft.Column(
            [
                theme.section_title("台账匹配"),
                ft.Text(
                    "导入 Excel 文件后选择匹配列，执行匹配并导出结果。",
                    size=13,
                    color=theme.TEXT_SECONDARY,
                ),

                # ── 文件操作 ──
                theme.module_card([file_row, progress_row], spacing=6),

                # ── 匹配配置（折叠） ──
                match_config_collapsible,

                # ── 操作栏 ──
                action_rows,

                # ── 视图切换 ──
                view_segment,

                # ── 数据表格 ──
                ft.Container(
                    content=table_wrapper,
                    border=ft.Border.all(1, theme.BORDER),
                    border_radius=theme.RADIUS_MD,
                    padding=4,
                    bgcolor=theme.SURFACE_HIGH,
                    expand=True,
                ),

                # ── 分页 ──
                ft.Row(
                    [prev_btn, page_label, next_btn],
                    spacing=4,
                    alignment=ft.MainAxisAlignment.CENTER,
                ),
            ],
            spacing=theme.SPACING_MD,
            expand=True,
        ),
        padding=12,
        border=ft.Border.all(1, theme.BORDER),
        border_radius=theme.RADIUS_LG,
        bgcolor=theme.SURFACE,
        expand=True,
    )

    refs = {"build_table": build_table}
    return container, refs
